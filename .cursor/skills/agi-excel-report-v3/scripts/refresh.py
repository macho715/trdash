#!/usr/bin/env python3
# Refresh derived fields validation (lightweight)
import argparse
from openpyxl import load_workbook


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.xlsx)
    required = ["8_EVENT_MAP", "9_SEGMENT_DELTA_TR1-TR3", "10_EVIDENCE_LOG", "12_GATE_EVAL", "13_WINDOWS_ANALYSIS"]
    missing = [s for s in required if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Missing sheets: {missing}")

    for s in required:
        ws = wb[s]
        if ws.max_row < 1:
            raise SystemExit(f"Empty header: {s}")

    print("OK: structure validated")


if __name__ == "__main__":
    main()
