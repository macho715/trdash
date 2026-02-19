#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGI Delay Analysis Excel Template v3.0 (Sheets 0~18)
- Keeps 0~7 if base workbook provided
- Adds CORE 8~13 + ADV 14~18 with schema, DV, formats, and core formulas
- Timezone is documented (no tz conversion in Excel; expects LT string)
"""
import argparse
import datetime as dt
from pathlib import Path
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

TZ_DEFAULT = "Asia/Dubai"

SHEETS_EXISTING = [
    "0_README", "1_SSOT_INPUTS", "2_ACTUAL_MILESTONES", "3_ORIGINAL_PLAN",
    "4_DELAY_ANALYSIS_TR1-TR3", "5_FORECAST_TR4-TR7", "6_COMPARISON_TR1-TR7", "7_FLOAT_SUMMARY"
]

SHEETS_CORE = [
    "8_EVENT_MAP", "9_SEGMENT_DELTA_TR1-TR3", "10_EVIDENCE_LOG", "11_CAUSE_TAXONOMY",
    "12_GATE_EVAL", "13_WINDOWS_ANALYSIS"
]
SHEETS_ADV = [
    "14_5WHY_TR2_TR3", "15_FISHBONE_6M", "16_BUTFOR_SCENARIOS", "17_QSRA_INPUTS", "18_QSRA_OUTPUT"
]

DV_LISTS = {
    "SourceType": ["OFFICIAL", "AIS", "CHAT"],
    "Reliability": ["HIGH", "MED", "LOW"],
    "Verdict": ["OK", "LIMIT", "FAIL", "VERIFY"],
    "Category6M": ["Man", "Machine", "Method", "Material", "Measurement", "Environment"],
    "DistType": ["NORM", "TRI", "LOGN"],
    "MethodTag": ["APAB_WINDOW", "TIME_SLICE"],
}

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)


def _set_cols(ws, headers):
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CELL_ALIGN
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(36, len(str(h)) + 2))
    ws.freeze_panes = "A2"


def _style_body(ws, max_col):
    for r in range(2, 300):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def _add_dv_list(ws, col_letter, values, start_row=2, end_row=300):
    dv = DataValidation(type="list", formula1='"{}"'.format(",".join(values)), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def _clone_sheet_with_styles(src_ws, dst_wb, name):
    dst_ws = dst_wb.create_sheet(title=name)
    dst_ws.sheet_view.showGridLines = True
    for row in src_ws.iter_rows():
        for cell in row:
            col = cell.column if hasattr(cell, "column") else cell.col_idx
            new_cell = dst_ws.cell(row=cell.row, column=col, value=cell.value)
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                except Exception:
                    pass
            new_cell.number_format = cell.number_format
    for col_dim in list(src_ws.column_dimensions.keys()):
        if col_dim in dst_ws.column_dimensions:
            dst_ws.column_dimensions[col_dim].width = src_ws.column_dimensions[col_dim].width
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    return dst_ws


def build(wb: Workbook, tz: str):
    if "0_README" in wb.sheetnames:
        ws0 = wb["0_README"]
    else:
        ws0 = wb.create_sheet("0_README", 0)

    ws0["A1"] = "AGI Delay Analysis / SSOT v1.1 — Excel Extension v3.0 Template"
    ws0["A2"] = f"Timezone (LT): {tz}"
    ws0["A3"] = "DateTime format: YYYY-MM-DD HH:MM (LT); if UTC needed, use separate *_UTC columns."
    ws0["A4"] = "Principle: Event → Segment → Window → Cause → Evidence traceability."
    ws0["A6"] = "CORE sheets: 8~13, ADV sheets: 14~18 (optional)."

    for name in SHEETS_CORE + SHEETS_ADV:
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    if "1_SSOT_INPUTS" in wb.sheetnames:
        ws1 = wb["1_SSOT_INPUTS"]
    else:
        ws1 = wb.create_sheet("1_SSOT_INPUTS")

    if ws1.max_row < 2:
        _set_cols(ws1, ["Param", "Value", "Unit", "Notes", "Source"])
        _style_body(ws1, 5)

    recommended = [
        ("TZ", tz, "LT", "", "User"),
        ("WIND_MAX_KT", "", "kt", "Gate threshold", ""),
        ("WAVE_MAX_FT", "", "ft", "Gate threshold", ""),
        ("VIS_MIN_M", "", "m", "Gate threshold", ""),
        ("AGI_OPS_ASSUME_DAYS", "2.00", "days", "Use only if S4 missing", "Assumption"),
        ("EVID_W_OFFICIAL", "3.00", "", "", "SSOT"),
        ("EVID_W_AIS", "2.00", "", "", "SSOT"),
        ("EVID_W_CHAT", "1.00", "", "", "SSOT"),
    ]
    start_row = ws1.max_row + 2
    ws1.cell(row=start_row, column=1, value="-- Recommended Params (fill blanks) --").font = Font(bold=True)
    for i, row in enumerate(recommended, start_row + 1):
        for j, v in enumerate(row, 1):
            ws1.cell(row=i, column=j, value=v)

    ws8 = wb["8_EVENT_MAP"]
    _set_cols(ws8, ["EventKey", "EventName", "SegmentID", "SegmentName", "TR", "Planned_TS", "Actual_TS", "EvidenceID_Primary"])
    _style_body(ws8, 8)
    _add_dv_list(ws8, "C", ["S1", "S2", "S3", "S4", "S5"])

    ws9 = wb["9_SEGMENT_DELTA_TR1-TR3"]
    _set_cols(ws9, ["TR", "SegmentID", "Plan_Start", "Plan_End", "Actual_Start", "Actual_End", "Plan_Dur_h", "Actual_Dur_h", "ΔDur_h", "Driver_Top1", "EvidenceID_Top"])
    _style_body(ws9, 11)
    _add_dv_list(ws9, "B", ["S1", "S2", "S3", "S4", "S5"])
    for r in range(2, 300):
        ws9[f"G{r}"] = f'=IF(OR(C{r}="",D{r}=""),"",(D{r}-C{r})*24)'
        ws9[f"H{r}"] = f'=IF(OR(E{r}="",F{r}=""),"",(F{r}-E{r})*24)'
        ws9[f"I{r}"] = f'=IF(OR(G{r}="",H{r}=""),"",H{r}-G{r})'
        ws9[f"G{r}"].number_format = "0.00"
        ws9[f"H{r}"].number_format = "0.00"
        ws9[f"I{r}"].number_format = "0.00"

    ws10 = wb["10_EVIDENCE_LOG"]
    _set_cols(ws10, ["EvidenceID", "TR", "TS_LT", "WindowID", "CauseCode", "SourceType", "SourceName", "LocationRef", "Excerpt", "Weight", "Reliability", "LinkOrFile", "Notes"])
    _style_body(ws10, 13)
    _add_dv_list(ws10, "F", DV_LISTS["SourceType"])
    _add_dv_list(ws10, "K", DV_LISTS["Reliability"])
    for r in range(2, 300):
        ws10[f"J{r}"].number_format = "0.00"

    ws11 = wb["11_CAUSE_TAXONOMY"]
    _set_cols(ws11, ["CauseCode", "Category6M", "ShortDesc", "TypicalGate", "DefaultLikelihood", "Notes"])
    _style_body(ws11, 6)
    _add_dv_list(ws11, "B", DV_LISTS["Category6M"])
    for r in range(2, 200):
        ws11[f"E{r}"].number_format = "0.00"
    seed = [
        ("WX_FOG", "Environment", "Dense fog / low visibility", "VIS_MIN_M", "0.20", ""),
        ("WX_WIND", "Environment", "Strong wind / gust", "WIND_MAX_KT", "0.20", ""),
        ("WX_WAVE", "Environment", "High wave / sea state", "WAVE_MAX_FT", "0.15", ""),
        ("BERTH_OCCUPIED", "Method", "Berth occupied / allocation", "", "0.15", ""),
        ("PILOT_DELAY", "Man", "Pilot delay", "", "0.10", ""),
        ("DOC_PTW", "Method", "PTW / permit delay", "", "0.10", ""),
        ("DOC_MWS", "Method", "Survey/MWS delay", "", "0.05", ""),
        ("EQUIP_LINKSPAN", "Machine", "Linkspan constraint", "", "0.05", ""),
        ("EQUIP_SPMT", "Machine", "SPMT constraint", "", "0.05", ""),
        ("SAFETY_STOP", "Method", "Safety stop / hold point", "", "0.05", ""),
    ]
    for i, row in enumerate(seed, 2):
        for j, v in enumerate(row, 1):
            ws11.cell(row=i, column=j, value=v)

    ws12 = wb["12_GATE_EVAL"]
    _set_cols(ws12, ["GateID", "GateValue", "EventTS", "METAR_Wind_kt", "METAR_Gust_kt", "METAR_Vis_m", "NCM_Wind_kt", "NCM_Wave_ft", "Verdict", "EvidenceID"])
    _style_body(ws12, 10)
    _add_dv_list(ws12, "I", DV_LISTS["Verdict"])
    for r in range(2, 300):
        ws12[f"I{r}"] = f'=IF(OR(A{r}="",B{r}="",C{r}=""),"VERIFY","OK")'

    ws13 = wb["13_WINDOWS_ANALYSIS"]
    _set_cols(ws13, ["WindowID", "TR", "Start_LT", "End_LT", "ΔDelay_h", "TopCause1", "TopCause2", "EvidenceIDs", "MethodTag"])
    _style_body(ws13, 9)
    _add_dv_list(ws13, "I", DV_LISTS["MethodTag"])
    for r in range(2, 200):
        ws13[f"E{r}"].number_format = "0.00"

    ws14 = wb["14_5WHY_TR2_TR3"]
    _set_cols(ws14, ["TR", "Symptom", "Why1", "Why2", "Why3", "Why4", "Why5", "RootCause", "Countermeasure", "EvidenceIDs"])
    _style_body(ws14, 10)

    ws15 = wb["15_FISHBONE_6M"]
    _set_cols(ws15, ["TR", "Effect", "Man", "Machine", "Method", "Material", "Measurement", "Environment", "EvidenceIDs"])
    _style_body(ws15, 9)

    ws16 = wb["16_BUTFOR_SCENARIOS"]
    _set_cols(ws16, ["ScenarioID", "TR", "RemoveCauseCode", "Recalc_ETD", "Recalc_JD", "ΔFinish_days", "MethodRef", "Notes"])
    _style_body(ws16, 8)
    for r in range(2, 200):
        ws16[f"F{r}"].number_format = "0.00"

    ws17 = wb["17_QSRA_INPUTS"]
    _set_cols(ws17, ["SegmentID", "DistType", "P50_h", "P80_h", "P90_h", "Min_h", "Max_h", "SeasonalityFactor"])
    _style_body(ws17, 8)
    _add_dv_list(ws17, "B", DV_LISTS["DistType"])
    for r in range(2, 200):
        for col in ["C", "D", "E", "F", "G", "H"]:
            ws17[f"{col}{r}"].number_format = "0.00"

    ws18 = wb["18_QSRA_OUTPUT"]
    _set_cols(ws18, ["TR", "Finish_P50", "Finish_P80", "Finish_P90", "Buffer_for_P80_days"])
    _style_body(ws18, 5)
    for r in range(2, 200):
        ws18[f"E{r}"].number_format = "0.00"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True, help="output xlsx path")
    ap.add_argument("--base", default="", help="optional base xlsx to copy 0~7 sheets")
    ap.add_argument("--tz", default=TZ_DEFAULT)
    args = ap.parse_args()

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if args.base:
        wb_base = load_workbook(args.base, data_only=False)
        wb = Workbook()
        wb.remove(wb.active)
        for name in SHEETS_EXISTING:
            if name in wb_base.sheetnames:
                _clone_sheet_with_styles(wb_base[name], wb, name)
        for name in SHEETS_EXISTING:
            if name not in wb.sheetnames:
                wb.create_sheet(name)
    else:
        wb = Workbook()
        wb.active.title = "0_README"
        for name in SHEETS_EXISTING[1:]:
            wb.create_sheet(name)

    build(wb, args.tz)
    wb.save(out_path.as_posix())
    print(f"OK: saved {out_path}")


if __name__ == "__main__":
    main()
