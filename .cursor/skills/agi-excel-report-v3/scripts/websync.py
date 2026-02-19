#!/usr/bin/env python3
"""
Web evidence sync (skeleton):
- In Cursor agent runtime, use web tool to fetch links/PDFs and store in evidence_bundle/
- Then write EvidenceID rows into 10_EVIDENCE_LOG and link to 12_GATE_EVAL
This script is a placeholder for local-run; actual web retrieval is handled by the agent toolchain.
"""
import argparse
from pathlib import Path
from openpyxl import load_workbook


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--bundle_dir", default="evidence_bundle")
    args = ap.parse_args()

    Path(args.bundle_dir).mkdir(parents=True, exist_ok=True)

    wb = load_workbook(args.xlsx)
    if "10_EVIDENCE_LOG" not in wb.sheetnames:
        raise SystemExit("Missing sheet: 10_EVIDENCE_LOG")

    idx = Path(args.bundle_dir) / "00_index.md"
    if not idx.exists():
        idx.write_text("# Evidence Bundle Index\n\n- Retrieved by agent\n", encoding="utf-8")

    wb.save(args.xlsx)
    print("OK: bundle stub created; agent should populate downloads + excel rows.")


if __name__ == "__main__":
    main()
