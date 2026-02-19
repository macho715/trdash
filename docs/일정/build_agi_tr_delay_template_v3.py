"""
AGI TR Delay Template v3.0 Builder
Generates AGI_TR_Delay_Template_v3_0.xlsx (SSOT + Evidence QA + Forensic + QSRA)
References: AACE 29R-03, ASQ Fishbone, PSNet 5-Whys, NOAA METAR, NCM Al Bahar
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
from pathlib import Path

# -------------------------
# Styles (minimal + stable)
# -------------------------
thin = Side(style="thin", color="C0C0C0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")  # yellow
FILL_CALC = PatternFill("solid", fgColor="F2F2F2")  # gray
FILL_NOTE = PatternFill("solid", fgColor="E2EFDA")  # light green
FILL_OK = PatternFill("solid", fgColor="C6E0B4")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_FAIL = PatternFill("solid", fgColor="F8CBAD")
FILL_VERIFY = PatternFill("solid", fgColor="D9D9D9")

AL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = float(w)


def make_title(ws, text: str, end_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, text)
    c.font = TITLE_FONT
    c.fill = FILL_TITLE
    c.alignment = AL_LEFT
    c.border = BORDER
    ws.row_dimensions[1].height = 22


def make_header(ws, row: int, headers: list[str]):
    ws.row_dimensions[row].height = 18
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = HEADER_FONT
        c.fill = FILL_HEADER
        c.alignment = AL_CENTER
        c.border = BORDER


def add_table_safe(ws, name: str, start_row: int, end_row: int, end_col: int):
    try:
        ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
    except Exception:
        pass


def dv_list(ws, cell_range: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def cf_fill_text(ws, rng: str, text: str, fill):
    """Conditional format: fill color when cell equals text. Uses row-relative ref."""
    parts = rng.split(":")
    tl = parts[0]
    col = "".join(ch for ch in tl if ch.isalpha())
    row = "".join(ch for ch in tl if ch.isdigit())
    formula = f'=${col}{row}="{text}"'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=fill))


def cf_scale(ws, rng: str):
    rule = ColorScaleRule(
        start_type="min", start_color="C6E0B4",
        mid_type="percentile", mid_value=50, mid_color="FFF2CC",
        end_type="max", end_color="F8CBAD"
    )
    ws.conditional_formatting.add(rng, rule)


def build_table_sheet(
    ws,
    sheet_title: str,
    headers: list[str],
    n_rows: int,
    widths: dict,
    input_cols: set[int],
    formula_cols: dict | None = None,
    table_name: str | None = None
):
    formula_cols = formula_cols or {}
    make_title(ws, sheet_title, max(6, len(headers)))
    make_header(ws, 3, headers)
    first = 4
    for r in range(first, first + n_rows):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = AL_LEFT if (
                c == 1 or any(k in h.lower() for k in ["name", "desc", "excerpt", "notes", "remarks", "link", "source"])
            ) else AL_CENTER
            cell.fill = FILL_INPUT if c in input_cols else FILL_CALC
            if c in formula_cols:
                cell.value = formula_cols[c](r)
                cell.fill = FILL_CALC
    set_widths(ws, widths)
    ws.freeze_panes = "A4"
    if table_name:
        add_table_safe(ws, table_name, 3, first + n_rows - 1, len(headers))


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # ---- hidden lists for DV ----
    wsL = wb.create_sheet("_LISTS")
    wsL.sheet_state = "hidden"
    make_title(wsL, "Lists for Data Validation (do not edit)", 8)

    lists = {
        "SourceType": ["OFFICIAL", "AIS", "CHAT", "DOC", "OTHER"],
        "Reliability": ["HIGH", "MED", "LOW"],
        "Verdict": ["OK", "LIMIT", "FAIL", "VERIFY"],
        "DistType": ["NORM", "TRI", "LOGN", "EMPIRICAL"],
        "Category6M": ["Man", "Machine", "Method", "Material", "Measurement", "Environment"],
    }

    r = 3
    for nm, vals in lists.items():
        wsL.cell(r, 1, nm).font = Font(bold=True)
        for j, v in enumerate(vals, start=2):
            wsL.cell(r, j, v)
        col_s = get_column_letter(2)
        col_e = get_column_letter(1 + len(vals))
        ref = f"'_LISTS'!${col_s}${r}:${col_e}${r}"
        wb.defined_names[nm] = DefinedName(name=nm, attr_text=ref)
        r += 2
    set_widths(wsL, {"A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14})

    # ---- sheets ----
    sheets = [
        "0_README", "1_SSOT_INPUTS", "2_ACTUAL_MILESTONES", "3_ORIGINAL_PLAN", "4_DELAY_ANALYSIS_TR1-TR3",
        "5_FORECAST_TR4-TR7", "6_COMPARISON_TR1-TR7", "7_FLOAT_SUMMARY",
        "8_EVENT_MAP", "9_SEGMENT_DELTA_TR1-TR3", "10_EVIDENCE_LOG", "11_CAUSE_TAXONOMY",
        "12_GATE_EVAL", "13_WINDOWS_ANALYSIS", "14_5WHY_TR2_TR3", "15_FISHBONE_6M",
        "16_BUTFOR_SCENARIOS", "17_QSRA_INPUTS", "18_QSRA_OUTPUT"
    ]
    for s in sheets:
        wb.create_sheet(s)

    # 0_README
    ws = wb["0_README"]
    make_title(ws, "AGI TR Delay Analysis Template v3.0 (Blank) — SSOT + Evidence QA + Forensic + QSRA", 10)
    ws["A3"] = "How to use"; ws["A3"].font = Font(bold=True)
    ws["A4"] = "1) Fill YELLOW cells only. 2) Use LT(UTC+4). 3) Every claim links EvidenceID. 4) Use Segment+Windows+GateEval."
    ws["A4"].fill = FILL_NOTE
    ws["A4"].border = BORDER
    ws["A4"].alignment = AL_LEFT
    ws["A6"] = "Key references"; ws["A6"].font = Font(bold=True)
    ws["A7"] = "AACEI 29R-03; SCL Protocol; ASQ Fishbone; PSNet 5-Whys caution; NOAA METAR decode key."
    ws["A7"].fill = FILL_NOTE
    ws["A7"].border = BORDER
    ws["A7"].alignment = AL_LEFT
    ws["A9"] = "Version"; ws["B9"] = "v3.0"
    ws["A10"] = "Generated"; ws["B10"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    set_widths(ws, {"A": 24, "B": 40, "C": 18, "D": 18, "E": 18})
    ws.freeze_panes = "A3"

    # 1_SSOT_INPUTS
    ws = wb["1_SSOT_INPUTS"]
    make_title(ws, "SSOT Inputs (fill yellow)", 8)
    ssot_headers = ["Param", "Value", "Unit", "Notes", "Source", "Owner", "Required(Y/N)", "Default"]
    make_header(ws, 3, ssot_headers)
    rows = [
        ("TZ", "Asia/Dubai", "", "Timezone", "SSOT", "SCT", "Y", "Asia/Dubai"),
        ("TR3_ETD_LT", "", "LT", "User fixed ETD", "Ops", "SCT", "Y", "2026-02-19 03:00"),
        ("WIND_MAX_KT", 15.00, "kt", "Stop if >", "HM/HSE", "Port", "Y", "15.00"),
        ("WAVE_MAX_FT", "", "ft", "Stop if > (if used)", "NCM", "Marine", "N", ""),
        ("VIS_MIN_M", "", "m", "Stop if < (fog)", "METAR", "Marine", "N", ""),
        ("AGI_OPS_ASSUME_DAYS", 2.00, "days", "Assumption if missing actual", "Origin", "SCT", "Y", "2.00"),
        ("EVID_W_OFFICIAL", 3.00, "", "Evidence weight", "Rule", "SCT", "Y", "3"),
        ("EVID_W_AIS", 2.00, "", "Evidence weight", "Rule", "SCT", "Y", "2"),
        ("EVID_W_CHAT", 1.00, "", "Evidence weight", "Rule", "SCT", "Y", "1"),
        ("FORECAST_SAMPLES_N", 5000, "", "MonteCarlo samples", "Rule", "SCT", "N", "5000"),
    ]
    for i, rowv in enumerate(rows, start=4):
        for c, v in enumerate(rowv, start=1):
            cell = ws.cell(i, c, v)
            cell.border = BORDER
            cell.alignment = AL_LEFT if c in (1, 4, 5, 6) else AL_CENTER
            cell.fill = FILL_INPUT if c == 2 else FILL_CALC
    set_widths(ws, {"A": 26, "B": 24, "C": 10, "D": 40, "E": 14, "F": 12, "G": 14, "H": 18})
    ws.freeze_panes = "A4"

    # 2_ACTUAL_MILESTONES
    ws = wb["2_ACTUAL_MILESTONES"]
    hdr = ["TR", "EventKey", "EventName", "TS_LT", "TS_UTC", "Location", "SourceType", "EvidenceID", "Notes"]
    build_table_sheet(ws, "Actual Milestones (TR1-TR7)", hdr, 300,
                      {"A": 5, "B": 18, "C": 32, "D": 18, "E": 18, "F": 16, "G": 12, "H": 16, "I": 40},
                      set(range(1, 10)), table_name="ActualMilestones")
    dv_list(ws, "G4:G303", "=SourceType")

    # 3_ORIGINAL_PLAN
    ws = wb["3_ORIGINAL_PLAN"]
    hdr = ["TR", "EventKey", "Planned_TS", "Planned_Duration_h", "Baseline_Source", "Notes"]
    build_table_sheet(ws, "Original Plan", hdr, 200,
                      {"A": 5, "B": 18, "C": 18, "D": 18, "E": 18, "F": 40},
                      set(range(1, 7)), table_name="OriginalPlan")

    # 4_DELAY_ANALYSIS_TR1-TR3
    ws = wb["4_DELAY_ANALYSIS_TR1-TR3"]
    hdr = ["TR", "Milestone", "Planned_TS", "Actual_TS", "ΔDays", "ΔHours", "PrimaryCause", "EvidenceID", "Remarks"]
    form = {
        5: lambda r: f'=IF(AND(C{r}<>"",D{r}<>""),(D{r}-C{r}),"")',
        6: lambda r: f'=IF(E{r}<>"",E{r}*24,"")',
    }
    build_table_sheet(ws, "Delay Analysis (TR1-TR3)", hdr, 120,
                      {"A": 5, "B": 18, "C": 18, "D": 18, "E": 10, "F": 10, "G": 16, "H": 16, "I": 40},
                      {1, 2, 3, 4, 7, 8, 9}, formula_cols=form, table_name="DelayTR13")
    cf_scale(ws, "E4:F123")

    # 9_SEGMENT_DELTA_TR1-TR3 (핵심)
    ws = wb["9_SEGMENT_DELTA_TR1-TR3"]
    hdr = ["TR", "SegmentID", "Plan_Start", "Plan_End", "Actual_Start", "Actual_End", "Plan_Dur_h", "Actual_Dur_h",
           "ΔDur_h", "Driver_Top1", "EvidenceID_Top"]
    form = {
        7: lambda r: f'=IF(AND(C{r}<>"",D{r}<>""),(D{r}-C{r})*24,"")',
        8: lambda r: f'=IF(AND(E{r}<>"",F{r}<>""),(F{r}-E{r})*24,"")',
        9: lambda r: f'=IF(AND(G{r}<>"",H{r}<>""),(H{r}-G{r}),"")',
    }
    build_table_sheet(ws, "Segment Delta TR1-TR3", hdr, 120,
                      {"A": 5, "B": 10, "C": 18, "D": 18, "E": 18, "F": 18, "G": 12, "H": 12, "I": 10, "J": 16, "K": 16},
                      {1, 2, 3, 4, 5, 6, 10, 11}, formula_cols=form, table_name="SegDelta")
    cf_scale(ws, "I4:I123")

    # 10_EVIDENCE_LOG (핵심)
    ws = wb["10_EVIDENCE_LOG"]
    hdr = ["EvidenceID", "TR", "TS_LT", "WindowID", "CauseCode", "SourceType", "SourceName", "LocationRef", "Excerpt",
           "Weight", "Reliability", "LinkOrFile", "Notes"]
    build_table_sheet(ws, "Evidence Log", hdr, 400,
                      {"A": 16, "B": 5, "C": 18, "D": 12, "E": 16, "F": 12, "G": 18, "H": 14, "I": 52, "J": 8, "K": 10,
                       "L": 28, "M": 30},
                      {1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13}, table_name="EvidenceLog")
    dv_list(ws, "F4:F403", "=SourceType")
    dv_list(ws, "K4:K403", "=Reliability")
    cf_fill_text(ws, "F4:F403", "OFFICIAL", FILL_OK)
    cf_fill_text(ws, "F4:F403", "AIS", FILL_NOTE)
    cf_fill_text(ws, "F4:F403", "CHAT", FILL_WARN)

    # 12_GATE_EVAL (Verdict 색)
    ws = wb["12_GATE_EVAL"]
    hdr = ["GateID", "GateValue", "EventTS", "METAR_Wind_kt", "METAR_Gust_kt", "METAR_Vis_m", "NCM_Wind_kt",
           "NCM_Wave_ft", "Verdict", "EvidenceID"]
    build_table_sheet(ws, "Gate Eval", hdr, 250,
                      {"A": 16, "B": 12, "C": 18, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 10, "J": 16},
                      set(range(1, 11)), table_name="GateEval")
    dv_list(ws, "I4:I253", "=Verdict")
    cf_fill_text(ws, "I4:I253", "OK", FILL_OK)
    cf_fill_text(ws, "I4:I253", "LIMIT", FILL_WARN)
    cf_fill_text(ws, "I4:I253", "FAIL", FILL_FAIL)
    cf_fill_text(ws, "I4:I253", "VERIFY", FILL_VERIFY)

    for s in sheets:
        wb[s].sheet_view.showGridLines = True

    out_path = Path(__file__).parent / "AGI_TR_Delay_Template_v3_0.xlsx"
    wb.save(str(out_path))
    print(f"OK: {out_path}")


if __name__ == "__main__":
    main()
