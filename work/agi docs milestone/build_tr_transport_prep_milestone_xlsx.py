"""Build AGI_TR_Transport_Prep_Document_Milestones.xlsx from folder SSOT-style sources."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "AGI_TR_Transport_Prep_Document_Milestones.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, min_w: float = 10, max_w: float = 55) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        m = min_w
        for row in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > m:
                m = min(len(s) * 0.9 + 2, max_w)
        ws.column_dimensions[letter].width = m


def main() -> None:
    wb = Workbook()
    # --- Control ---
    ws_c = wb.active
    ws_c.title = "Control"
    ws_c["A1"] = "AGI TR — Transport prep & document milestones (control)"
    ws_c["A1"].font = Font(bold=True, size=12)
    ws_c["A3"] = "Trip / Cycle (e.g. TR5)"
    ws_c["B3"] = "TRn (user)"
    ws_c["A4"] = "D0 = Planned load-out date (anchor)"
    ws_c["B4"] = date(2026, 4, 26)  # default placeholder; user edits
    ws_c["A5"] = "D_sail (optional; sail-away anchor)"
    ws_c["B5"] = ""
    ws_c["A6"] = "D_agi_loadin (optional)"
    ws_c["B6"] = ""
    for addr in ("B4", "B5", "B6"):
        ws_c[addr].number_format = "yyyy-mm-dd"
    ws_c["A8"] = "Notes"
    ws_c["B8"] = (
        "Planned_Due in Master = D0 + Offset_days. "
        "Official HM/Maqta/ADNOC dates override any template. "
        "Cross-ref: AGI_TR_Submission_Documents_Milestone_Tracker (WhatsApp), "
        "AGI_TR_Submission_Milestones (PTW/V2), AGI_Email_Document_Submission_Milestone_FINAL (email)."
    )
    ws_c["B8"].alignment = WRAP
    ws_c.merge_cells("B8:M10")

    # --- TR_Prep_Master ---
    ws_m = wb.create_sheet("TR_Prep_Master")
    headers = [
        "No",
        "Gate",
        "Phase",
        "Package",
        "Document_or_Submission",
        "When_to_submit_rule",
        "Offset_days_vs_D0",
        "Planned_Due",
        "Owner_Prepare",
        "Owner_Submit_Track",
        "Approver_or_Verify",
        "Priority",
        "Hold_Y",
        "DOC_ID",
        "Evidence_IDs",
        "Risk_if_missing",
        "Next_action_hint",
    ]
    ws_m.append(headers)
    style_header(ws_m, 1, len(headers))

    # Rows: consolidated from Document_Register + prep ideas (offset vs D0 load-out anchor)
    rows: list[tuple] = [
        (1, "G0", "Prep / Engineering", "Engineering", "Ramp / Linkspan certificate (survey)", "D-7~D-2; each voyage if changed", -5, "Mammoet / Inspector / Aries", "OFCO", "HM / Port / MWS", "Critical", "Y", "DOC-001", "E001;E017;E018", "PTW/load-out hold", "Chase inspector-signed pack; revision-control"),
        (2, "G0", "Prep / Engineering", "MWS", "MWS confirmation / suitability (bow strength)", "Before loading / D0 hold", -1, "MWS / Aries / Samsung", "OFCO / Samsung", "HM / MWS", "Critical", "Y", "DOC-011", "E004;E038", "Load-out No-Go", "Separate from COA; file formal confirmation"),
        (3, "G1", "Permit / Meeting", "Permit", "PTW — Hot Work", "D-5~D-1; voyage/trip-wise", -3, "OFCO / Samsung / Mammoet", "OFCO submit", "Maqta / AD Ports HSE / HM", "Critical", "Y", "DOC-002", "E018;E069", "Hot work stop", "Validity cover slip + shift extension"),
        (4, "G1", "Permit / Meeting", "Permit", "PTW — Working Over Water", "D-5~D-1; voyage/trip-wise", -3, "OFCO / Samsung / Mammoet", "OFCO submit", "Maqta / AD Ports HSE / HM", "Critical", "Y", "DOC-003", "E018;E066", "Ramp/jetty interface stop", "Bundle with marine PTW set"),
        (5, "G1", "Permit / Meeting", "Permit", "Land PTW / SPMT heavy transport", "D-5~D0; trip-wise", -2, "OFCO / Samsung", "OFCO submit", "Maqta / AD Ports", "Critical", "Y", "DOC-004", "E007;E067;E074", "No SPMT move", "Do not assume prior trip reuse"),
        (6, "G1", "Permit / Meeting", "HSE / Ops", "MOS / MS (load in&out)", "D-10~D-5; HM comments", -7, "Mammoet", "Samsung / OFCO", "ADNOC / HM", "High", "N", "DOC-005", "E008;E085", "Approval delay", "Single revision set with RA"),
        (7, "G1", "Permit / Meeting", "HSE / Ops", "Risk Assessment / JSA", "D-5 before pre-op", -4, "Mammoet / Samsung", "OFCO", "HSE / HM", "High", "N", "DOC-006", "E022;E045", "PTW reject", "Match MOS revision"),
        (8, "G1", "Permit / Meeting", "Cargo", "Stowage Plan", "D-5 before load-out mtg", -4, "Mammoet / Eng", "OFCO", "HM / MWS", "High", "N", "DOC-007", "E045", "Deck review delay", "With drawings one pack"),
        (9, "G1", "Permit / Meeting", "Meeting", "Pre-ops / FRA minutes", "D-4~D-1", -3, "OFCO / Samsung / Mammoet", "OFCO circulate", "AD Ports / HM / HSE", "High", "N", "DOC-020", "E021;E052", "Approval drift", "Store minutes + action IDs"),
        (10, "G2", "Load-out readiness", "Access", "Gate pass / CICPA / operator access", "D-2~D0", -1, "OFCO / Samsung / Mammoet", "OFCO apply", "Port security", "Critical", "Y", "DOC-018", "E032;E073", "Crew lockout", "Per-person + operator list"),
        (11, "G2", "Load-out readiness", "Port", "Berth request / allocation / Maqta update", "D-2~D0", -1, "OFCO / LCT / Port Ops", "OFCO", "MZP / VTS / Maqta", "Critical", "N", "DOC-019", "E023;E053", "Schedule slip", "Lay-by risk note"),
        (12, "G2", "Load-out readiness", "Tide / Wx", "Tide window + weather forecast", "Daily + D0 AM", -1, "Mammoet / LCT / Port", "OFCO share", "MWS / HM / Samsung", "Critical", "Y", "DOC-021", "E034;E077", "No-Go", "Observed vs forecast log"),
        (13, "G2", "Load-out readiness", "Equipment", "Equipment booking (crane/fork/MAFI/gang) ≥24h", "D-1 / 24h before", -1, "Mammoet / OFCO / Port", "OFCO book", "GC Ops", "Medium", "N", "DOC-023", "E051;E062", "Standby cost", "Release/off-hire evidence"),
        (14, "G2", "Load-out readiness", "Cargo", "Cargo drawings / loading drawings", "D-5", -4, "Mammoet / Eng", "OFCO", "HM / MWS", "Medium", "N", "DOC-009", "E045", "Queries", "Align to stowage"),
        (15, "G3", "Load-out / securing", "Marine", "Toolbox talk + attendance", "Before each operation", 0, "Mammoet / Samsung / MWS", "OFCO file", "HSE / MWS", "High", "N", "DOC-024", "E009", "Start gate", "D0 first task"),
        (16, "G3", "Load-out / securing", "Sea fastening", "Sea fastening certificate / sail-away readiness", "After lashing / before sail", 0, "Mammoet / LCT Master / MWS", "OFCO", "AGI / HM", "Critical", "Y", "DOC-012", "E011;E055", "Cast-off delay", "LCT Master sign-off chain"),
        (17, "G3", "Load-out / securing", "Inspection", "Lashing completion / LCT Master inspection", "After load-out", 0, "Mammoet / LCT", "Samsung / MWS", "HM", "Critical", "Y", "DOC-025", "E012;E089", "Sail-away fail", "Photo + checklist"),
        (18, "G4", "Sail-away", "Cargo", "Manifest (final) + voyage number", "Before cast-off", 0, "Samsung / Shipping / LCT", "OFCO / Customs", "HM / AGI", "Critical", "Y", "DOC-008", "E039;E043", "Cast-off / AGI prep", "Revision freeze"),
        (19, "G4", "Sail-away", "MWS", "MWS COA (voyage)", "Prior sailing", 0, "MWS (NAS)", "OFCO / Samsung", "HM / LCT", "Critical", "Y", "DOC-010", "E003;E006;E013", "Departure stop", "Not same as suitability memo"),
        (20, "G4", "Sail-away", "Marine", "Vessel sailing consent", "Before cast-off", 0, "OFCO / LCT / HM", "OFCO", "HM / VTS", "High", "N", "DOC-013", "E044", "Clearance delay", "Bundle COA+manifest"),
        (21, "G4", "Sail-away", "Customs", "Custom exit / clearance evidence", "Before cast-off", 0, "Samsung / OFCO / Customs", "OFCO", "Port / Customs", "High", "N", "DOC-014", "E040", "Hold", "Match manifest"),
        (22, "G4", "Sail-away", "Marine", "Voyage plan / ETA notice", "Pre-sail + arrival", 0, "LCT / OFCO / Samsung", "OFCO", "AGI / VTS", "Medium", "N", "DOC-015", "E043", "Coordination", "ETA log"),
        (23, "G5", "AGI load-in", "AGI", "LOP (linkspan-only scope typical)", "Before AGI offload", 2, "ALS / SCT / Mammoet", "OFCO", "AOF SPA / ADNOC", "High", "N", "DOC-016", "E014;E015", "Lifting stop", "Confirm scope vs site rule"),
        (24, "G5", "AGI load-in", "AGI", "ADNOC approved TPC / lifting cert", "Before AGI lifting", 2, "Mammoet / Lifting", "OFCO", "ADNOC", "High", "Y", "DOC-017", "E016", "No lift", "TPC register per item"),
        (25, "G5", "AGI load-in", "AGI", "AGI berth / load-in plan / site readiness", "Before arrival", 1, "Samsung / AGI / LCT", "AGI team", "AGI HM / ADNOC", "High", "N", "DOC-026", "E041;E086", "Offload delay", "Security override note"),
        (26, "G6", "Return / next TR", "Demob", "MOS/PTW updates for next cycle", "D+3~D+7", 5, "Mammoet / Samsung", "OFCO", "HM / HSE", "High", "N", "DOC-005;DOC-022", "E085", "Next TR delay", "Lessons: FW/ballast in MOS"),
        (27, "G1", "Commercial / NOC", "NOC", "NOC support pack (license, award letter, RA+ERP)", "As per AD Maritime gate", -5, "Samsung", "OFCO", "AD Maritime", "High", "N", "(MS-027~041 pattern)", "", "NOC block", "From email milestone pack"),
        (28, "G0", "Vessel / PAN", "PAN / ATLP", "PAN set (crew list, registry, nav license, MDOH…)", "Before MZP berthing declare", -10, "LCT / Master", "OFCO", "Port", "High", "Y", "(MS-007)", "msg-65", "Berth stall", "ATLP reference store"),
        (29, "Idea", "Governance", "SSOT", "Document revision matrix (supersedes)", "Continuous", "", "Doc controller (Samsung/OFCO)", "Shared drive", "All", "Medium", "N", "", "", "Wrong revision on jetty", "One PDF name rule + watermark"),
        (30, "Idea", "Governance", "Evidence", "Email + portal receipt IDs in one column", "Each submit", "", "Submitter", "OFCO archive", "Internal", "Medium", "N", "", "", "Dispute", "Pastes Maqta ref #"),
        (31, "Idea", "Ops", "24h rule", "Port equipment requests", "D-1 18:00 style deadline", -1, "Mammoet", "OFCO", "GC Ops", "High", "N", "", "E062", "Standby", "Countdown plan ties PTW shifts"),
        (32, "Idea", "Permit", "Validity overlap", "PTW pack covers weather slip (+7d buffer)", "When scheduling", -7, "OFCO / Mammoet", "OFCO", "HSE", "High", "N", "", "E019", "Expiry mid-ops", "Mammoet 1-week permit lesson"),
    ]

    r0 = 2
    for i, tup in enumerate(rows):
        r = r0 + i
        (
            no,
            gate,
            phase,
            pkg,
            doc,
            when_rule,
            off,
            prep,
            subm,
            appr,
            prio,
            hold,
            did,
            eid,
            risk,
            nxt,
        ) = tup
        ws_m.cell(r, 1, no)
        ws_m.cell(r, 2, gate)
        ws_m.cell(r, 3, phase)
        ws_m.cell(r, 4, pkg)
        ws_m.cell(r, 5, doc)
        ws_m.cell(r, 6, when_rule)
        ws_m.cell(r, 7, off if isinstance(off, int) else "")
        # Planned_Due: blank if no D0 or no numeric offset
        col_g = get_column_letter(7)
        ws_m.cell(
            r,
            8,
            f'=IF(OR(Control!$B$4="",{col_g}{r}=""),"",Control!$B$4+{col_g}{r})',
        )
        ws_m.cell(r, 9, prep)
        ws_m.cell(r, 10, subm)
        ws_m.cell(r, 11, appr)
        ws_m.cell(r, 12, prio)
        ws_m.cell(r, 13, hold)
        ws_m.cell(r, 14, did)
        ws_m.cell(r, 15, eid)
        ws_m.cell(r, 16, risk)
        ws_m.cell(r, 17, nxt)
        for c in range(1, 18):
            ws_m.cell(r, c).border = BORDER
            ws_m.cell(r, c).alignment = WRAP
        ws_m.cell(r, 8).number_format = "yyyy-mm-dd"

    ws_m.freeze_panes = "A2"

    # --- Gate_Checklist ---
    ws_g = wb.create_sheet("Gate_Checklist")
    ws_g.append(["Gate", "Phase", "Typical_timing", "Milestone_check", "Linked_DOC_ids", "Owner_of_integrated_pack"])
    style_header(ws_g, 1, 6)
    gates = [
        ("G0", "Prep / Engineering", "D-14~D-7", "Linkspan cert path; MWS suitability; engineering basis", "DOC-001;DOC-011", "Samsung / Mammoet lead; OFCO interface"),
        ("G1", "Permit / Meeting", "D-5~D-1", "PTW x3; MOS/RA; pre-op/FRA; berth plan", "DOC-002~007;020", "OFCO submit; Mammoet/Samsung draft"),
        ("G2", "Load-out readiness", "D0 AM", "Gate pass; equipment; tide/weather; MWS attendance; LCT alongside", "DOC-018;019;021;023", "OFCO + Port Ops"),
        ("G3", "Load-out / securing", "D0", "Ramp crossing; lashing; sea fastening; inspections", "DOC-012;024;025", "Mammoet + LCT"),
        ("G4", "Sail-away", "D0/D+1", "Manifest; customs exit; voyage #; sailing consent; COA", "DOC-008;010;013~015", "OFCO + MWS + Customs"),
        ("G5", "AGI load-in", "D+1~D+3", "LOP/TPC; AGI berth; offload method", "DOC-016;017;026", "Samsung AGI team + ALS"),
        ("G6", "Return / next TR", "D+3~D+7", "Demob/remob; next TR MOS/PTW refresh", "DOC-005;022;027", "Mammoet / OFCO"),
    ]
    for g in gates:
        ws_g.append(list(g))
    for row in range(2, ws_g.max_row + 1):
        for c in range(1, 7):
            ws_g.cell(row=row, column=c).border = BORDER
            ws_g.cell(row=row, column=c).alignment = WRAP

    # --- RACI ---
    ws_r = wb.create_sheet("RACI_by_party")
    ws_r.append(["Party", "Primary_roles", "Typical_submit_track", "Typical_prepare_review"])
    style_header(ws_r, 1, 4)
    raci = [
        ("OFCO", "Port/HM interface; PTW submission; gate pass; berth; evidence archive", "Submit/track to Maqta/HM/Port", "Receipts, portal refs"),
        ("Samsung / SCT", "Owner docs; NOC support; coordination; AGI site", "Sign indemnities; align manifest", "Approve drafts internally"),
        ("Mammoet (MMT)", "Method/stowage; simulations; deck ops; seafastening execution", "Draft pack to OFCO", "RA/MOS/PTW attachments"),
        ("MWS", "COA; suitability; attendance", "Issue COA to OFCO/LCT", "Confirm calculations vs Aries"),
        ("LCT Bushra / KFS", "Marine docs; voyage; mooring; ramp evidence", "Master certs/PAN", "Stability/voyage plan inputs"),
        ("HM / AD Ports / Customs", "Approvals", "N/A", "Review"),
    ]
    for row in raci:
        ws_r.append(row)
    for row in range(2, ws_r.max_row + 1):
        for c in range(1, 5):
            ws_r.cell(row=row, column=c).border = BORDER
            ws_r.cell(row=row, column=c).alignment = WRAP

    # --- Ideas ---
    ws_i = wb.create_sheet("Ideas_extras")
    ws_i.append(["Topic", "Recommendation"])
    style_header(ws_i, 1, 2)
    ideas = [
        ("Single trip folder", "Per TR: 00_Index + 01_PTWHM + 02_Marine + 03_Customs + 04_AGI + 99_Receipts"),
        ("Revision SSOT", "Filename: DOCID_rev_date_author; obsolete folder, not delete"),
        ("Two-clock deadlines", "Store UAE local + UTC columns for sail-away / COA cutoffs"),
        ("Permit buffer", "Request PTW validity to cover +7d weather/berth slip (project lesson)"),
        ("Operator matrix", "SPMT operator names ↔ gate pass IDs ↔ shift (avoid last-minute ID block)"),
        ("Weather gate", "Attach METAR/NCM snapshot ID to DOC-021 row when Go/No-Go"),
        ("AGI exception path", "Emergency voyages: separate tracker row; do not reuse normal TR checklist"),
        ("Weekly governance", "Friday 09:00: red rows in Master + berth + PTW expiry review"),
    ]
    for row in ideas:
        ws_i.append(row)
    for row in range(2, ws_i.max_row + 1):
        for c in range(1, 3):
            ws_i.cell(row=row, column=c).border = BORDER
            ws_i.cell(row=row, column=c).alignment = WRAP

    autosize(ws_m)
    autosize(ws_g)
    autosize(ws_r)
    autosize(ws_i)
    ws_c.column_dimensions["A"].width = 28
    ws_c.column_dimensions["B"].width = 18

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
