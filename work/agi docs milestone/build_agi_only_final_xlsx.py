"""Build AGI-only TR transport preparation document milestone workbook.

Source: TR_Transport_Preparation_Document_Milestone_AGI_DAS.md
Output: TR_Transport_Preparation_Document_Milestone_AGI_FINAL.xlsx
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
SOURCE = BASE / "TR_Transport_Preparation_Document_Milestone_AGI_DAS.md"
OUT = BASE / "TR_Transport_Preparation_Document_Milestone_AGI_FINAL.xlsx"

EXPECTED_COUNTS = {
    "01_Master_AGI": 159,
    "02_AGI_Email_MS": 86,
    "03_AGI_Doc_Register": 27,
    "04_AGI_TR2_40": 40,
    "05_AGI_Evidence": 263,
    "06_AGI_Gate_Hold": 7,
    "07_Source_Map": 6,
    "08_Open_Actions": 8,
}

HEADER_FILL = PatternFill("solid", fgColor="17365D")
TITLE_FILL = PatternFill("solid", fgColor="0B1020")
SUBTITLE_FILL = PatternFill("solid", fgColor="EAF2F8")
CRITICAL_FILL = PatternFill("solid", fgColor="F8696B")
HIGH_FILL = PatternFill("solid", fgColor="F4B183")
HOLD_FILL = PatternFill("solid", fgColor="FFD966")
OPEN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def split_markdown_row(line: str) -> list[str]:
    text = line.strip()
    if text.startswith("|"):
        text = text[1:]
    if text.endswith("|"):
        text = text[:-1]
    cells: list[str] = []
    buf: list[str] = []
    i = 0
    while i < len(text):
        ch = text[i]
        if ch == "\\" and i + 1 < len(text):
            buf.append(text[i + 1])
            i += 2
            continue
        if ch == "|":
            cells.append(clean_cell("".join(buf)))
            buf = []
        else:
            buf.append(ch)
        i += 1
    cells.append(clean_cell("".join(buf)))
    return cells


def clean_cell(value: str) -> str:
    value = value.strip()
    if value == "NaN":
        return ""
    value = value.replace("\\_", "_").replace("\\*", "*")
    value = value.replace("filecite", "filecite:").replace("", "")
    return value


def is_separator(row: list[str]) -> bool:
    return bool(row) and all(re.fullmatch(r":?-{2,}:?", c.strip()) for c in row if c.strip())


def load_sections() -> dict[str, list[str]]:
    lines = SOURCE.read_text(encoding="utf-8", errors="replace").splitlines()
    sections: dict[str, list[str]] = {}
    current: str | None = None
    for line in lines:
        if line.startswith("## "):
            current = line[3:].strip()
            sections[current] = []
            continue
        if current:
            sections[current].append(line)
    return sections


def parse_section(sections: dict[str, list[str]], name: str, first_header: str) -> tuple[list[str], list[list[str]]]:
    rows = [split_markdown_row(line) for line in sections[name] if line.strip().startswith("|")]
    rows = [row for row in rows if not is_separator(row)]
    header_idx = None
    for idx, row in enumerate(rows):
        if row and row[0] == first_header:
            header_idx = idx
            break
    if header_idx is None:
        raise ValueError(f"Cannot find header '{first_header}' in section {name}")
    headers = rows[header_idx]
    data: list[list[str]] = []
    for row in rows[header_idx + 1 :]:
        if not any(cell for cell in row):
            continue
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        if len(row) > len(headers):
            row = row[: len(headers) - 1] + [" | ".join(row[len(headers) - 1 :])]
        data.append(row)
    return headers, data


def make_unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for idx, header in enumerate(headers, start=1):
        base = header or f"Column_{idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        result.append(base if count == 1 else f"{base}_{count}")
    return result


def safe_table_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_]", "_", name)


def add_data_sheet(
    wb: Workbook,
    title: str,
    source_note: str,
    headers: list[str],
    rows: list[list[str]],
    table_name: str,
) -> None:
    ws = wb.create_sheet(title)
    headers = make_unique_headers(headers)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = TITLE_FILL
    ws["A2"] = source_note
    ws["A2"].fill = SUBTITLE_FILL
    ws["A2"].alignment = WRAP
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_row = ws.max_row
    last_col = ws.max_column
    style_header(ws, 4, last_col)
    for row in ws.iter_rows(min_row=5, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    ref = f"A4:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=safe_table_name(table_name), ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.auto_filter.ref = ref
    ws.freeze_panes = "A5"
    add_common_conditional_formats(ws, headers, last_row)
    autosize(ws)


def style_header(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def add_common_conditional_formats(ws, headers: list[str], last_row: int) -> None:
    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        rng = f"{col}5:{col}{last_row}"
        lowered = header.lower()
        if "priority" in lowered:
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'OR(UPPER({col}5)="CRITICAL",UPPER({col}5)="C")'], fill=CRITICAL_FILL),
            )
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'OR(UPPER({col}5)="HIGH",UPPER({col}5)="H")'], fill=HIGH_FILL),
            )
        if "hold" in lowered or header == "Hold_Flag":
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'OR(UPPER({col}5)="Y",{col}5=1)'], fill=HOLD_FILL),
            )
        if "open_flag" in lowered:
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["1"], fill=OPEN_FILL))
        if "status" in lowered:
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'ISNUMBER(SEARCH("Completed",{col}5))'], fill=OK_FILL),
            )


def autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            width = max(width, min(55, len(str(value)) * 0.9 + 2))
        ws.column_dimensions[letter].width = width
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 18 if row > 4 else 24


def build_dashboard(wb: Workbook, counts: dict[str, int]) -> None:
    ws = wb.active
    ws.title = "00_Dashboard"
    ws["A1"] = "TR Transport Preparation - AGI Document Milestone Final"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = TITLE_FILL
    ws.merge_cells("A1:J1")
    ws["A2"] = "Scope: AGI only. DAS rows/sheets/sources excluded. Generated from AGI_DAS source on " + datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells("A2:J2")
    ws["A2"].alignment = WRAP
    ws["A2"].fill = SUBTITLE_FILL

    dashboard_rows = [
        ("KPI", "Formula / Value", "Interpretation"),
        ("AGI master milestone/document rows", "='01_Master_AGI'!A5*0+COUNTA('01_Master_AGI'!C5:C163)", "Expected 159 AGI rows"),
        ("AGI email milestone rows", "=COUNTA('02_AGI_Email_MS'!A5:A90)", "Expected 86 rows"),
        ("AGI document register rows", "=COUNTA('03_AGI_Doc_Register'!A5:A31)", "Expected 27 rows"),
        ("AGI TR2 detailed milestone rows", "=COUNTA('04_AGI_TR2_40'!A5:A44)", "Expected 40 rows"),
        ("AGI evidence rows", "=COUNTA('05_AGI_Evidence'!A5:A267)", "Expected 263 evidence entries"),
        ("Open / pending / review rows", "=COUNTIF('01_Master_AGI'!S5:S163,1)", "Expected 112 from AGI master Open_Flag"),
        ("Hold / critical rows", "=COUNTIF('01_Master_AGI'!T5:T163,1)", "Expected 101 from AGI master Hold_Flag"),
        ("AGI gate rows", "=COUNTA('06_AGI_Gate_Hold'!A5:A11)", "AGI G0-G6 only"),
        ("AGI source files mapped", "=COUNTA('07_Source_Map'!A5:A10)", "S-AGI sources only"),
        ("AGI open actions", "=COUNTA('08_Open_Actions'!A5:A12)", "Expected 8 open actions"),
    ]
    for r, row in enumerate(dashboard_rows, start=4):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = WRAP
        if r == 4:
            style_header(ws, r, 3)

    ws["E4"] = "Source Track"
    ws["F4"] = "Rows"
    style_header(ws, 4, 6)
    source_rows = [
        ("Email Milestone", 86),
        ("Document Register", 27),
        ("TR2 Detailed 40", 40),
        ("Method/Basis rows", 6),
        ("Evidence", 263),
        ("Open Actions", 8),
    ]
    for r, (name, value) in enumerate(source_rows, start=5):
        ws.cell(r, 5, name)
        ws.cell(r, 6, value)
        ws.cell(r, 5).border = BORDER
        ws.cell(r, 6).border = BORDER

    chart = BarChart()
    chart.title = "AGI source row coverage"
    chart.y_axis.title = "Rows"
    chart.x_axis.title = "Source"
    data = Reference(ws, min_col=6, min_row=4, max_row=10)
    cats = Reference(ws, min_col=5, min_row=5, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "E13")

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "A4"


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("09_README")
    rows = [
        ("Field", "Value"),
        ("Workbook purpose", "AGI-only TR transport preparation document milestone control workbook."),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source", "TR_Transport_Preparation_Document_Milestone_AGI_DAS.md"),
        ("AGI basis", "AGI email milestone tracker; AGI document register; AGI TR2 40-line submission milestone file; AGI method/basis rows."),
        ("DAS handling", "DAS milestones, gates, sources, and dashboard references are excluded from this AGI final version."),
        ("Important caveat", "Trackers derived from uploaded messages/reports do not replace official HM/Maqta/ADNOC/MWS approval documents."),
        ("Included sheets", "00_Dashboard, 01_Master_AGI, 02_AGI_Email_MS, 03_AGI_Doc_Register, 04_AGI_TR2_40, 05_AGI_Evidence, 06_AGI_Gate_Hold, 07_Source_Map, 08_Open_Actions, 09_README."),
        ("Filter policy", "Do not delete rows by raw DAS text. Some AGI evidence rows mention DAS in message content and are intentionally retained."),
    ]
    ws["A1"] = "README - AGI Final Workbook Scope and Caveats"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = TITLE_FILL
    ws.merge_cells("A1:B1")
    ws.append([])
    ws.append([])
    for row in rows:
        ws.append(row)
    style_header(ws, 4, 2)
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    tab = Table(displayName="ReadmeTbl", ref=f"A4:B{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    ws.auto_filter.ref = tab.ref
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100


def validate_workbook(path: Path, counts: dict[str, int]) -> list[str]:
    wb = load_workbook(path, data_only=False)
    errors: list[str] = []
    expected_sheets = [
        "00_Dashboard",
        "01_Master_AGI",
        "02_AGI_Email_MS",
        "03_AGI_Doc_Register",
        "04_AGI_TR2_40",
        "05_AGI_Evidence",
        "06_AGI_Gate_Hold",
        "07_Source_Map",
        "08_Open_Actions",
        "09_README",
    ]
    for sheet in expected_sheets:
        if sheet not in wb.sheetnames:
            errors.append(f"Missing sheet: {sheet}")
    for sheet in wb.sheetnames:
        if re.search(r"(^|_)DAS($|_)", sheet.upper()):
            errors.append(f"DAS sheet remains: {sheet}")
    for sheet, expected in EXPECTED_COUNTS.items():
        actual = counts.get(sheet)
        if actual != expected:
            errors.append(f"{sheet} row count {actual} != {expected}")
    if "01_Master_AGI" in wb.sheetnames:
        ws = wb["01_Master_AGI"]
        site_values = [ws.cell(row=r, column=1).value for r in range(5, 5 + EXPECTED_COUNTS["01_Master_AGI"])]
        if any(value != "AGI" for value in site_values):
            errors.append("01_Master_AGI has non-AGI Site values")
    if "06_AGI_Gate_Hold" in wb.sheetnames:
        ws = wb["06_AGI_Gate_Hold"]
        gates = [ws.cell(row=r, column=2).value for r in range(5, 12)]
        if gates != [f"G{i}" for i in range(7)]:
            errors.append(f"AGI gates mismatch: {gates}")
    if "07_Source_Map" in wb.sheetnames:
        ws = wb["07_Source_Map"]
        source_ids = [ws.cell(row=r, column=1).value for r in range(5, 11)]
        if any(not str(value).startswith("S-AGI") for value in source_ids):
            errors.append(f"Source map includes non-AGI source: {source_ids}")
    for ws in wb.worksheets:
        if ws.title != "00_Dashboard" and len(ws.tables) == 0:
            errors.append(f"No Excel table in {ws.title}")
        if ws.freeze_panes is None:
            errors.append(f"No freeze pane in {ws.title}")
    return errors


def main() -> None:
    sections = load_sections()
    wb = Workbook()
    counts: dict[str, int] = {}
    build_dashboard(wb, counts)

    master_headers, master_rows = parse_section(sections, "01_Master_Combined", "Site")
    master_rows = [row for row in master_rows if row[0] == "AGI"]
    counts["01_Master_AGI"] = len(master_rows)
    add_data_sheet(wb, "01_Master_AGI", "AGI rows only from 01_Master_Combined. DAS rows excluded by Site column.", master_headers, master_rows, "MasterAGITbl")

    email_headers, email_rows = parse_section(sections, "02_AGI_Email_MS", "No")
    counts["02_AGI_Email_MS"] = len(email_rows)
    add_data_sheet(wb, "02_AGI_Email_MS", "AGI email-derived milestone tracker.", email_headers, email_rows, "AGIEmailMSTbl")

    doc_headers, doc_rows = parse_section(sections, "03_AGI_Doc_Register", "Doc_ID")
    counts["03_AGI_Doc_Register"] = len(doc_rows)
    add_data_sheet(wb, "03_AGI_Doc_Register", "AGI WhatsApp-derived document register.", doc_headers, doc_rows, "AGIDocRegisterTbl")

    tr2_headers, tr2_rows = parse_section(sections, "04_AGI_TR2_40", "No")
    counts["04_AGI_TR2_40"] = len(tr2_rows)
    add_data_sheet(wb, "04_AGI_TR2_40", "AGI TR2 detailed 40-line submission milestone reference.", tr2_headers, tr2_rows, "AGITR2MilestonesTbl")

    ev_headers, ev_rows = parse_section(sections, "05_AGI_Evidence", "Evidence Source")
    counts["05_AGI_Evidence"] = len(ev_rows)
    add_data_sheet(wb, "05_AGI_Evidence", "AGI evidence log. Rows may mention DAS in message content and are retained.", ev_headers, ev_rows, "AGIEvidenceTbl")

    gate_headers, gate_rows = parse_section(sections, "07_Gate_Hold_Matrix", "Site")
    gate_rows = [row for row in gate_rows if row[0] == "AGI"]
    counts["06_AGI_Gate_Hold"] = len(gate_rows)
    add_data_sheet(wb, "06_AGI_Gate_Hold", "AGI G0-G6 gate and hold matrix only.", gate_headers, gate_rows, "AGIGateHoldTbl")

    source_headers, source_rows = parse_section(sections, "08_Source_Map", "Source_ID")
    source_rows = [row for row in source_rows if row[0].startswith("S-AGI")]
    counts["07_Source_Map"] = len(source_rows)
    add_data_sheet(wb, "07_Source_Map", "AGI source map only. S-DAS sources excluded.", source_headers, source_rows, "AGISourceMapTbl")

    action_headers, action_rows = parse_section(sections, "09_Open_Actions", "Action_ID")
    counts["08_Open_Actions"] = len(action_rows)
    add_data_sheet(wb, "08_Open_Actions", "AGI open actions from tracker.", action_headers, action_rows, "AGIOpenActionsTbl")

    build_readme(wb)
    wb.save(OUT)

    errors = validate_workbook(OUT, counts)
    if errors:
        raise SystemExit("Validation failed:\n" + "\n".join(errors))
    print(f"Wrote {OUT}")
    for sheet, expected in EXPECTED_COUNTS.items():
        print(f"PASS {sheet}: {counts[sheet]} rows")


if __name__ == "__main__":
    main()
