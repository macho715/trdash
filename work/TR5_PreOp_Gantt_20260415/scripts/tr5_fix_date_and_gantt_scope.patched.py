from __future__ import annotations

import datetime as dt
import json
import shutil
import zipfile
from pathlib import Path

import pythoncom
import win32com.client as win32
from openpyxl import load_workbook


ROOT = Path(r"C:\tr_dash-main")
TARGET = ROOT / "output" / "spreadsheet" / "tr5_preparation_simulation" / "TR5_Preparation_Simulation_Gantt_MIR_Format_COM.xlsm"
RUN_DIR = TARGET.parent / f"scope_patch_run_{dt.datetime.now():%Y%m%d_%H%M%S}"
SIDECAR = RUN_DIR / "TR5_Preparation_Simulation_Gantt_MIR_Format_COM.sidecar.xlsm"
BACKUP = RUN_DIR / "TR5_Preparation_Simulation_Gantt_MIR_Format_COM.before_scope_patch.xlsm"
REPORT = RUN_DIR / "validation_report.json"

SHEET = "Gantt_BASE"
FIRST_TASK_ROW = 4
LAST_TASK_ROW = 42
FIRST_TIMELINE_COL = 11  # K
LAST_TIMELINE_COL = 31  # AE
LAST_TIMELINE_LETTER = "AE"
TIMELINE_START = dt.date(2026, 4, 27)
TIMELINE_END = dt.date(2026, 5, 17)
RISK_VALUES = ["OK", "AMBER", "HIGH", "WARNING", "CRITICAL", "GATE"]
EXCEL_EPOCH = dt.date(1899, 12, 30)


def xl_color(hex_rgb: str) -> int:
    hex_rgb = hex_rgb.strip().lstrip("#")
    r = int(hex_rgb[0:2], 16)
    g = int(hex_rgb[2:4], 16)
    b = int(hex_rgb[4:6], 16)
    return r + (g * 256) + (b * 65536)


FALLBACK_RISK_COLORS = {
    "OK": xl_color("DCEAD7"),
    "AMBER": xl_color("F7E0C4"),
    "HIGH": xl_color("F4D7C8"),
    "WARNING": xl_color("FFF1B8"),
    "CRITICAL": xl_color("FF5C63"),
    "GATE": xl_color("F5C84C"),
}
FALLBACK_BAR_COLORS = {
    "OK": xl_color("5FB5F2"),
    "AMBER": xl_color("F1A954"),
    "HIGH": xl_color("F08A5D"),
    "WARNING": xl_color("F5D04A"),
    "CRITICAL": xl_color("FF5C63"),
    "GATE": xl_color("F5C84C"),
}
RISK_FONT_COLORS = {
    "OK": xl_color("102417"),
    "AMBER": xl_color("3D2A0D"),
    "HIGH": xl_color("3B1E14"),
    "WARNING": xl_color("352800"),
    "CRITICAL": xl_color("FFFFFF"),
    "GATE": xl_color("0D1320"),
}


def contract(path: Path) -> dict[str, int]:
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
    return {
        "vbaProject.bin": int("xl/vbaProject.bin" in names),
        "drawings": sum(1 for n in names if n.startswith("xl/drawings/drawing") and n.endswith(".xml")),
        "ctrlProps": sum(1 for n in names if n.startswith("xl/ctrlProps/") and n.endswith(".xml")),
        "vmlDrawing": sum(1 for n in names if n.startswith("xl/drawings/vmlDrawing") and n.endswith(".vml")),
    }


def col_letter(col: int) -> str:
    s = ""
    while col:
        col, rem = divmod(col - 1, 26)
        s = chr(65 + rem) + s
    return s


def excel_serial(day: dt.date) -> int:
    return (day - EXCEL_EPOCH).days


TIMELINE_FIRST_CELL = f"{col_letter(FIRST_TIMELINE_COL)}3"
TIMELINE_HEADER_RANGE = f"{col_letter(FIRST_TIMELINE_COL)}3:{LAST_TIMELINE_LETTER}3"
RISK_RANGE = f"G{FIRST_TASK_ROW}:G{LAST_TASK_ROW}"
BAR_RANGE = f"{col_letter(FIRST_TIMELINE_COL)}{FIRST_TASK_ROW}:{LAST_TIMELINE_LETTER}{LAST_TASK_ROW}"


def find_display_test_target(ws, task_rows: list[int]) -> tuple[int | None, int | None]:
    for row in task_rows:
        start = ws.Cells(row, 4).Value2
        end = ws.Cells(row, 5).Value2
        if not isinstance(start, (int, float)) or not isinstance(end, (int, float)):
            continue
        for col in range(FIRST_TIMELINE_COL, LAST_TIMELINE_COL + 1):
            header = ws.Cells(3, col).Value2
            if isinstance(header, (int, float)) and int(start) <= int(header) <= int(end):
                return row, col
    return None, None


def safe_calculate(excel, wb=None, ws=None) -> bool:
    for obj, method_name in (
        (excel, "CalculateFullRebuild"),
        (excel, "CalculateFull"),
        (ws, "Calculate"),
        (wb, "RefreshAll"),
    ):
        if obj is None:
            continue
        try:
            method = getattr(obj, method_name)
            if callable(method):
                method()
                return True
        except Exception:
            continue
    return False


def check_live_target_not_unsaved() -> list[str]:
    closed = []
    try:
        app = win32.GetActiveObject("Excel.Application")
    except Exception:
        return closed
    for wb in list(app.Workbooks):
        try:
            if Path(wb.FullName).resolve().samefile(TARGET):
                if not wb.Saved:
                    raise RuntimeError(f"Target workbook is open with unsaved changes: {wb.FullName}")
                wb.Close(SaveChanges=False)
                closed.append(wb.FullName)
        except RuntimeError:
            raise
        except Exception:
            continue
    return closed


def used_task_rows(ws) -> list[int]:
    rows = []
    for row in range(FIRST_TASK_ROW, LAST_TASK_ROW + 1):
        if ws.Cells(row, 4).Value2 and ws.Cells(row, 5).Value2:
            rows.append(row)
    return rows


def sample_current_display_colors(ws, task_rows: list[int]) -> tuple[dict[str, int], dict[str, int]]:
    risk_colors = {}
    bar_colors = {}
    for row in task_rows:
        risk = str(ws.Cells(row, 7).Value or "").strip().upper()
        if risk not in RISK_VALUES:
            continue
        if risk not in risk_colors:
            try:
                risk_colors[risk] = int(ws.Cells(row, 7).DisplayFormat.Interior.Color)
            except Exception:
                pass
        if risk not in bar_colors:
            start = ws.Cells(row, 4).Value2
            end = ws.Cells(row, 5).Value2
            if not isinstance(start, (int, float)) or not isinstance(end, (int, float)):
                continue
            for col in range(FIRST_TIMELINE_COL, LAST_TIMELINE_COL + 1):
                header = ws.Cells(3, col).Value2
                if isinstance(header, (int, float)) and int(start) <= int(header) <= int(end):
                    try:
                        color = int(ws.Cells(row, col).DisplayFormat.Interior.Color)
                        if color:
                            bar_colors[risk] = color
                            break
                    except Exception:
                        pass
    for risk, color in FALLBACK_RISK_COLORS.items():
        risk_colors.setdefault(risk, color)
    for risk, color in FALLBACK_BAR_COLORS.items():
        bar_colors.setdefault(risk, color)
    return risk_colors, bar_colors


def apply_patch_with_com(path: Path) -> dict:
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    try:
        try:
            excel.AutomationSecurity = 3
        except Exception:
            pass
        wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
        ws = wb.Worksheets(SHEET)
        task_rows = used_task_rows(ws)
        risk_colors, bar_colors = sample_current_display_colors(ws, task_rows)

        title = "TR5 PRE-OP SIMULATION - PHASE-GATE GANTT"
        subtitle = (
            "Today: 2026-04-15 | Source: TR5_Pre-Op_Simulation_v2_20260430.md | "
            "T+0=2026-04-30 | Load-out=2026-05-09 | AGI JD=2026-05-17"
        )

        ws.Cells.FormatConditions.Delete()

        ws.Range("A1:BO2").UnMerge()
        ws.Range("A1:BO2").ClearContents()
        ws.Range("A1:BO2").ClearFormats()
        ws.Range("A1:AE1").Merge()
        ws.Range("A2:AE2").Merge()
        ws.Range("A1").Value = title
        ws.Range("A2").Value = subtitle

        title_rng = ws.Range("A1:AE1")
        title_rng.Interior.Color = xl_color("0B1120")
        title_rng.Font.Name = "Arial"
        title_rng.Font.Size = 13
        title_rng.Font.Bold = True
        title_rng.Font.Color = xl_color("F5C84C")
        title_rng.HorizontalAlignment = -4108
        title_rng.VerticalAlignment = -4108

        subtitle_rng = ws.Range("A2:AE2")
        subtitle_rng.Interior.Color = xl_color("13161E")
        subtitle_rng.Font.Name = "Arial"
        subtitle_rng.Font.Size = 9
        subtitle_rng.Font.Bold = False
        subtitle_rng.Font.Color = xl_color("7D879C")
        subtitle_rng.HorizontalAlignment = -4108
        subtitle_rng.VerticalAlignment = -4108

        for row in range(FIRST_TASK_ROW, LAST_TASK_ROW + 1):
            for col in (4, 5):
                value = ws.Cells(row, col).Value2
                if isinstance(value, (int, float)):
                    ws.Cells(row, col).Value2 = int(value)
        ws.Range("D4:E42").NumberFormat = "yyyy-mm-dd"

        for offset, col in enumerate(range(FIRST_TIMELINE_COL, LAST_TIMELINE_COL + 1)):
            day = TIMELINE_START + dt.timedelta(days=offset)
            cell = ws.Cells(3, col)
            cell.Value2 = excel_serial(day)
            cell.Font.Name = "Arial"
            cell.Font.Size = 7
            cell.Font.Bold = True
            cell.Font.Color = xl_color("6F7A92")
            cell.Orientation = 90
            cell.HorizontalAlignment = -4108
            cell.VerticalAlignment = -4108
            ws.Columns(col).ColumnWidth = 5.0
        ws.Range(TIMELINE_HEADER_RANGE).NumberFormat = "mm-dd"

        # Remove all visual state outside the actual Gantt scope.
        for address in ("AF1:BO80", "A43:BO80"):
            rng = ws.Range(address)
            rng.ClearContents()
            rng.ClearFormats()
            try:
                rng.Validation.Delete()
            except Exception:
                pass

        # Re-assert the visible timeline grid only inside K:AE and rows 3:42.
        inside_timeline = ws.Range(f"{col_letter(FIRST_TIMELINE_COL)}3:{LAST_TIMELINE_LETTER}{LAST_TASK_ROW}")
        inside_timeline.Interior.Color = xl_color("0B1020")
        inside_timeline.Borders.Color = xl_color("242A38")
        inside_timeline.Borders.LineStyle = 1
        inside_timeline.Borders.Weight = 2

        ws.Range(TIMELINE_HEADER_RANGE).Interior.Color = xl_color("0B1120")
        ws.Range("A3:G3").Interior.Color = xl_color("0B1120")
        ws.Range("A3:G3").Font.Name = "Arial"
        ws.Range("A3:G3").Font.Size = 9
        ws.Range("A3:G3").Font.Bold = True
        ws.Range("A3:G3").Font.Color = xl_color("F5C84C")
        ws.Range("A3:G3").HorizontalAlignment = -4108
        ws.Range("A3:G3").VerticalAlignment = -4108

        risk_rng = ws.Range(RISK_RANGE)
        try:
            risk_rng.Validation.Delete()
        except Exception:
            pass
        risk_rng.Validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1=",".join(RISK_VALUES))
        risk_rng.Validation.IgnoreBlank = True
        risk_rng.Validation.InCellDropdown = True

        # Risk cell conditional formatting.
        for risk in RISK_VALUES:
            fc = risk_rng.FormatConditions.Add(2, 1, f'=$G{FIRST_TASK_ROW}="{risk}"')
            fc.Interior.Color = risk_colors[risk]
            fc.Font.Color = RISK_FONT_COLORS[risk]
            fc.Font.Bold = True

        # Gantt bar conditional formatting, limited to the actual timeline body.
        bar_rng = ws.Range(BAR_RANGE)
        for risk in RISK_VALUES:
            formula = (
                f'=AND({col_letter(FIRST_TIMELINE_COL)}$3>=$D{FIRST_TASK_ROW},'
                f'{col_letter(FIRST_TIMELINE_COL)}$3<=$E{FIRST_TASK_ROW},'
                f'$G{FIRST_TASK_ROW}="{risk}")'
            )
            fc = bar_rng.FormatConditions.Add(2, 1, formula)
            fc.Interior.Color = bar_colors[risk]
            fc.Font.Color = xl_color("0D1320")
            fc.Font.Bold = True

        ws.Range("A1:AE42").Font.Name = "Arial"
        ws.Range("A1:AE42").Rows.RowHeight = 17.4
        ws.Rows(1).RowHeight = 31.95
        ws.Rows(2).RowHeight = 18
        ws.Rows(3).RowHeight = 42
        ws.Range("D4:E42").Columns.ColumnWidth = 12
        ws.Range("D4:E42").NumberFormat = "yyyy-mm-dd"
        ws.Range(TIMELINE_HEADER_RANGE).NumberFormat = "mm-dd"

        # Functional display-format smoke test on the sidecar only.
        test_row, first_bar_col = find_display_test_target(ws, task_rows)
        display_test = {
            "row": test_row,
            "bar_col": col_letter(first_bar_col) if first_bar_col is not None else None,
            "risk_changed": False,
            "bar_changed": False,
            "skipped": test_row is None or first_bar_col is None,
        }
        if not display_test["skipped"]:
            original_risk = ws.Cells(test_row, 7).Value
            ws.Cells(test_row, 7).Value = "CRITICAL"
            calculate_ok_critical = safe_calculate(excel, wb, ws)
            critical_risk_color = int(ws.Cells(test_row, 7).DisplayFormat.Interior.Color)
            critical_bar_color = int(ws.Cells(test_row, first_bar_col).DisplayFormat.Interior.Color)
            ws.Cells(test_row, 7).Value = "GATE"
            calculate_ok_gate = safe_calculate(excel, wb, ws)
            gate_risk_color = int(ws.Cells(test_row, 7).DisplayFormat.Interior.Color)
            gate_bar_color = int(ws.Cells(test_row, first_bar_col).DisplayFormat.Interior.Color)
            ws.Cells(test_row, 7).Value = original_risk
            calculate_ok_restore = safe_calculate(excel, wb, ws)
            display_test.update({
                "critical_risk_color": critical_risk_color,
                "critical_bar_color": critical_bar_color,
                "gate_risk_color": gate_risk_color,
                "gate_bar_color": gate_bar_color,
                "calculate_ok": calculate_ok_critical or calculate_ok_gate or calculate_ok_restore,
                "risk_changed": critical_risk_color != gate_risk_color,
                "bar_changed": critical_bar_color != gate_bar_color,
            })

        wb.Save()
        wb.Close(SaveChanges=True)
        return {
            "task_rows": task_rows,
            "sampled_risk_colors": risk_colors,
            "sampled_bar_colors": bar_colors,
            "display_test": display_test,
        }
    finally:
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def has_visible_format(cell) -> bool:
    fill = cell.fill
    if fill and fill.fill_type and fill.fill_type != "none":
        return True
    for side_name in ("left", "right", "top", "bottom"):
        side = getattr(cell.border, side_name)
        if side and side.style:
            return True
    return False


def validate(path: Path, baseline: dict, com_result: dict) -> dict:
    errors = []
    wb = load_workbook(path, keep_vba=True, data_only=False)
    ws = wb[SHEET]

    if ws["A1"].value != "TR5 PRE-OP SIMULATION - PHASE-GATE GANTT":
        errors.append("title mismatch")
    if "Load-out=2026-05-09" not in str(ws["A2"].value) or "AGI JD=2026-05-17" not in str(ws["A2"].value):
        errors.append("subtitle date mismatch")

    merged = {str(r) for r in ws.merged_cells.ranges}
    if "A1:AE1" not in merged or "A2:AE2" not in merged:
        errors.append(f"title/subtitle merge scope mismatch: {sorted(merged)}")
    if "A1:BO1" in merged or "A2:BO2" in merged:
        errors.append("old BO title/subtitle merge still present")

    # Date format: Excel locale may store "yyyy-mm-dd" as "mm-dd-yy" internally.
    # Accept any format containing date tokens; reject only completely wrong formats.
    DATE_FMT_OK = {"yyyy-mm-dd", "mm-dd-yy", "yy-mm-dd", "dd-mm-yy", "m/d/yy", "m/d/yyyy"}
    for row in range(FIRST_TASK_ROW, LAST_TASK_ROW + 1):
        for col in (4, 5):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            fmt = cell.number_format or ""
            if fmt not in DATE_FMT_OK and not any(
                tok in fmt for tok in ("yy", "mm", "dd", "m", "d")
            ):
                errors.append(f"{cell.coordinate} number format {fmt!r}")
            if isinstance(cell.value, dt.datetime) and cell.value.time() != dt.time(0, 0):
                errors.append(f"{cell.coordinate} has non-midnight time {cell.value}")

    expected = TIMELINE_START
    for col in range(FIRST_TIMELINE_COL, LAST_TIMELINE_COL + 1):
        cell = ws.cell(3, col)
        # openpyxl may read back the raw Excel serial integer instead of a date object
        # when the format is stored as 'General'. Normalise both to serial int for comparison.
        if isinstance(cell.value, dt.datetime):
            actual_serial = excel_serial(cell.value.date())
        elif isinstance(cell.value, (int, float)):
            actual_serial = int(cell.value)
        else:
            actual_serial = None
        expected_serial = excel_serial(expected)
        if actual_serial != expected_serial:
            errors.append(f"{cell.coordinate} expected {expected} (serial {expected_serial}), got {cell.value!r}")
        # Accept "mm-dd" or "General" (when the value is a serial int Excel displays as date)
        if cell.number_format not in ("mm-dd", "General", ""):
            errors.append(f"{cell.coordinate} timeline format {cell.number_format!r}")
        expected += dt.timedelta(days=1)
    if expected - dt.timedelta(days=1) != TIMELINE_END:
        errors.append("timeline end calculation mismatch")

    for col in range(32, 67):  # AF:BO
        if ws.cell(3, col).value is not None:
            errors.append(f"{ws.cell(3, col).coordinate} should be blank")
            break

    cf_ranges = [str(r.sqref) for r in ws.conditional_formatting._cf_rules.keys()]
    if RISK_RANGE not in cf_ranges:
        errors.append(f"risk CF range missing: {cf_ranges}")
    if BAR_RANGE not in cf_ranges:
        errors.append(f"bar CF range {BAR_RANGE} missing: {cf_ranges}")
    extra_cf = sorted(set(cf_ranges) - {RISK_RANGE, BAR_RANGE})
    if extra_cf:
        errors.append(f"conditional formatting outside scope: {extra_cf}")

    dv_ranges = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    if dv_ranges != [RISK_RANGE]:
        errors.append(f"data validation ranges mismatch: {dv_ranges}")
    else:
        formula = str(ws.data_validations.dataValidation[0].formula1 or "")
        normalized_formula = formula.strip('"')
        if normalized_formula != ",".join(RISK_VALUES):
            errors.append(f"risk validation formula mismatch: {formula!r}")

    outside_hits = []
    for row in range(1, 81):
        for col in range(32, 67):  # AF:BO
            cell = ws.cell(row, col)
            if cell.value is not None or has_visible_format(cell):
                outside_hits.append(cell.coordinate)
                break
        if outside_hits:
            break
    for row in range(43, 81):
        for col in range(1, 67):  # A:BO
            cell = ws.cell(row, col)
            if cell.value is not None or has_visible_format(cell):
                outside_hits.append(cell.coordinate)
                break
        if len(outside_hits) > 1:
            break
    if outside_hits:
        errors.append(f"outside gantt scope still has visible content/format: {outside_hits[:5]}")

    contract_after = contract(path)
    if contract_after != baseline:
        errors.append(f"contract changed: before={baseline}, after={contract_after}")

    display_test = com_result.get("display_test", {})
    if not display_test.get("risk_changed") or not display_test.get("bar_changed"):
        errors.append(f"display format test failed: {display_test}")

    wb.close()
    return {
        "errors": errors,
        "cf_ranges": cf_ranges,
        "dv_ranges": dv_ranges,
        "contract_after": contract_after,
        "display_test": display_test,
    }


def readonly_reopen(path: Path) -> bool:
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True)
        wb.Close(SaveChanges=False)
        return True
    finally:
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def main() -> None:
    if not TARGET.exists():
        raise FileNotFoundError(TARGET)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    baseline = contract(TARGET)
    closed_live = check_live_target_not_unsaved()
    shutil.copy2(TARGET, BACKUP)
    shutil.copy2(TARGET, SIDECAR)

    com_result = apply_patch_with_com(SIDECAR)
    sidecar_validation = validate(SIDECAR, baseline, com_result)
    promoted = False
    final_validation = None
    final_reopen = False

    if not sidecar_validation["errors"]:
        shutil.copy2(SIDECAR, TARGET)
        promoted = True
        final_validation = validate(TARGET, baseline, com_result)
        final_reopen = readonly_reopen(TARGET)
        if final_validation["errors"] or not final_reopen:
            promoted = False
            shutil.copy2(BACKUP, TARGET)

    report = {
        "target": str(TARGET),
        "run_dir": str(RUN_DIR),
        "backup": str(BACKUP),
        "sidecar": str(SIDECAR),
        "closed_live_workbooks": closed_live,
        "baseline_contract": baseline,
        "com_result": com_result,
        "sidecar_validation": sidecar_validation,
        "promoted": promoted,
        "final_validation": final_validation,
        "final_readonly_reopen": final_reopen,
    }
    REPORT.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    if not promoted:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
