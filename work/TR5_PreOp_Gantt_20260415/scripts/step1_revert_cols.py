"""
step1_revert_cols_openpyxl.py  — Excel 없이 실행
openpyxl로 K/L 컬럼(Apr25/Apr26) 삭제 후 저장
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSM = r'C:\tr_dash-main\work\TR5_PreOp_Gantt_20260415\excel\TR5_PreOp_Gantt_20260415_162140.xlsm'
HEADER_ROW = 3; DATE_COL_START = 11
APR25=46137; APR26=46138; APR27=46139

print("Loading workbook (no Excel needed)...")
wb = load_workbook(XLSM, keep_vba=True)
ws = wb['Gantt_BASE']

k3 = ws.cell(HEADER_ROW, DATE_COL_START).value
l3 = ws.cell(HEADER_ROW, DATE_COL_START+1).value
m3 = ws.cell(HEADER_ROW, DATE_COL_START+2).value
print(f"K3={k3!r}  L3={l3!r}  M3={m3!r}")

def is_apr25(v):
    if isinstance(v, int) and v == APR25: return True
    try:
        if hasattr(v,'year') and v.year==2026 and v.month==4 and v.day==25: return True
    except: pass
    return False

def is_apr26(v):
    if isinstance(v, int) and v == APR26: return True
    try:
        if hasattr(v,'year') and v.year==2026 and v.month==4 and v.day==26: return True
    except: pass
    return False

if is_apr25(k3):
    print("Detected Apr25 at K3 — deleting 2 inserted columns...")
    ws.delete_cols(DATE_COL_START, 2)
    k3_after = ws.cell(HEADER_ROW, DATE_COL_START).value
    print(f"After: K3={k3_after!r}")
    wb.save(XLSM)
    print("Saved. Column revert complete.")
elif is_apr26(k3):
    print("Detected Apr26 at K3 — deleting 2 inserted columns...")
    ws.delete_cols(DATE_COL_START, 2)
    wb.save(XLSM)
    print("Saved.")
else:
    print(f"K3 is not Apr25/Apr26 ({k3!r}) — no revert needed.")
    wb.close() if hasattr(wb,'close') else None
print("Step 1 done.")
