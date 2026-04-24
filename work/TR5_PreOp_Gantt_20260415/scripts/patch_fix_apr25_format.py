"""
patch_fix_apr25_format.py
K3/L3 values stored as datetime by previous script.
Fix: convert to integer date serials matching existing cells, apply MM-DD number format.
Apr 27 = 46139 → Apr 25 = 46137, Apr 26 = 46138
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

XLSM = r'C:\tr_dash-main\work\TR5_PreOp_Gantt_20260415\excel\TR5_PreOp_Gantt_20260415_162140.xlsm'
HEADER_ROW = 3
DATE_COL_START = 11

print("Loading workbook...")
wb = load_workbook(XLSM, keep_vba=True)
ws = wb['Gantt_BASE']

# Inspect M3 to get the serial for Apr 27 and derive Apr25/Apr26
m3_val = ws.cell(HEADER_ROW, DATE_COL_START + 2).value   # M3 = Apr 27 (shifted right by 2)
print(f"M3 (Apr27 reference): {m3_val!r}")

# Apr 27 serial = m3_val; Apr25 = m3_val - 2; Apr26 = m3_val - 1
if isinstance(m3_val, int):
    apr25_serial = m3_val - 2
    apr26_serial = m3_val - 1
else:
    # Fallback: hardcoded serials
    apr25_serial = 46137
    apr26_serial = 46138
print(f"Apr25 serial={apr25_serial}, Apr26 serial={apr26_serial}")

# Read format from M3 to match style
ref_cell = ws.cell(HEADER_ROW, DATE_COL_START + 2)
ref_fmt = ref_cell.number_format

# Determine the correct number format for date display
# If ref uses "General", explicitly set "MM-DD" on new cells
if ref_fmt in ("General", "@", "", None):
    use_fmt = "MM-DD"  # explicit date format for the header
else:
    use_fmt = ref_fmt
print(f"Using number_format: {use_fmt!r}")

# Patch K3 and L3
for i, serial in enumerate([apr25_serial, apr26_serial]):
    cell = ws.cell(HEADER_ROW, DATE_COL_START + i)
    cell.value = serial
    cell.number_format = use_fmt
    # Also set column K/L format to match date columns
    col_letter = get_column_letter(DATE_COL_START + i)
    # Copy column width from M
    # openpyxl: ws.column_dimensions access
    ws.column_dimensions[col_letter].width = ws.column_dimensions[get_column_letter(DATE_COL_START + 2)].width or 5.14
    print(f"  {col_letter}3 = serial {serial}  fmt={use_fmt}")

print("\nVerification (cols 11..17):")
for c in range(DATE_COL_START, DATE_COL_START + 7):
    cell = ws.cell(HEADER_ROW, c)
    ltr = get_column_letter(c)
    print(f"  {ltr}3 = {cell.value!r}  fmt={cell.number_format}")

wb.save(XLSM)
print("\nSaved. Done.")
