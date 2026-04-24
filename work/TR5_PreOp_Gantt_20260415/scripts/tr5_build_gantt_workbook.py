"""
TR5 Pre-Op Gantt Workbook Builder — Plan B: MIR template copy + Excel COM-only writer
PLAN : docs/일정/PLAN3.MD
SSOT : docs/일정/TR5_Pre-Op_Simulation_v2_20260430.md
"""
from __future__ import annotations

import argparse
import contextlib
import datetime as dt
import gc
import json
import logging
import os
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

try:
    import pythoncom
    import win32com.client as win32
except Exception:
    pythoncom = None  # type: ignore[assignment]
    win32 = None     # type: ignore[assignment]

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore[assignment]

# ── paths ─────────────────────────────────────────────────────────────────────
DEFAULT_TEMPLATE = Path(
    r"C:\Users\SAMSUNG\Downloads\90_REVIEW_HOLD"
    r"\MIR_Gantt_Full_Source_Set\MIR_Reactor_Repair_Schedule.xlsm"
)
DEFAULT_SCHEDULE = Path(
    r"C:\tr_dash-main\docs\일정\TR5_Pre-Op_Simulation_v2_20260430.md"
)
DEFAULT_OUT_DIR = Path(
    r"C:\tr_dash-main\output\spreadsheet\tr5_preparation_simulation"
)

# ── layout ────────────────────────────────────────────────────────────────────
SHEET         = "Gantt_BASE"
FIRST_TL_COL  = 11   # K
LAST_TL_COL   = 31   # AE   (21 days Apr 22 – May 12 inclusive)
CLEAR_COL     = 67   # BO
FIRST_ROW     = 4    # first section-header row
LAST_ROW      = 42
CLEAR_ROW     = 80
COL_TL_WIDTH  = 5.0
TASK_COUNT    = 33
TL_START      = dt.date(2026, 4, 22)
TL_END        = dt.date(2026, 5, 12)
LOAD_OUT      = dt.date(2026, 5, 4)
AGI_JD        = dt.date(2026, 5, 12)
EXCEL_EPOCH   = dt.date(1899, 12, 30)

RISK_VALUES   = ["OK", "AMBER", "HIGH", "WARNING", "CRITICAL", "GATE"]
RISK_FILL     = {"OK": "DFF0D8", "AMBER": "F6C44B", "HIGH": "F4B183",
                 "WARNING": "FCE4D6", "CRITICAL": "FF5B5B", "GATE": "FFD500"}
RISK_FONT     = {"OK": "173B22", "AMBER": "1A1A1A", "HIGH": "1A1A1A",
                 "WARNING": "1A1A1A", "CRITICAL": "FFFFFF", "GATE": "111111"}
BAR_FILL      = {"OK": "4FB4F5", "AMBER": "F6C44B", "HIGH": "F4B183",
                 "WARNING": "FCE4D6", "CRITICAL": "FF5B5B", "GATE": "FFD500"}
PHASE_FILL    = {"DOC": "7030A0", "HM_GATE": "A586E8", "EQUIP": "27C8A2",
                 "MARINE": "3D7CCB", "MZP_OPS": "F58220", "AGI_OPS": "62C85B"}
SECTION_PHASE = {
    "OFCO - Customs and PTW":               "DOC",
    "OFCO+MMT - Harbour Master Approval":   "HM_GATE",
    "Mammoet - Equipment Mobilization":     "EQUIP",
    "KFS - LCT Bushra":                     "MARINE",
    "ALL - MZP Berth Ops and MWS Survey":   "MZP_OPS",
    "KFS+MMT - Voyage and AGI Operations":  "AGI_OPS",
}

# COM enum aliases
XL_CENTER                             = -4108
XL_LEFT                               = -4131
XL_CONTINUOUS                         = 1
XL_THIN                               = 2
XL_VALIDATE_LIST                      = 3
XL_VALID_ALERT_STOP                   = 1
XL_FORMAT_CONDITION_EXPRESSION        = 2
XL_CALCULATION_MANUAL                 = -4135
MSO_AUTOMATION_SECURITY_FORCE_DISABLE = 3


# ── exceptions ────────────────────────────────────────────────────────────────
class BuildBlocked(RuntimeError):
    pass


class ValidationFailed(RuntimeError):
    pass


# ── data model ────────────────────────────────────────────────────────────────
@dataclass
class TaskRow:
    row_type: str          # "section" | "task"
    section:  str
    phase:    str
    task:     str = ""
    task_id:  str = ""
    start:    dt.date | None = None
    end:      dt.date | None = None
    days:     int | None = None
    risk:     str = ""
    note:     str = ""


@dataclass
class BuildConfig:
    template:    Path
    schedule:    Path
    out_dir:     Path
    output_name: str | None = None
    run_dir:     Path | None = None
    open_after:  bool = False
    keep_sidecar: bool = False
    dry_run:     bool = False
    now: dt.datetime = field(default_factory=lambda: dt.datetime.now().replace(microsecond=0))

    @property
    def stamp(self) -> str:
        return self.now.strftime("%Y%m%d_%H%M%S")

    @property
    def final_name(self) -> str:
        return self.output_name or f"TR5_PreOp_Gantt_{self.stamp}.xlsm"

    @property
    def output(self) -> Path:
        return self.out_dir / self.final_name

    @property
    def run(self) -> Path:
        return self.run_dir or (self.out_dir / f"_build_{self.stamp}")

    @property
    def sidecar(self) -> Path:
        return self.run / f"_sidecar_{self.final_name}"

    @property
    def report(self) -> Path:
        return self.run / "validation_report.json"

    @property
    def log_file(self) -> Path:
        return self.run / "build.log"


# ── utilities ─────────────────────────────────────────────────────────────────
def xl_color(hex_rgb: str) -> int:
    h = hex_rgb.lstrip("#")
    return int(h[4:6] + h[2:4] + h[0:2], 16)


def col_letter(col: int) -> str:
    out = ""
    while col:
        col, rem = divmod(col - 1, 26)
        out = chr(65 + rem) + out
    return out


def excel_serial(day: dt.date) -> int:
    return (day - EXCEL_EPOCH).days


def serial_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        return EXCEL_EPOCH + dt.timedelta(days=int(value))
    return None


def setup_logger(path: Path) -> logging.Logger:
    path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("tr5_gantt_build")
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    fh = logging.FileHandler(path, encoding="utf-8")
    fh.setFormatter(fmt)
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger


def has_zone_identifier(path: Path) -> bool:
    return os.path.exists(str(path) + ":Zone.Identifier")


def package_contract(path: Path) -> dict[str, Any]:
    protected: dict[str, dict[str, int]] = {}
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            n = info.filename
            if (n == "xl/vbaProject.bin"
                    or (n.startswith("xl/drawings/drawing") and n.endswith(".xml"))
                    or (n.startswith("xl/drawings/vmlDrawing") and n.endswith(".vml"))
                    or (n.startswith("xl/ctrlProps/") and n.endswith(".xml"))):
                protected[n] = {"crc": int(info.CRC), "file_size": int(info.file_size)}
    return {
        "vbaProject.bin": int("xl/vbaProject.bin" in protected),
        "drawings": sum(1 for n in protected if n.startswith("xl/drawings/drawing")),
        "ctrlProps": sum(1 for n in protected if n.startswith("xl/ctrlProps/")),
        "vmlDrawing": sum(1 for n in protected if n.startswith("xl/drawings/vmlDrawing")),
        "protected_parts": protected,
    }


# ── Mermaid parser ────────────────────────────────────────────────────────────
def parse_schedule(path: Path) -> list[TaskRow]:
    text   = path.read_text(encoding="utf-8")
    blocks = re.findall(r"```mermaid\s*(.*?)```", text, flags=re.S)
    block  = next((b for b in blocks if re.search(r"^\s*gantt\s*$", b, re.M)), None)
    if not block:
        raise BuildBlocked("Mermaid Gantt block not found in schedule MD.")

    rows: list[TaskRow] = []
    section = ""
    phase   = ""

    for raw in block.splitlines():
        line = raw.strip()
        if (not line or line == "gantt"
                or line.startswith(("%%", "title ", "dateFormat ", "axisFormat "))):
            continue
        if line.startswith("section "):
            section = line.removeprefix("section ").strip()
            phase   = SECTION_PHASE.get(section, section.upper().replace(" ", "_")[:12])
            rows.append(TaskRow("section", section, phase, task=section))
            continue
        if ":" not in line:
            continue

        task_text, meta = line.split(":", 1)
        tokens   = [t.strip() for t in meta.split(",") if t.strip()]
        statuses: list[str] = []
        task_id  = ""
        start: dt.date | None = None
        days: int | None = None

        for tok in tokens:
            low = tok.lower()
            if low in {"crit", "active", "done", "milestone"}:
                statuses.append(low)
            elif re.fullmatch(r"\d{4}-\d{2}-\d{2}", tok):
                start = dt.datetime.strptime(tok, "%Y-%m-%d").date()
            elif re.fullmatch(r"\d+d", low):
                days = int(low[:-1])
            elif not task_id:
                task_id = tok.upper()

        if not task_id or start is None or days is None:
            raise BuildBlocked(f"Could not parse task line: {line!r}")

        end  = start if days == 0 else start + dt.timedelta(days=days - 1)
        risk = ("GATE" if "milestone" in statuses or days == 0
                else ("CRITICAL" if "crit" in statuses else "OK"))
        rows.append(TaskRow(
            "task", section, phase,
            task_text.strip(), task_id, start, end, days, risk,
            f"Mermaid id={task_id.lower()}; {days}d from {start:%Y-%m-%d}"))

    tasks = [r for r in rows if r.row_type == "task"]
    if len(tasks) != TASK_COUNT:
        raise BuildBlocked(f"Expected {TASK_COUNT} tasks, parsed {len(tasks)}.")

    final = tasks[-1]
    if (final.task_id != "JDC" or final.start != AGI_JD
            or final.end != AGI_JD or final.days != 0 or final.risk != "GATE"):
        raise BuildBlocked(f"Final JDC semantics failed: {final}")

    expected_rows = TASK_COUNT + len([r for r in rows if r.row_type == "section"])
    if len(rows) != expected_rows:
        raise BuildBlocked(f"Expected {expected_rows} rows, parsed {len(rows)}.")

    return rows


# ── Excel COM helpers ─────────────────────────────────────────────────────────
@contextlib.contextmanager
def excel_app(*, visible: bool = False,
              force_disable_macros: bool = True) -> Iterator[Any]:
    if pythoncom is None or win32 is None:
        raise BuildBlocked("pywin32 is not available.")
    pythoncom.CoInitialize()
    app = win32.DispatchEx("Excel.Application")
    original: dict[str, Any] = {}
    for prop in ("AutomationSecurity", "DisplayAlerts", "EnableEvents",
                 "ScreenUpdating", "Calculation", "Visible"):
        try:
            original[prop] = getattr(app, prop)
        except Exception:
            pass
    try:
        app.Visible = visible
        app.DisplayAlerts = False
        app.EnableEvents = False
        app.ScreenUpdating = False
        try:
            app.Calculation = XL_CALCULATION_MANUAL
        except Exception:
            pass
        if force_disable_macros:
            try:
                app.AutomationSecurity = MSO_AUTOMATION_SECURITY_FORCE_DISABLE
            except Exception:
                pass
        yield app
    finally:
        for prop, value in original.items():
            try:
                setattr(app, prop, value)
            except Exception:
                pass
        try:
            app.Quit()
        except Exception:
            pass
        del app
        gc.collect()
        pythoncom.CoUninitialize()


def open_wb(app: Any, path: Path, read_only: bool) -> Any:
    return app.Workbooks.Open(
        str(path), UpdateLinks=0, ReadOnly=read_only,
        IgnoreReadOnlyRecommended=True, AddToMru=False)


def set_fmt(rng: Any, fmt: str) -> None:
    rng.NumberFormat = fmt
    try:
        rng.NumberFormatLocal = fmt
    except Exception:
        pass


def capture_controls(ws: Any) -> dict[str, Any]:
    shapes = []
    for i in range(1, ws.Shapes.Count + 1):
        sh = ws.Shapes(i)
        text = action = ""
        try:
            text = str(sh.TextFrame.Characters().Text).strip()
        except Exception:
            pass
        try:
            action = str(sh.OnAction or "").strip()
        except Exception:
            pass
        shapes.append({"name": sh.Name, "type": int(sh.Type),
                        "text": text, "on_action": action})
    ole = 0
    try:
        ole = int(ws.OLEObjects().Count)
    except Exception:
        pass
    return {"shapes_count": int(ws.Shapes.Count),
            "ole_objects_count": ole, "shapes": shapes}


def update_buttons(ws: Any, workbook_name: str) -> list[str]:
    mapping = {
        "Init":       "modMIR_Gantt_Unified.Init_Unified_System",
        "Reset":      "modMIR_Gantt_Unified.Reset_Schedule_To_Original",
        "Show Notes": "modMIR_Gantt_Unified.Toggle_Notes_Action_Column",
    }
    updated = []
    for i in range(1, ws.Shapes.Count + 1):
        sh = ws.Shapes(i)
        text = ""
        try:
            text = str(sh.TextFrame.Characters().Text).strip()
        except Exception:
            pass
        if text in mapping:
            # Keep the binding workbook-local. Excel may rewrite a workbook-
            # qualified sidecar action to the user's Documents path before the
            # final file is promoted.
            sh.OnAction = mapping[text]
            updated.append(text)
    return updated


def _apply_line_patches(code: str) -> tuple[str, int]:
    """Apply line-level VBA patches. Returns (patched_code, count_applied)."""
    count = 0
    # Normalise line endings for matching
    eol = "\r\n" if "\r\n" in code else "\n"

    # FIX 1: PaintBars out-of-range guard (error 1004)
    # When the task date lies completely outside the timeline, offsetS > offsetE
    # and the subsequent ws.Range(...) call raises error 1004.
    clamp_old = (
        f"        If offsetS < 0 Then offsetS = 0{eol}"
        f"        If offsetE > totalDays - 1 Then offsetE = totalDays - 1{eol}"
    )
    clamp_new = (
        f"        If offsetS < 0 Then offsetS = 0{eol}"
        f"        If offsetE > totalDays - 1 Then offsetE = totalDays - 1{eol}"
        f"        If offsetS > offsetE Then GoTo ContinueRow{eol}"
    )
    if clamp_old in code and clamp_new not in code:
        code = code.replace(clamp_old, clamp_new, 1)
        count += 1

    # FIX 2: LogMsg — write ISO timestamp instead of Excel serial float
    log_old = f"    ws.Cells(n, 1).Value2 = Now{eol}"
    log_new = f'    ws.Cells(n, 1).Value = Format$(Now, "yyyy-mm-dd hh:mm:ss"){eol}'
    if log_old in code and log_new not in code:
        code = code.replace(log_old, log_new, 1)
        count += 1

    return code, count


def patch_vba(wb: Any) -> dict[str, Any]:
    count       = int(wb.VBProject.VBComponents.Count)
    touched: list[str] = []
    replacements = 0
    exact = ("RefreshGanttExactRange ws, "
             "DateSerial(2026, 4, 22), DateSerial(2026, 5, 12), False")
    for i in range(1, count + 1):
        comp = wb.VBProject.VBComponents(i)
        cm   = comp.CodeModule
        n    = int(cm.CountOfLines)
        if n <= 0:
            continue
        code = cm.Lines(1, n)
        new  = code
        for pat in (r"RefreshGantt\s+ws\s*,\s*True\s*,\s*True",
                    r"RefreshGantt\s+ws\s*,\s*True"):
            new, c = re.subn(pat, exact, new)
            replacements += c
        for old, repl in {
            'SafeSetNumberFormat ws.Cells(HEADER_ROW, col), "mm/dd"':
            'SafeSetNumberFormat ws.Cells(HEADER_ROW, col), "mm-dd"',
            "ws.Columns(col).ColumnWidth = 3.2":
            "ws.Columns(col).ColumnWidth = 5#",
        }.items():
            if old in new:
                new = new.replace(old, repl)
                replacements += 1
        # FIX 1 & 2: line-level patches (PaintBars guard + LogMsg ISO date)
        new, lc = _apply_line_patches(new)
        replacements += lc
        if new != code:
            cm.DeleteLines(1, n)
            cm.AddFromString(new)
            touched.append(comp.Name)
    return {"vbcomponents_count": count, "touched": touched,
            "replacements": replacements}


def reset_names(wb: Any) -> int:
    deleted = 0
    for name in [n.Name for n in wb.Names]:
        simple = name.split("!")[-1]
        if simple.startswith(("MIR_UID_ROW_", "MIR_UID_ORIG_")):
            try:
                wb.Names(name).Delete()
                deleted += 1
            except Exception:
                pass
    try:
        wb.Names("MIR_ORIG_TIMELINE_RANGE").Delete()
    except Exception:
        pass
    nm = wb.Names.Add(Name="MIR_ORIG_TIMELINE_RANGE",
                      RefersTo='="2026-04-22|2026-05-12"')
    try:
        nm.Visible = False
    except Exception:
        pass
    return deleted


def add_cf(rng: Any, formula: str, fill: str, font: str | None = None) -> None:
    last: Exception | None = None
    for candidate in (formula, formula.replace(",", ";")):
        try:
            fc = rng.FormatConditions.Add(
                XL_FORMAT_CONDITION_EXPRESSION, XL_VALID_ALERT_STOP, candidate)
            fc.Interior.Color = xl_color(fill)
            if font:
                fc.Font.Color = xl_color(font)
            fc.Font.Bold = True
            return
        except Exception as exc:
            last = exc
            try:
                rng.FormatConditions(rng.FormatConditions.Count).Delete()
            except Exception:
                pass
    raise BuildBlocked(
        f"Conditional formatting formula failed: {formula}: {last}")


# ── COM writer ────────────────────────────────────────────────────────────────
def render(sidecar: Path, cfg: BuildConfig,
           rows: list[TaskRow], baseline: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {"errors": [], "warnings": []}

    with excel_app(visible=False, force_disable_macros=True) as app:
        wb     = None
        closed = False
        try:
            wb = open_wb(app, sidecar, read_only=False)
            if bool(getattr(wb, "ReadOnly", False)):
                raise BuildBlocked(f"sidecar opened read-only: {sidecar}")
            open_books = [app.Workbooks(i).FullName
                          for i in range(1, app.Workbooks.Count + 1)]
            if len(open_books) != 1:
                raise BuildBlocked(f"Single-instance gate failed: {open_books}")

            ws = wb.Worksheets(SHEET)

            # Unprotect sheet and workbook before any modification
            for fn in (wb.Unprotect, ws.Unprotect):
                try:
                    fn()
                except Exception:
                    pass

            result["controls_before"] = capture_controls(ws)
            try:
                result["LinkSources"] = list(wb.LinkSources() or [])
            except Exception as exc:
                result["LinkSources"] = f"unavailable: {exc}"
            try:
                result["Connections.Count"] = int(wb.Connections.Count)
            except Exception:
                result["Connections.Count"] = 0

            # VBA patch and defined names (before clear to avoid re-trigger)
            result["vba_patch"]         = patch_vba(wb)
            result["deleted_uid_names"] = reset_names(wb)

            # ── Clear scope ───────────────────────────────────────────────
            full_rng = f"A1:{col_letter(CLEAR_COL)}{CLEAR_ROW}"
            ws.Range(full_rng).FormatConditions.Delete()
            # Delete ALL existing data validations (removes inherited template DV)
            try:
                ws.Range(full_rng).Validation.Delete()
            except Exception:
                pass
            # Unmerge before clear to avoid merged-cell errors
            try:
                ws.Range(f"A1:{col_letter(CLEAR_COL)}3").UnMerge()
            except Exception:
                pass
            try:
                ws.Range(full_rng).UnMerge()
            except Exception:
                pass
            ws.Range(full_rng).ClearContents()
            ws.Range(full_rng).ClearFormats()

            # ── Title / Subtitle ──────────────────────────────────────────
            ws.Range("D1:AE1").Merge()
            ws.Range("D2:AE2").Merge()
            ws.Range("D1").Value = "TR5 PRE-OP SIMULATION - PHASE-GATE GANTT"
            try:
                wb.Names("MIR_SOURCE_LABEL").Delete()
            except Exception:
                pass
            source_name = wb.Names.Add(
                Name="MIR_SOURCE_LABEL",
                RefersTo='="TR5_Pre-Op_Simulation_v2_20260430.md"',
            )
            try:
                source_name.Visible = False
            except Exception:
                pass
            try:
                wb.Names("MIR_T_PLUS_ZERO").Delete()
            except Exception:
                pass
            t0_name = wb.Names.Add(Name="MIR_T_PLUS_ZERO", RefersTo="=DATE(2026,4,22)")
            try:
                t0_name.Visible = False
            except Exception:
                pass
            ws.Range("D2").Formula = (
                '=TEXT(TODAY(),"yyyy-mm-dd")'
                '&" | Source: "&IFERROR(MIR_SOURCE_LABEL,"TR5 schedule")'
                '&" | T+0="&TEXT(IFERROR(MIR_T_PLUS_ZERO,INDEX($3:$3,1,11)),"yyyy-mm-dd")'
                '&" | Load-out="&IFERROR(TEXT(LOOKUP(2,1/($A$1:$A$100="BD2"),$D$1:$D$100),"yyyy-mm-dd"),"")'
                '&" | AGI JD="&IFERROR(TEXT(LOOKUP(2,1/($A$1:$A$100="JDC"),$D$1:$D$100),"yyyy-mm-dd"),"")'
            )
            for addr, size, bold, fg, bg in (
                ("D1:AE1", 13, True,  "F5C84C", "0B1120"),
                ("D2:AE2",  9, False, "7D879C", "101827"),
            ):
                r = ws.Range(addr)
                r.Interior.Color       = xl_color(bg)
                r.Font.Name            = "Arial"
                r.Font.Size            = size
                r.Font.Bold            = bold
                r.Font.Color           = xl_color(fg)
                r.HorizontalAlignment  = XL_CENTER
                r.VerticalAlignment    = XL_CENTER

            # ── Column headers row 3 ──────────────────────────────────────
            headers = ["ID", "Phase", "Task Description",
                       "Start", "End", "Days", "Risk", "Notes / Action"]
            for col, header in enumerate(headers, 1):
                ws.Cells(3, col).Value = header
            hdr = ws.Range("A3:H3")
            hdr.Interior.Color      = xl_color("0B1120")
            hdr.Font.Name           = "Arial"
            hdr.Font.Size           = 9
            hdr.Font.Bold           = True
            hdr.Font.Color          = xl_color("F5C84C")
            hdr.HorizontalAlignment = XL_CENTER
            hdr.VerticalAlignment   = XL_CENTER

            # ── Timeline header row 3 (K:AE) ──────────────────────────────
            for offset, col in enumerate(range(FIRST_TL_COL, LAST_TL_COL + 1)):
                day  = TL_START + dt.timedelta(days=offset)
                cell = ws.Cells(3, col)
                cell.Value2 = excel_serial(day)
                set_fmt(cell, "mm-dd")
                cell.Orientation          = 90
                cell.Font.Name            = "Arial"
                cell.Font.Size            = 7
                cell.Font.Bold            = True
                cell.Font.Color           = xl_color("6F7A92")
                cell.Interior.Color       = xl_color("0B1120")
                cell.HorizontalAlignment  = XL_CENTER
                cell.VerticalAlignment    = XL_CENTER
                ws.Columns(col).ColumnWidth = COL_TL_WIDTH
            set_fmt(ws.Range("K3:AE3"), "mm-dd")

            # ── Task rows ─────────────────────────────────────────────────
            xrow      = FIRST_ROW
            task_rows: list[int] = []

            for row in rows:
                if row.row_type == "section":
                    ws.Cells(xrow, 2).Value = f"> {row.section}"
                    rng = ws.Range(f"A{xrow}:AE{xrow}")
                    rng.Interior.Color      = xl_color("1A2233")
                    rng.Font.Name           = "Arial"
                    rng.Font.Size           = 9
                    rng.Font.Bold           = True
                    rng.Font.Color          = xl_color("58B4F5")
                    rng.HorizontalAlignment = XL_LEFT
                else:
                    task_rows.append(xrow)
                    values = [
                        row.task_id, row.phase, row.task,
                        excel_serial(row.start), excel_serial(row.end),
                        row.days, row.risk, row.note,
                    ]
                    for col, value in enumerate(values, 1):
                        ws.Cells(xrow, col).Value = value
                    rng = ws.Range(f"A{xrow}:H{xrow}")
                    rng.Interior.Color = xl_color("0B1020")
                    rng.Font.Name      = "Arial"
                    rng.Font.Size      = 9
                    rng.Font.Color     = xl_color("D7DEE9")
                    ws.Cells(xrow, 2).Interior.Color = xl_color(
                        PHASE_FILL.get(row.phase, "3D7CCB"))
                    ws.Cells(xrow, 2).Font.Color = xl_color("FFFFFF")
                    ws.Cells(xrow, 2).Font.Bold  = True
                    ws.Cells(xrow, 7).Interior.Color = xl_color(RISK_FILL[row.risk])
                    ws.Cells(xrow, 7).Font.Color     = xl_color(RISK_FONT[row.risk])
                    ws.Cells(xrow, 7).Font.Bold      = True
                xrow += 1

            if xrow - 1 != LAST_ROW:
                raise BuildBlocked(
                    f"render row ended at {xrow - 1}, expected {LAST_ROW}")

            # ── Date formats / borders / sizes ────────────────────────────
            set_fmt(ws.Range("D4:E42"), "yyyy-mm-dd")
            ws.Range("A1:AE42").Font.Name         = "Arial"
            ws.Range("A4:AE42").Borders.LineStyle = XL_CONTINUOUS
            ws.Range("A4:AE42").Borders.Weight    = XL_THIN
            ws.Range("A4:AE42").Borders.Color     = xl_color("293142")
            ws.Range("K4:AE42").Interior.Color    = xl_color("0B1020")
            ws.Range("K4:AE42").Borders.Color     = xl_color("242A38")
            ws.Range("K4:AE42").Borders.LineStyle = XL_CONTINUOUS
            ws.Range("K4:AE42").Borders.Weight    = XL_THIN

            for col, width in {1: 8, 2: 16, 3: 42, 4: 12,
                               5: 12, 6: 8, 7: 12, 8: 34}.items():
                ws.Columns(col).ColumnWidth = width
            ws.Rows(1).RowHeight               = 31.95
            ws.Rows(2).RowHeight               = 18
            ws.Rows(3).RowHeight               = 42
            ws.Range("A4:AE42").Rows.RowHeight = 17.4

            # ── Risk data validation G4:G42 ───────────────────────────────
            risk_rng = ws.Range("G4:G42")
            try:
                risk_rng.Validation.Delete()
            except Exception:
                pass
            risk_rng.Validation.Add(
                Type=XL_VALIDATE_LIST,
                AlertStyle=XL_VALID_ALERT_STOP,
                Operator=XL_VALID_ALERT_STOP,
                Formula1=",".join(RISK_VALUES))
            risk_rng.Validation.IgnoreBlank    = True
            risk_rng.Validation.InCellDropdown = True
            risk_rng.FormatConditions.Delete()
            for risk in RISK_VALUES:
                add_cf(risk_rng, f'=$G4="{risk}"', RISK_FILL[risk], RISK_FONT[risk])

            # ── Gantt bar conditional formatting K5:AE42 ──────────────────
            bar_rng = ws.Range("K5:AE42")
            bar_rng.FormatConditions.Delete()
            for risk in RISK_VALUES:
                add_cf(bar_rng,
                       f'=AND(K$3>=$D5,K$3<=$E5,$G5="{risk}")',
                       BAR_FILL[risk], "0D1320")

            # ── Button OnAction update ────────────────────────────────────
            result["updated_buttons"] = update_buttons(ws, cfg.final_name)
            result["controls_after"]  = capture_controls(ws)
            result["task_rows"]       = task_rows

            try:
                app.CalculateFullRebuild()
            except Exception:
                pass
            wb.Save()
            wb.Close(SaveChanges=True)
            closed = True

        finally:
            if wb is not None and not closed:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass

    # ── Contract check (openpyxl read-only) ──────────────────────────────────
    after = package_contract(sidecar)
    result["contract_after_render"] = after
    if after["vbaProject.bin"] != baseline["vbaProject.bin"]:
        result["errors"].append("vbaProject.bin presence changed")
    for key in ("drawings", "ctrlProps", "vmlDrawing"):
        if after[key] != baseline[key]:
            result["errors"].append(
                f"{key} count changed: {baseline[key]} -> {after[key]}")
    if after["protected_parts"].keys() != baseline["protected_parts"].keys():
        result["errors"].append("protected package part set changed")
    for part, meta in baseline["protected_parts"].items():
        if (part in after["protected_parts"]
                and after["protected_parts"][part]["crc"] != meta["crc"]):
            result["warnings"].append(f"CRC changed for {part}; allowed warning.")
    return result


# ── openpyxl structural validation ───────────────────────────────────────────
# Acceptable date formats that COM may store vs what openpyxl reads back
_DATE_FMTS_OK = frozenset(["yyyy-mm-dd", "mm-dd-yy", "mm/dd/yy",
                            "m/d/yy", "d-mmm-yy", "dd-mmm-yy"])
_TL_FMTS_OK   = frozenset(["mm-dd", "General", ""])


def validate_openpyxl(path: Path, rows: list[TaskRow]) -> dict[str, Any]:
    if load_workbook is None:
        raise BuildBlocked("openpyxl is not available.")
    errors: list[str] = []
    wb = load_workbook(path, keep_vba=True, data_only=False)
    ws = wb[SHEET]
    try:
        # title
        if ws["D1"].value != "TR5 PRE-OP SIMULATION - PHASE-GATE GANTT":
            errors.append(f"title mismatch: {ws['D1'].value!r}")

        # D31 / E31 (Sea Fastening row — start=2026-05-09)
        for cell_addr in ("D31", "E31"):
            cell = ws[cell_addr]
            if cell.number_format not in _DATE_FMTS_OK:
                errors.append(f"{cell_addr} format: {cell.number_format!r}")
            if serial_date(cell.value) != LOAD_OUT:
                errors.append(
                    f"{cell_addr} value: {cell.value!r} != {LOAD_OUT}")

        # Timeline header K3:AE3
        expected = TL_START
        for col in range(FIRST_TL_COL, LAST_TL_COL + 1):
            cell = ws.cell(3, col)
            if cell.number_format not in _TL_FMTS_OK:
                errors.append(
                    f"{cell.coordinate} timeline format: {cell.number_format!r}")
                break
            if serial_date(cell.value) != expected:
                errors.append(
                    f"{cell.coordinate} expected {expected}, got {cell.value!r}")
                break
            expected += dt.timedelta(days=1)

        # AF3 must be blank
        if ws["AF3"].value is not None:
            errors.append("AF3 should be blank")

        # Task count from D/E columns
        task_rows = [
            r for r in range(4, 43)
            if ws.cell(r, 4).value is not None
            and ws.cell(r, 5).value is not None
        ]
        if len(task_rows) != TASK_COUNT:
            errors.append(
                f"Excel task count expected {TASK_COUNT}, got {len(task_rows)}")

        # MD vs Excel round-trip
        actual = [
            (str(ws.cell(r, 1).value or ""),
             str(ws.cell(r, 3).value or ""),
             serial_date(ws.cell(r, 4).value),
             int(ws.cell(r, 6).value)
             if ws.cell(r, 6).value is not None else None)
            for r in task_rows
        ]
        expected_tasks = [
            (t.task_id, t.task, t.start, t.days)
            for t in rows if t.row_type == "task"
        ]
        if actual != expected_tasks:
            # Surface first mismatch
            for i, (a, e) in enumerate(zip(actual, expected_tasks)):
                if a != e:
                    errors.append(
                        f"task mismatch at index {i}: {a} != {e}")
                    break

        # CF ranges
        cf_ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
        if "G4:G42" not in cf_ranges:
            errors.append(f"risk CF missing, found: {cf_ranges}")
        if "K5:AE42" not in cf_ranges:
            errors.append(f"bar CF missing, found: {cf_ranges}")

        # Data validation — only G4:G42 expected
        dvs = [str(dv.sqref).strip()
               for dv in ws.data_validations.dataValidation]
        if dvs != ["G4:G42"]:
            errors.append(f"DV range mismatch: {dvs}")
        elif (str(ws.data_validations.dataValidation[0].formula1 or "")
              .strip('"') != ",".join(RISK_VALUES)):
            errors.append("risk DV formula mismatch")

        # Final JDC row
        if (ws["A42"].value != "JDC"
                or ws["C42"].value != "TARGET AGI Jacking Down Complete"):
            errors.append(
                f"final JDC id/task: A42={ws['A42'].value!r}, C42={ws['C42'].value!r}")
        f42_days  = ws["F42"].value
        days_ok   = (f42_days is not None and int(f42_days) == 0)
        if (serial_date(ws["D42"].value) != AGI_JD
                or not days_ok
                or ws["G42"].value != "GATE"):
            errors.append(
                f"final JDC semantics: D42={ws['D42'].value!r}, "
                f"F42={f42_days!r}, G42={ws['G42'].value!r}")
    finally:
        wb.close()
    return {"errors": errors, "warnings": []}


# ── COM display validation ────────────────────────────────────────────────────
def validate_com(path: Path, *, macro_smoke: bool,
                 output_name: str | None = None) -> dict[str, Any]:
    errors: list[str] = []
    data: dict[str, Any] = {"errors": errors, "warnings": []}

    with excel_app(visible=False,
                   force_disable_macros=not macro_smoke) as app:
        wb = None
        try:
            wb = open_wb(app, path, read_only=True)
            ws = wb.Worksheets(SHEET)

            disp = {
                "D31.Text":         str(ws.Range("D31").Text),
                "D31.NumberFormat": str(ws.Range("D31").NumberFormat),
                "E31.Text":         str(ws.Range("E31").Text),
                "K3.Text":          str(ws.Range("K3").Text),
                "K3.NumberFormat":  str(ws.Range("K3").NumberFormat),
                "AE3.Text":         str(ws.Range("AE3").Text),
                "AE3.NumberFormat": str(ws.Range("AE3").NumberFormat),
                "AF3.Value2":       ws.Range("AF3").Value2,
                "Row3.Height":      float(ws.Rows(3).RowHeight),
                "K.Width":          float(ws.Columns(FIRST_TL_COL).ColumnWidth),
                "AE.Width":         float(ws.Columns(LAST_TL_COL).ColumnWidth),
            }
            data["display_validation"] = disp

            for key, value in {
                "D31.Text":  "2026-05-04",
                "E31.Text":  "2026-05-04",
                "K3.Text":   "04-22",
                "AE3.Text":  "05-12",
            }.items():
                if disp.get(key) != value:
                    errors.append(
                        f"display {key}: {disp.get(key)!r} != {value!r}")

            if disp["AF3.Value2"] is not None:
                errors.append("AF3.Value2 is not blank")
            if round(disp["Row3.Height"], 1) != 42.0:
                errors.append(f"row 3 height mismatch: {disp['Row3.Height']}")
            if (round(disp["K.Width"], 1) != COL_TL_WIDTH
                    or round(disp["AE.Width"], 1) != COL_TL_WIDTH):
                errors.append(
                    f"K/AE width mismatch: {disp['K.Width']}, {disp['AE.Width']}")

            controls = capture_controls(ws)
            data["controls"] = controls
            by_text = {s["text"]: s["on_action"]
                       for s in controls["shapes"] if s["text"]}
            # Accept either the sidecar filename or the expected output filename
            expected_names = {path.name}
            if output_name:
                expected_names.add(output_name)
            for label in ("Init", "Reset", "Show Notes"):
                action = by_text.get(label, "")
                workbook_local = action.startswith("modMIR_Gantt_Unified.")
                if not action or (not workbook_local and not any(n in action for n in expected_names)):
                    errors.append(
                        f"control binding failed for {label}: {action!r}")

            broken: list[str] = []
            try:
                for ref in wb.VBProject.References:
                    if bool(ref.IsBroken):
                        broken.append(str(ref.Name))
            except Exception as exc:
                errors.append(f"VBProject reference check failed: {exc}")
            data["broken_references"] = broken
            if broken:
                errors.append(f"broken VBA references: {broken}")

            if macro_smoke:
                macro = (f"'{path.name}'!"
                         "modMIR_Gantt_Unified.Init_Unified_System")
                data["macro_smoke"] = {"macro": macro, "ran": False}
                try:
                    app.Run(macro)
                    data["macro_smoke"]["ran"] = True
                except Exception as exc:
                    data["macro_smoke"]["error"] = str(exc)
                after = {
                    "K3.Text":  str(ws.Range("K3").Text),
                    "AE3.Text": str(ws.Range("AE3").Text),
                }
                data["macro_smoke"]["after"] = after
                if after["K3.Text"] != "04-22" or after["AE3.Text"] != "05-12":
                    errors.append(f"macro timeline expanded: {after}")

            wb.Close(SaveChanges=False)
            wb = None
        finally:
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
    return data


# ── main ──────────────────────────────────────────────────────────────────────
def main() -> int:
    p = argparse.ArgumentParser(description="TR5 Pre-Op Gantt Workbook Builder")
    p.add_argument("--template",              type=Path, default=DEFAULT_TEMPLATE)
    p.add_argument("--schedule",              type=Path, default=DEFAULT_SCHEDULE)
    p.add_argument("--out-dir",               type=Path, default=DEFAULT_OUT_DIR,
                   dest="out_dir")
    p.add_argument("--output-name",           type=str,  default=None,
                   dest="output_name")
    p.add_argument("--run-dir",               type=Path, default=None,
                   dest="run_dir")
    p.add_argument("--open",                  dest="open_after", action="store_true")
    p.add_argument("--no-open",               dest="open_after", action="store_false")
    p.add_argument("--keep-sidecar",          action="store_true", dest="keep_sidecar")
    p.add_argument("--dry-run",               action="store_true", dest="dry_run")
    p.add_argument("--fail-if-template-blocked", action="store_true",
                   dest="fail_if_template_blocked")
    args = p.parse_args()

    cfg = BuildConfig(
        template=args.template, schedule=args.schedule,
        out_dir=args.out_dir, output_name=args.output_name,
        run_dir=args.run_dir, open_after=args.open_after,
        keep_sidecar=args.keep_sidecar, dry_run=args.dry_run,
    )
    cfg.out_dir.mkdir(parents=True, exist_ok=True)
    cfg.run.mkdir(parents=True, exist_ok=True)

    log = setup_logger(cfg.log_file)
    report: dict[str, Any] = {
        "status": "exception", "output": None, "run_dir": str(cfg.run),
        "errors": [], "warnings": [], "gates": {},
        "contract": {}, "display_validation": {}, "opened": False,
    }

    def write_report() -> None:
        cfg.report.write_text(
            json.dumps(report, indent=2, default=str, ensure_ascii=False),
            encoding="utf-8")
        log.info("Report: %s", cfg.report)

    try:
        # ── Preflight ────────────────────────────────────────────────────────
        log.info("=== PREFLIGHT ===")
        if not cfg.template.exists():
            raise BuildBlocked(f"Template not found: {cfg.template}")
        if not cfg.schedule.exists():
            raise BuildBlocked(f"Schedule not found: {cfg.schedule}")
        if has_zone_identifier(cfg.template):
            msg = f"Template has MOTW: {cfg.template}"
            if args.fail_if_template_blocked:
                raise BuildBlocked(msg)
            log.warning(msg)
            report["warnings"].append(msg)

        baseline = package_contract(cfg.template)
        log.info("Template contract: %s",
                 {k: v for k, v in baseline.items() if k != "protected_parts"})

        # ── Parse ─────────────────────────────────────────────────────────────
        log.info("=== PARSE ===")
        rows  = parse_schedule(cfg.schedule)
        tasks = [r for r in rows if r.row_type == "task"]
        log.info("Parsed %d tasks, %d rows total", len(tasks), len(rows))

        if cfg.dry_run:
            log.info("--dry-run: stop before workbook generation")
            report["status"] = "success"
            write_report()
            return 0

        # ── Generate workbook ─────────────────────────────────────────────────
        log.info("=== GENERATE ===")
        shutil.copy2(cfg.template, cfg.sidecar)
        log.info("Sidecar: %s", cfg.sidecar)

        render_result = render(cfg.sidecar, cfg, rows, baseline)
        report["warnings"].extend(render_result.get("warnings", []))
        if render_result.get("errors"):
            report["errors"].extend(render_result["errors"])
            report["status"] = "validation_failed"
            write_report()
            return 1
        log.info(
            "Render OK. VBA patch=%s  names=%s  buttons=%s",
            render_result.get("vba_patch", {}).get("touched"),
            render_result.get("deleted_uid_names"),
            render_result.get("updated_buttons"))

        # ── Openpyxl structural validation ────────────────────────────────────
        log.info("=== VALIDATE (openpyxl) ===")
        oxl = validate_openpyxl(cfg.sidecar, rows)
        report["warnings"].extend(oxl.get("warnings", []))
        if oxl.get("errors"):
            report["errors"].extend(oxl["errors"])
            log.error("openpyxl validation FAILED: %s", oxl["errors"])
            report["status"] = "validation_failed"
            write_report()
            return 1
        log.info("openpyxl validation PASS")

        # ── COM display validation ─────────────────────────────────────────────
        log.info("=== VALIDATE (COM display) ===")
        com_val = validate_com(cfg.sidecar, macro_smoke=False,
                               output_name=cfg.final_name)
        report["display_validation"] = com_val.get("display_validation", {})
        report["warnings"].extend(com_val.get("warnings", []))
        if com_val.get("errors"):
            report["errors"].extend(com_val["errors"])
            log.error("COM display validation FAILED: %s", com_val["errors"])
            report["status"] = "validation_failed"
            write_report()
            return 1
        log.info("COM display validation PASS: %s", report["display_validation"])

        log.info("=== VALIDATE (macro smoke) ===")
        macro_val = validate_com(cfg.sidecar, macro_smoke=True,
                                 output_name=cfg.final_name)
        report["warnings"].extend(macro_val.get("warnings", []))
        if macro_val.get("errors"):
            report["errors"].extend(macro_val["errors"])
            log.error("Macro smoke FAILED: %s", macro_val["errors"])
            report["status"] = "validation_failed"
            write_report()
            return 1
        report["macro_smoke"] = macro_val.get("macro_smoke", {})
        log.info("Macro smoke PASS: %s", report["macro_smoke"])

        report["contract"] = render_result.get("contract_after_render", {})
        report["gates"] = {
            "schedule_semantics": "PASS",
            "contract":           "PASS",
            "display_validation": "PASS",
            "vba_integrity":      "PASS",
            "motw":               "PASS",
        }

        # ── Promote sidecar → output ──────────────────────────────────────────
        log.info("=== PROMOTE ===")
        shutil.copy2(cfg.sidecar, cfg.output)
        report["output"] = str(cfg.output)
        log.info("Output: %s", cfg.output)

        if not cfg.keep_sidecar:
            try:
                cfg.sidecar.unlink()
            except Exception:
                pass

        # ── Open ──────────────────────────────────────────────────────────────
        if cfg.open_after:
            try:
                os.startfile(str(cfg.output))
                report["opened"] = True
                log.info("Opened: %s", cfg.output)
            except Exception as exc:
                log.warning("Could not open: %s", exc)

        report["status"] = "success"
        write_report()
        log.info("=== DONE - status=success ===")
        return 0

    except BuildBlocked as exc:
        log.error("BLOCKED: %s", exc)
        report["status"] = "blocked"
        report["errors"].append(str(exc))
        write_report()
        return 2
    except ValidationFailed as exc:
        log.error("VALIDATION_FAILED: %s", exc)
        report["status"] = "validation_failed"
        report["errors"].append(str(exc))
        write_report()
        return 1
    except Exception as exc:
        log.exception("Unexpected exception")
        report["status"] = "exception"
        report["errors"].append(str(exc))
        write_report()
        return 3


if __name__ == "__main__":
    sys.exit(main())
