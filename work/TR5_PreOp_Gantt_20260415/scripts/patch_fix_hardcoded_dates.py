"""
patch_fix_hardcoded_dates.py
Step A: delete_cols(11,2) — revert Apr25/Apr26 column insertion
Step B: COM VBA patch — replace all 6 RefreshGanttExactRange hardcoded calls
        with RefreshGantt ws, False
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pythoncom, win32com.client as win32, gc

XLSM = r'C:\tr_dash-main\work\TR5_PreOp_Gantt_20260415\excel\TR5_PreOp_Gantt_20260415_162140.xlsm'
HEADER_ROW = 3
DATE_COL_START = 11  # K

# ── Step A: openpyxl — revert 2-column insertion ─────────────────────────────
print("Step A: Reverting Apr25/Apr26 column insertion via openpyxl...")
wb_opy = load_workbook(XLSM, keep_vba=True)
ws_opy = wb_opy['Gantt_BASE']

k3_before = ws_opy.cell(HEADER_ROW, DATE_COL_START).value
l3_before = ws_opy.cell(HEADER_ROW, DATE_COL_START + 1).value
m3_before = ws_opy.cell(HEADER_ROW, DATE_COL_START + 2).value
print(f"  Before: K3={k3_before!r}  L3={l3_before!r}  M3={m3_before!r}")

# Only delete if K3 looks like Apr25 (46137) or L3 looks like Apr26 (46138)
# and M3 looks like Apr27 (46139)
need_revert = False
if isinstance(k3_before, int) and k3_before == 46137:
    need_revert = True
    print("  K3=46137 (Apr25) detected — reverting 2 cols")
elif hasattr(k3_before, 'year') and k3_before.year == 2026 and k3_before.month == 4 and k3_before.day == 25:
    need_revert = True
    print("  K3=datetime(Apr25) detected — reverting 2 cols")
else:
    print(f"  K3 is not Apr25 ({k3_before!r}) — no revert needed")

if need_revert:
    ws_opy.delete_cols(DATE_COL_START, 2)
    k3_after = ws_opy.cell(HEADER_ROW, DATE_COL_START).value
    print(f"  After delete: K3={k3_after!r}")
    wb_opy.save(XLSM)
    print("  Saved (openpyxl).")
else:
    print("  No openpyxl change needed.")

del wb_opy
print()

# ── Step B: COM VBA patch ─────────────────────────────────────────────────────
print("Step B: Patching VBA via COM...")

OLD_EXACT = "RefreshGanttExactRange ws, DateSerial(2026, 4, 27), DateSerial(2026, 5, 17), False"
NEW_GANTT  = "RefreshGantt ws, False"

def normalize(s: str) -> str:
    return s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\r\n")

pythoncom.CoInitialize()
app = win32.DispatchEx("Excel.Application")
app.Visible = False; app.DisplayAlerts = False; app.EnableEvents = False
app.AutomationSecurity = 3

try:
    wb = app.Workbooks.Open(XLSM, UpdateLinks=0, ReadOnly=False)
    cm = wb.VBProject.VBComponents("modMIR_Gantt_Unified").CodeModule
    n = cm.CountOfLines
    code = normalize(cm.Lines(1, n))

    count_before = code.count(OLD_EXACT)
    print(f"  Occurrences of hardcoded string: {count_before}")

    if count_before == 0:
        print("  [WARN] Pattern not found. Trying CRLF variant...")
        code = cm.Lines(1, n)
        count_before = code.count(OLD_EXACT)
        print(f"  After raw read: {count_before}")

    if count_before > 0:
        new_code = code.replace(OLD_EXACT, NEW_GANTT)
        replaced = code.count(OLD_EXACT) - new_code.count(OLD_EXACT)
        print(f"  Replaced {replaced} occurrences → saving VBA...")
        cm.DeleteLines(1, n)
        cm.InsertLines(1, new_code)
        print(f"  VBA module updated ({cm.CountOfLines} lines)")
    else:
        print("  [ERROR] Pattern still not found. Checking all occurrences of 'RefreshGanttExactRange':")
        lines = cm.Lines(1, n).splitlines()
        for i, ln in enumerate(lines):
            if "RefreshGanttExactRange" in ln:
                print(f"    {i+1:4}: {ln!r}")

    # Verify: count remaining hardcoded refs
    final_code = cm.Lines(1, cm.CountOfLines)
    remaining = final_code.count(OLD_EXACT)
    print(f"  Remaining hardcoded refs: {remaining}")

    wb.Save()
    print("  Workbook saved (COM).")
    wb.Close(False)

finally:
    app.Quit(); del app; gc.collect(); pythoncom.CoUninitialize()

print("\nAll done.")
