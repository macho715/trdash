
from __future__ import annotations

import argparse
import contextlib
import datetime as dt
import gc
import hashlib
import json
import logging
import os
import platform
import shutil
import sys
import time
import traceback
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterator

import openpyxl
import pythoncom
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# ---------------------------------------------------------------------------
# Defaults (kept compatible with the original script unless overridden)
# ---------------------------------------------------------------------------

DEFAULT_ROOT = Path(r"C:\tr_dash-main")
DEFAULT_TARGET = (
    DEFAULT_ROOT
    / "output"
    / "spreadsheet"
    / "tr5_preparation_simulation"
    / "TR5_Preparation_Simulation_Gantt_MIR_Format_COM.xlsm"
)

DEFAULT_SHEET = "Gantt_BASE"
DEFAULT_FIRST_TASK_ROW = 4
DEFAULT_LAST_TASK_ROW = 42
DEFAULT_FIRST_TIMELINE_COL = 11  # K
DEFAULT_LAST_TIMELINE_COL = 31  # AE
DEFAULT_CLEAR_UNTIL_ROW = 80
DEFAULT_CLEAR_UNTIL_COL = 67  # BO

DEFAULT_TODAY = dt.date(2026, 4, 15)
DEFAULT_TIMELINE_START = dt.date(2026, 4, 27)
DEFAULT_TIMELINE_END = dt.date(2026, 5, 17)
DEFAULT_SOURCE_LABEL = "TR5_Pre-Op_Simulation_v2_20260430.md"
DEFAULT_T_PLUS_ZERO = dt.date(2026, 4, 30)
DEFAULT_LOAD_OUT = dt.date(2026, 5, 9)
DEFAULT_AGI_JD = dt.date(2026, 5, 17)

EXCEL_EPOCH = dt.date(1899, 12, 30)
RISK_VALUES = ["OK", "AMBER", "HIGH", "WARNING", "CRITICAL", "GATE"]

# Excel constants used by COM (kept explicit to avoid makepy dependencies)
XL_CENTER = -4108
XL_CONTINUOUS = 1
XL_THIN = 2
XL_VALIDATE_LIST = 3
XL_VALID_ALERT_STOP = 1
XL_FORMAT_CONDITION_EXPRESSION = 2
MSO_AUTOMATION_SECURITY_FORCE_DISABLE = 3


def xl_color(hex_rgb: str) -> int:
    """Convert #RRGGBB to Excel BGR integer."""
    value = hex_rgb.strip().lstrip("#")
    if len(value) != 6:
        raise ValueError(f"Expected 6 hex chars, got {hex_rgb!r}")
    r = int(value[0:2], 16)
    g = int(value[2:4], 16)
    b = int(value[4:6], 16)
    return r + (g * 256) + (b * 65536)


FALLBACK_RISK_COLORS = {
    "OK": xl_color("DCEAD7"),
    "AMBER": xl_color("F7E0C4"),
    "HIGH": xl_color("F4D7C8"),
    "WARNING": xl_color("FFF1B8"),
    "CRITICAL": xl_color("FF5C63"),
    "GATE": xl_color("F5C84C"),
}
FALLBACK_BAR_COLORS = {
    "OK": xl_color("5FB5F2"),
    "AMBER": xl_color("F1A954"),
    "HIGH": xl_color("F08A5D"),
    "WARNING": xl_color("F5D04A"),
    "CRITICAL": xl_color("FF5C63"),
    "GATE": xl_color("F5C84C"),
}
RISK_FONT_COLORS = {
    "OK": xl_color("102417"),
    "AMBER": xl_color("3D2A0D"),
    "HIGH": xl_color("3B1E14"),
    "WARNING": xl_color("352800"),
    "CRITICAL": xl_color("FFFFFF"),
    "GATE": xl_color("0D1320"),
}


class PatchError(RuntimeError):
    """Raised when patching or validation cannot safely continue."""


@dataclass(frozen=True)
class PatchConfig:
    target: Path = DEFAULT_TARGET
    sheet: str = DEFAULT_SHEET
    first_task_row: int = DEFAULT_FIRST_TASK_ROW
    last_task_row: int = DEFAULT_LAST_TASK_ROW
    first_timeline_col: int = DEFAULT_FIRST_TIMELINE_COL
    last_timeline_col: int = DEFAULT_LAST_TIMELINE_COL
    clear_until_row: int = DEFAULT_CLEAR_UNTIL_ROW
    clear_until_col: int = DEFAULT_CLEAR_UNTIL_COL
    today: dt.date = DEFAULT_TODAY
    timeline_start: dt.date = DEFAULT_TIMELINE_START
    timeline_end: dt.date = DEFAULT_TIMELINE_END
    source_label: str = DEFAULT_SOURCE_LABEL
    t_plus_zero: dt.date = DEFAULT_T_PLUS_ZERO
    load_out: dt.date = DEFAULT_LOAD_OUT
    agi_jd: dt.date = DEFAULT_AGI_JD
    run_dir: Path | None = None
    promote: bool = True
    open_retries: int = 3
    retry_delay_seconds: float = 1.0
    close_saved_live_target: bool = True

    def __post_init__(self) -> None:
        if self.first_task_row > self.last_task_row:
            raise ValueError("first_task_row must be <= last_task_row")
        if self.first_timeline_col > self.last_timeline_col:
            raise ValueError("first_timeline_col must be <= last_timeline_col")
        expected_days = (self.timeline_end - self.timeline_start).days + 1
        actual_days = self.last_timeline_col - self.first_timeline_col + 1
        if expected_days != actual_days:
            raise ValueError(
                f"Timeline day count mismatch: dates imply {expected_days}, columns imply {actual_days}"
            )
        if self.clear_until_col < self.last_timeline_col:
            raise ValueError("clear_until_col must be >= last_timeline_col")
        if self.clear_until_row < self.last_task_row:
            raise ValueError("clear_until_row must be >= last_task_row")

    @property
    def parent_dir(self) -> Path:
        return self.target.parent

    @property
    def effective_run_dir(self) -> Path:
        if self.run_dir is not None:
            return self.run_dir
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.parent_dir / f"scope_patch_run_{stamp}"

    @property
    def target_name(self) -> str:
        return self.target.name

    @property
    def sidecar(self) -> Path:
        return self.effective_run_dir / f"{self.target.stem}.sidecar{self.target.suffix}"

    @property
    def backup(self) -> Path:
        return self.effective_run_dir / f"{self.target.stem}.before_scope_patch{self.target.suffix}"

    @property
    def report(self) -> Path:
        return self.effective_run_dir / "validation_report.json"

    @property
    def log_file(self) -> Path:
        return self.effective_run_dir / "scope_patch.log"

    @property
    def timeline_end_col_letter(self) -> str:
        return col_letter(self.last_timeline_col)

    @property
    def clear_until_col_letter(self) -> str:
        return col_letter(self.clear_until_col)

    @property
    def risk_range(self) -> str:
        return f"G{self.first_task_row}:G{self.last_task_row}"

    @property
    def bar_range(self) -> str:
        return (
            f"{col_letter(self.first_timeline_col)}{self.first_task_row + 1}:"
            f"{col_letter(self.last_timeline_col)}{self.last_task_row}"
        )

    @property
    def timeline_header_range(self) -> str:
        return (
            f"{col_letter(self.first_timeline_col)}3:"
            f"{col_letter(self.last_timeline_col)}3"
        )

    @property
    def title_range(self) -> str:
        return f"D1:{self.timeline_end_col_letter}1"

    @property
    def subtitle_range(self) -> str:
        return f"D2:{self.timeline_end_col_letter}2"

    @property
    def timeline_body_range(self) -> str:
        return (
            f"{col_letter(self.first_timeline_col)}3:"
            f"{col_letter(self.last_timeline_col)}{self.last_task_row}"
        )

    @property
    def active_scope_range(self) -> str:
        return f"A1:{self.timeline_end_col_letter}{self.last_task_row}"

    @property
    def patch_scope_range(self) -> str:
        return f"A1:{self.clear_until_col_letter}{self.clear_until_row}"

    @property
    def right_clear_range(self) -> str | None:
        if self.clear_until_col <= self.last_timeline_col:
            return None
        start = col_letter(self.last_timeline_col + 1)
        return f"{start}1:{self.clear_until_col_letter}{self.clear_until_row}"

    @property
    def bottom_clear_range(self) -> str | None:
        if self.clear_until_row <= self.last_task_row:
            return None
        start_row = self.last_task_row + 1
        return f"A{start_row}:{self.clear_until_col_letter}{self.clear_until_row}"

    @property
    def title(self) -> str:
        return "TR5 PRE-OP SIMULATION - PHASE-GATE GANTT"

    @property
    def subtitle(self) -> str:
        return (
            '=TEXT(TODAY(),"yyyy-mm-dd")'
            '&" | Source: "&IFERROR(MIR_SOURCE_LABEL,"TR5 schedule")'
            '&" | T+0="&TEXT(IFERROR(MIR_T_PLUS_ZERO,INDEX($3:$3,1,11)),"yyyy-mm-dd")'
            '&" | Load-out="&IFERROR(TEXT(LOOKUP(2,1/($A$1:$A$100="BD2"),$D$1:$D$100),"yyyy-mm-dd"),"")'
            '&" | AGI JD="&IFERROR(TEXT(LOOKUP(2,1/($A$1:$A$100="JDC"),$D$1:$D$100),"yyyy-mm-dd"),"")'
        )


def col_letter(col: int) -> str:
    value = ""
    while col:
        col, rem = divmod(col - 1, 26)
        value = chr(65 + rem) + value
    return value


def excel_serial(day: dt.date) -> int:
    return (day - EXCEL_EPOCH).days


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def same_path(a: Path, b: Path) -> bool:
    try:
        return a.resolve().samefile(b)
    except Exception:
        return str(a.resolve(strict=False)).casefold() == str(b.resolve(strict=False)).casefold()


def setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("tr5_scope_patch")
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    return logger


def protected_zip_parts(path: Path) -> dict[str, dict[str, int]]:
    """
    Track the key package parts that should not be silently corrupted:
    vba project, drawing XML, VML, and control property XML.
    """
    tracked: dict[str, dict[str, int]] = {}
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            name = info.filename
            if (
                name == "xl/vbaProject.bin"
                or (name.startswith("xl/drawings/drawing") and name.endswith(".xml"))
                or (name.startswith("xl/drawings/vmlDrawing") and name.endswith(".vml"))
                or (name.startswith("xl/ctrlProps/") and name.endswith(".xml"))
            ):
                tracked[name] = {
                    "crc": int(info.CRC),
                    "file_size": int(info.file_size),
                }
    return tracked


def package_contract(path: Path) -> dict[str, Any]:
    protected = protected_zip_parts(path)
    return {
        "vbaProject.bin": int("xl/vbaProject.bin" in protected),
        "drawings": sum(1 for name in protected if name.startswith("xl/drawings/drawing")),
        "ctrlProps": sum(1 for name in protected if name.startswith("xl/ctrlProps/")),
        "vmlDrawing": sum(1 for name in protected if name.startswith("xl/drawings/vmlDrawing")),
        "protected_parts": protected,
    }


@contextlib.contextmanager
def excel_application(visible: bool = False) -> Iterator[Any]:
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    original: dict[str, Any] = {}

    def _remember(prop: str) -> None:
        try:
            original[prop] = getattr(excel, prop)
        except Exception:
            pass

    try:
        for prop in ("Visible", "DisplayAlerts", "EnableEvents", "ScreenUpdating", "AutomationSecurity"):
            _remember(prop)
        excel.Visible = visible
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.ScreenUpdating = False
        try:
            excel.AutomationSecurity = MSO_AUTOMATION_SECURITY_FORCE_DISABLE
        except Exception:
            pass
        yield excel
    finally:
        for prop, value in original.items():
            try:
                setattr(excel, prop, value)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass
        del excel
        gc.collect()
        pythoncom.CoUninitialize()


def open_workbook_with_retry(excel: Any, path: Path, readonly: bool, retries: int, delay_seconds: float) -> Any:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            wb = excel.Workbooks.Open(
                str(path),
                UpdateLinks=0,
                ReadOnly=readonly,
                IgnoreReadOnlyRecommended=True,
            )
            if not readonly and bool(getattr(wb, "ReadOnly", False)):
                wb.Close(SaveChanges=False)
                raise PatchError(f"Workbook opened read-only unexpectedly: {path}")
            return wb
        except Exception as exc:  # pragma: no cover - COM only
            last_error = exc
            if attempt >= retries:
                break
            time.sleep(delay_seconds * attempt)
    raise PatchError(f"Failed to open workbook after {retries} attempts: {path}") from last_error


def set_number_format(rng: Any, fmt: str) -> None:
    rng.NumberFormat = fmt
    try:
        rng.NumberFormatLocal = fmt
    except Exception:
        pass


def safe_calculate(excel: Any, wb: Any = None, ws: Any = None) -> bool:
    for obj, method_name in (
        (excel, "CalculateFullRebuild"),
        (excel, "CalculateFull"),
        (ws, "Calculate"),
        (wb, "RefreshAll"),
    ):
        if obj is None:
            continue
        try:
            method = getattr(obj, method_name)
            if callable(method):
                method()
                return True
        except Exception:
            continue
    return False


def check_live_target_not_unsaved(target: Path, close_saved: bool) -> list[str]:
    closed: list[str] = []
    try:
        app = win32.GetActiveObject("Excel.Application")
    except Exception:
        return closed

    try:
        workbooks = list(app.Workbooks)
    except Exception:
        return closed

    for wb in workbooks:
        try:
            full_name = Path(str(wb.FullName))
        except Exception:
            continue
        try:
            if same_path(full_name, target):
                if not wb.Saved:
                    raise PatchError(f"Target workbook is open with unsaved changes: {wb.FullName}")
                if close_saved:
                    wb.Close(SaveChanges=False)
                    closed.append(str(wb.FullName))
                else:
                    raise PatchError(f"Target workbook is already open: {wb.FullName}")
        except PatchError:
            raise
        except Exception:
            continue
    return closed


def used_task_rows(ws: Any, cfg: PatchConfig) -> list[int]:
    rows: list[int] = []
    for row in range(cfg.first_task_row, cfg.last_task_row + 1):
        if ws.Cells(row, 4).Value2 and ws.Cells(row, 5).Value2:
            rows.append(row)
    return rows


def sample_current_display_colors(
    ws: Any,
    cfg: PatchConfig,
    task_rows: list[int],
) -> tuple[dict[str, int], dict[str, int]]:
    risk_colors: dict[str, int] = {}
    bar_colors: dict[str, int] = {}

    for row in task_rows:
        risk = str(ws.Cells(row, 7).Value or "").strip().upper()
        if risk not in RISK_VALUES:
            continue

        if risk not in risk_colors:
            try:
                risk_colors[risk] = int(ws.Cells(row, 7).DisplayFormat.Interior.Color)
            except Exception:
                pass

        if risk not in bar_colors:
            start = ws.Cells(row, 4).Value2
            end = ws.Cells(row, 5).Value2
            if not isinstance(start, (int, float)) or not isinstance(end, (int, float)):
                continue

            for col in range(cfg.first_timeline_col, cfg.last_timeline_col + 1):
                header = ws.Cells(3, col).Value2
                if isinstance(header, (int, float)) and int(start) <= int(header) <= int(end):
                    try:
                        color = int(ws.Cells(row, col).DisplayFormat.Interior.Color)
                        if color:
                            bar_colors[risk] = color
                            break
                    except Exception:
                        pass

    for risk, color in FALLBACK_RISK_COLORS.items():
        risk_colors.setdefault(risk, color)
    for risk, color in FALLBACK_BAR_COLORS.items():
        bar_colors.setdefault(risk, color)
    return risk_colors, bar_colors


def clear_visual_state_outside_scope(ws: Any, cfg: PatchConfig) -> None:
    for address in (cfg.right_clear_range, cfg.bottom_clear_range):
        if not address:
            continue
        rng = ws.Range(address)
        rng.ClearContents()
        rng.ClearFormats()
        try:
            rng.Validation.Delete()
        except Exception:
            pass


def apply_patch_with_com(path: Path, cfg: PatchConfig, logger: logging.Logger) -> dict[str, Any]:
    logger.info("Applying COM patch to sidecar: %s", path)
    with excel_application(visible=False) as excel:
        wb = None
        closed = False
        try:
            wb = open_workbook_with_retry(
                excel=excel,
                path=path,
                readonly=False,
                retries=cfg.open_retries,
                delay_seconds=cfg.retry_delay_seconds,
            )
            ws = wb.Worksheets(cfg.sheet)
            task_rows = used_task_rows(ws, cfg)
            if not task_rows:
                raise PatchError("No task rows with both start/end dates were found in the configured task range.")

            risk_colors, bar_colors = sample_current_display_colors(ws, cfg, task_rows)

            # Limit CF deletion to the patch area instead of the whole worksheet.
            ws.Range(cfg.patch_scope_range).FormatConditions.Delete()

            title_reset_range = f"A1:{cfg.clear_until_col_letter}2"
            ws.Range(title_reset_range).UnMerge()
            ws.Range(title_reset_range).ClearContents()
            ws.Range(title_reset_range).ClearFormats()

            ws.Range(cfg.title_range).Merge()
            ws.Range(cfg.subtitle_range).Merge()
            ws.Range("D1").Value = cfg.title
            try:
                wb.Names("MIR_SOURCE_LABEL").Delete()
            except Exception:
                pass
            source_name = wb.Names.Add(
                Name="MIR_SOURCE_LABEL",
                RefersTo=f'="{cfg.source_label}"',
            )
            try:
                source_name.Visible = False
            except Exception:
                pass
            try:
                wb.Names("MIR_T_PLUS_ZERO").Delete()
            except Exception:
                pass
            t0_name = wb.Names.Add(Name="MIR_T_PLUS_ZERO", RefersTo=f"=DATE({cfg.t_plus_zero.year},{cfg.t_plus_zero.month},{cfg.t_plus_zero.day})")
            try:
                t0_name.Visible = False
            except Exception:
                pass
            ws.Range("D2").Formula = cfg.subtitle

            title_rng = ws.Range(cfg.title_range)
            title_rng.Interior.Color = xl_color("0B1120")
            title_rng.Font.Name = "Arial"
            title_rng.Font.Size = 13
            title_rng.Font.Bold = True
            title_rng.Font.Color = xl_color("F5C84C")
            title_rng.HorizontalAlignment = XL_CENTER
            title_rng.VerticalAlignment = XL_CENTER

            subtitle_rng = ws.Range(cfg.subtitle_range)
            subtitle_rng.Interior.Color = xl_color("13161E")
            subtitle_rng.Font.Name = "Arial"
            subtitle_rng.Font.Size = 9
            subtitle_rng.Font.Bold = False
            subtitle_rng.Font.Color = xl_color("7D879C")
            subtitle_rng.HorizontalAlignment = XL_CENTER
            subtitle_rng.VerticalAlignment = XL_CENTER

            for row in range(cfg.first_task_row, cfg.last_task_row + 1):
                for col in (4, 5):
                    value = ws.Cells(row, col).Value2
                    if isinstance(value, (int, float)):
                        ws.Cells(row, col).Value2 = int(value)

            date_range = f"D{cfg.first_task_row}:E{cfg.last_task_row}"
            set_number_format(ws.Range(date_range), "yyyy-mm-dd")
            for row in range(cfg.first_task_row, cfg.last_task_row + 1):
                for col in (4, 5):
                    set_number_format(ws.Cells(row, col), "yyyy-mm-dd")

            for offset, col in enumerate(range(cfg.first_timeline_col, cfg.last_timeline_col + 1)):
                day = cfg.timeline_start + dt.timedelta(days=offset)
                cell = ws.Cells(3, col)
                cell.Value2 = excel_serial(day)
                cell.Font.Name = "Arial"
                cell.Font.Size = 7
                cell.Font.Bold = True
                cell.Font.Color = xl_color("6F7A92")
                cell.Orientation = 90
                cell.HorizontalAlignment = XL_CENTER
                cell.VerticalAlignment = XL_CENTER
                ws.Columns(col).ColumnWidth = 5.0

            set_number_format(ws.Range(cfg.timeline_header_range), "mm-dd")
            for col in range(cfg.first_timeline_col, cfg.last_timeline_col + 1):
                set_number_format(ws.Cells(3, col), "mm-dd")

            clear_visual_state_outside_scope(ws, cfg)

            inside_timeline = ws.Range(cfg.timeline_body_range)
            inside_timeline.Interior.Color = xl_color("0B1020")
            inside_timeline.Borders.Color = xl_color("242A38")
            inside_timeline.Borders.LineStyle = XL_CONTINUOUS
            inside_timeline.Borders.Weight = XL_THIN

            ws.Range(cfg.timeline_header_range).Interior.Color = xl_color("0B1120")
            header_rng = ws.Range("A3:G3")
            header_rng.Interior.Color = xl_color("0B1120")
            header_rng.Font.Name = "Arial"
            header_rng.Font.Size = 9
            header_rng.Font.Bold = True
            header_rng.Font.Color = xl_color("F5C84C")
            header_rng.HorizontalAlignment = XL_CENTER
            header_rng.VerticalAlignment = XL_CENTER

            risk_rng = ws.Range(cfg.risk_range)
            try:
                risk_rng.Validation.Delete()
            except Exception:
                pass
            risk_rng.Validation.Add(
                Type=XL_VALIDATE_LIST,
                AlertStyle=XL_VALID_ALERT_STOP,
                Operator=XL_VALID_ALERT_STOP,
                Formula1=",".join(RISK_VALUES),
            )
            risk_rng.Validation.IgnoreBlank = True
            risk_rng.Validation.InCellDropdown = True

            for risk in RISK_VALUES:
                fc = risk_rng.FormatConditions.Add(
                    XL_FORMAT_CONDITION_EXPRESSION,
                    XL_VALID_ALERT_STOP,
                    f'=$G{cfg.first_task_row}="{risk}"',
                )
                fc.Interior.Color = risk_colors[risk]
                fc.Font.Color = RISK_FONT_COLORS[risk]
                fc.Font.Bold = True

            bar_rng = ws.Range(cfg.bar_range)
            for risk in RISK_VALUES:
                formula = f'=AND({col_letter(cfg.first_timeline_col)}$3>=$D{cfg.first_task_row + 1},' \
                          f'{col_letter(cfg.first_timeline_col)}$3<=$E{cfg.first_task_row + 1},' \
                          f'$G{cfg.first_task_row + 1}="{risk}")'
                fc = bar_rng.FormatConditions.Add(
                    XL_FORMAT_CONDITION_EXPRESSION,
                    XL_VALID_ALERT_STOP,
                    formula,
                )
                fc.Interior.Color = bar_colors[risk]
                fc.Font.Color = xl_color("0D1320")
                fc.Font.Bold = True

            ws.Range(cfg.active_scope_range).Font.Name = "Arial"
            ws.Range(cfg.active_scope_range).Rows.RowHeight = 17.4
            ws.Rows(1).RowHeight = 31.95
            ws.Rows(2).RowHeight = 18
            ws.Rows(3).RowHeight = 42
            ws.Range(date_range).Columns.ColumnWidth = 12
            set_number_format(ws.Range(date_range), "yyyy-mm-dd")
            set_number_format(ws.Range(cfg.timeline_header_range), "mm-dd")

            # Display-format smoke test on the sidecar only.
            test_row = task_rows[0]
            original_risk = ws.Cells(test_row, 7).Value
            first_bar_col: int | None = None

            try:
                ws.Cells(test_row, 7).Value = "CRITICAL"
                calculate_ok_critical = safe_calculate(excel, wb, ws)
                critical_risk_color = int(ws.Cells(test_row, 7).DisplayFormat.Interior.Color)

                start = ws.Cells(test_row, 4).Value2
                end = ws.Cells(test_row, 5).Value2
                for col in range(cfg.first_timeline_col, cfg.last_timeline_col + 1):
                    header = ws.Cells(3, col).Value2
                    if isinstance(header, (int, float)) and isinstance(start, (int, float)) and isinstance(end, (int, float)):
                        if int(start) <= int(header) <= int(end):
                            first_bar_col = col
                            break
                if first_bar_col is None:
                    raise PatchError(f"Could not locate first gantt bar column for task row {test_row}")

                critical_bar_color = int(ws.Cells(test_row, first_bar_col).DisplayFormat.Interior.Color)

                ws.Cells(test_row, 7).Value = "GATE"
                calculate_ok_gate = safe_calculate(excel, wb, ws)
                gate_risk_color = int(ws.Cells(test_row, 7).DisplayFormat.Interior.Color)
                gate_bar_color = int(ws.Cells(test_row, first_bar_col).DisplayFormat.Interior.Color)
            finally:
                ws.Cells(test_row, 7).Value = original_risk
                calculate_ok_restore = safe_calculate(excel, wb, ws)

            wb.Save()
            wb.Close(SaveChanges=True)
            closed = True

            return {
                "task_rows": task_rows,
                "sampled_risk_colors": risk_colors,
                "sampled_bar_colors": bar_colors,
                "display_test": {
                    "row": test_row,
                    "bar_col": col_letter(first_bar_col or cfg.first_timeline_col),
                    "critical_risk_color": critical_risk_color,
                    "critical_bar_color": critical_bar_color,
                    "gate_risk_color": gate_risk_color,
                    "gate_bar_color": gate_bar_color,
                    "calculate_ok": calculate_ok_critical or calculate_ok_gate or calculate_ok_restore,
                    "risk_changed": critical_risk_color != gate_risk_color,
                    "bar_changed": critical_bar_color != gate_bar_color,
                },
            }
        finally:
            if wb is not None and not closed:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass


def has_visible_format(cell: Any) -> bool:
    fill = cell.fill
    if fill and fill.fill_type and fill.fill_type != "none":
        return True
    for side_name in ("left", "right", "top", "bottom"):
        side = getattr(cell.border, side_name)
        if side and side.style:
            return True
    return False


def cell_date_value(value: Any) -> Any:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        return EXCEL_EPOCH + dt.timedelta(days=int(value))
    return value


def sqref_areas(sqref: str) -> list[tuple[int, int, int, int]]:
    return [range_boundaries(part) for part in str(sqref).split()]


def conditional_format_ranges(ws: Any) -> list[str]:
    return [str(cf.sqref) for cf in ws.conditional_formatting]


def validate(path: Path, baseline: dict[str, Any], com_result: dict[str, Any], cfg: PatchConfig) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []

    wb = load_workbook(path, keep_vba=True, data_only=False)
    ws = wb[cfg.sheet]

    if ws["D1"].value != cfg.title:
        errors.append("title mismatch")

    subtitle_value = str(ws["D2"].value)
    required_fragments = [
        f"Today: {cfg.today:%Y-%m-%d}",
        f"Source: {cfg.source_label}",
        f"T+0={cfg.t_plus_zero:%Y-%m-%d}",
        f"Load-out={cfg.load_out:%Y-%m-%d}",
        f"AGI JD={cfg.agi_jd:%Y-%m-%d}",
    ]
    for fragment in required_fragments:
        if fragment not in subtitle_value:
            errors.append(f"subtitle fragment missing: {fragment}")

    merged = {str(r) for r in ws.merged_cells.ranges}
    if cfg.title_range not in merged or cfg.subtitle_range not in merged:
        errors.append(f"title/subtitle merge scope mismatch: {sorted(merged)}")

    legacy_title_merge = f"A1:{cfg.clear_until_col_letter}1"
    legacy_subtitle_merge = f"A2:{cfg.clear_until_col_letter}2"
    if legacy_title_merge in merged and legacy_title_merge != cfg.title_range:
        errors.append("legacy title merge still present")
    if legacy_subtitle_merge in merged and legacy_subtitle_merge != cfg.subtitle_range:
        errors.append("legacy subtitle merge still present")

    for row in range(cfg.first_task_row, cfg.last_task_row + 1):
        for col in (4, 5):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            fmt = cell.number_format or ""
            if fmt != "yyyy-mm-dd" and fmt != "mm-dd-yy" and not all(tok in fmt.lower() for tok in ("m", "d")):
                errors.append(f"{cell.coordinate} number format {cell.number_format!r}")
            if isinstance(cell.value, dt.datetime) and cell.value.time() != dt.time(0, 0):
                errors.append(f"{cell.coordinate} has non-midnight time {cell.value}")

    expected = cfg.timeline_start
    for col in range(cfg.first_timeline_col, cfg.last_timeline_col + 1):
        cell = ws.cell(3, col)
        value = cell_date_value(cell.value)
        if value != expected:
            errors.append(f"{cell.coordinate} expected {expected}, got {cell.value!r}")
        if cell.number_format not in ("mm-dd", "General", ""):
            errors.append(f"{cell.coordinate} timeline format {cell.number_format!r}")
        expected += dt.timedelta(days=1)
    if expected - dt.timedelta(days=1) != cfg.timeline_end:
        errors.append("timeline end calculation mismatch")

    for col in range(cfg.last_timeline_col + 1, cfg.clear_until_col + 1):
        if ws.cell(3, col).value is not None:
            errors.append(f"{ws.cell(3, col).coordinate} should be blank")
            break

    cf_ranges = conditional_format_ranges(ws)
    expected_risk_area = range_boundaries(cfg.risk_range)
    expected_bar_area = range_boundaries(cfg.bar_range)
    found_risk = False
    found_bar = False
    bad_cf: list[str] = []

    for sqref in cf_ranges:
        for area in sqref_areas(sqref):
            if area == expected_risk_area:
                found_risk = True
            elif area == expected_bar_area:
                found_bar = True
            else:
                bad_cf.append(sqref)

    if not found_risk:
        errors.append(f"risk CF range missing: {cf_ranges}")
    if not found_bar:
        errors.append(f"bar CF range missing: {cf_ranges}")
    if bad_cf:
        errors.append(f"conditional formatting outside scope: {sorted(set(bad_cf))}")

    dv_ranges = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    if dv_ranges != [cfg.risk_range]:
        errors.append(f"data validation ranges mismatch: {dv_ranges}")
    else:
        formula = str(ws.data_validations.dataValidation[0].formula1 or "")
        normalized_formula = formula.strip('"')
        if normalized_formula != ",".join(RISK_VALUES):
            errors.append(f"risk validation formula mismatch: {formula!r}")

    outside_hits: list[str] = []
    for row in range(1, cfg.clear_until_row + 1):
        for col in range(cfg.last_timeline_col + 1, cfg.clear_until_col + 1):
            cell = ws.cell(row, col)
            if cell.value is not None or has_visible_format(cell):
                outside_hits.append(cell.coordinate)
                break
        if outside_hits:
            break

    for row in range(cfg.last_task_row + 1, cfg.clear_until_row + 1):
        for col in range(1, cfg.clear_until_col + 1):
            cell = ws.cell(row, col)
            if cell.value is not None or has_visible_format(cell):
                outside_hits.append(cell.coordinate)
                break
        if len(outside_hits) > 1:
            break

    if outside_hits:
        errors.append(f"outside gantt scope still has visible content/format: {outside_hits[:5]}")

    contract_after = package_contract(path)
    if contract_after["vbaProject.bin"] != baseline["vbaProject.bin"]:
        errors.append("vbaProject.bin presence changed")
    if contract_after["drawings"] != baseline["drawings"]:
        errors.append(f"drawing count changed: before={baseline['drawings']}, after={contract_after['drawings']}")
    if contract_after["ctrlProps"] != baseline["ctrlProps"]:
        errors.append(f"ctrlProps count changed: before={baseline['ctrlProps']}, after={contract_after['ctrlProps']}")
    if contract_after["vmlDrawing"] != baseline["vmlDrawing"]:
        errors.append(f"vmlDrawing count changed: before={baseline['vmlDrawing']}, after={contract_after['vmlDrawing']}")
    if contract_after["protected_parts"].keys() != baseline["protected_parts"].keys():
        errors.append("protected package part set changed")

    display_test = com_result.get("display_test", {})
    if not display_test.get("risk_changed") or not display_test.get("bar_changed"):
        errors.append(f"display format test failed: {display_test}")

    wb.close()
    return {
        "errors": errors,
        "warnings": warnings,
        "cf_ranges": cf_ranges,
        "dv_ranges": dv_ranges,
        "contract_after": contract_after,
        "display_test": display_test,
    }


def readonly_reopen(path: Path, cfg: PatchConfig) -> bool:
    with excel_application(visible=False) as excel:
        wb = None
        try:
            wb = open_workbook_with_retry(
                excel=excel,
                path=path,
                readonly=True,
                retries=cfg.open_retries,
                delay_seconds=cfg.retry_delay_seconds,
            )
            wb.Close(SaveChanges=False)
            wb = None
            return True
        finally:
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass


def promote_sidecar(sidecar: Path, target: Path) -> Path:
    temp_copy = target.with_name(target.name + ".promote.tmp")
    if temp_copy.exists():
        temp_copy.unlink()
    shutil.copy2(sidecar, temp_copy)
    os.replace(temp_copy, target)
    return target


def safe_json_value(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(k): safe_json_value(v) for k, v in value.items()}
    if isinstance(value, list):
        return [safe_json_value(v) for v in value]
    return value


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(safe_json_value(report), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def parse_date(value: str) -> dt.date:
    return dt.datetime.strptime(value, "%Y-%m-%d").date()


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Patch and validate TR5 Gantt scope workbook safely.")
    p.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--first-task-row", type=int, default=DEFAULT_FIRST_TASK_ROW)
    p.add_argument("--last-task-row", type=int, default=DEFAULT_LAST_TASK_ROW)
    p.add_argument("--first-timeline-col", type=int, default=DEFAULT_FIRST_TIMELINE_COL)
    p.add_argument("--last-timeline-col", type=int, default=DEFAULT_LAST_TIMELINE_COL)
    p.add_argument("--timeline-start", type=parse_date, default=DEFAULT_TIMELINE_START)
    p.add_argument("--timeline-end", type=parse_date, default=DEFAULT_TIMELINE_END)
    p.add_argument("--clear-until-row", type=int, default=DEFAULT_CLEAR_UNTIL_ROW)
    p.add_argument("--clear-until-col", type=int, default=DEFAULT_CLEAR_UNTIL_COL)
    p.add_argument("--today", type=parse_date, default=DEFAULT_TODAY)
    p.add_argument("--source-label", default=DEFAULT_SOURCE_LABEL)
    p.add_argument("--t-plus-zero", type=parse_date, default=DEFAULT_T_PLUS_ZERO)
    p.add_argument("--load-out", type=parse_date, default=DEFAULT_LOAD_OUT)
    p.add_argument("--agi-jd", type=parse_date, default=DEFAULT_AGI_JD)
    p.add_argument("--run-dir", type=Path, default=None)
    p.add_argument("--no-promote", action="store_true", help="Patch/validate the sidecar only; do not replace target.")
    p.add_argument("--open-retries", type=int, default=3)
    p.add_argument("--retry-delay-seconds", type=float, default=1.0)
    p.add_argument(
        "--fail-if-live-open",
        action="store_true",
        help="Fail if the target workbook is already open, even when it is saved.",
    )
    return p


def config_from_args(args: argparse.Namespace) -> PatchConfig:
    run_dir = args.run_dir
    if run_dir is None:
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir = args.target.parent / f"scope_patch_run_{stamp}"
    return PatchConfig(
        target=args.target,
        sheet=args.sheet,
        first_task_row=args.first_task_row,
        last_task_row=args.last_task_row,
        first_timeline_col=args.first_timeline_col,
        last_timeline_col=args.last_timeline_col,
        clear_until_row=args.clear_until_row,
        clear_until_col=args.clear_until_col,
        today=args.today,
        timeline_start=args.timeline_start,
        timeline_end=args.timeline_end,
        source_label=args.source_label,
        t_plus_zero=args.t_plus_zero,
        load_out=args.load_out,
        agi_jd=args.agi_jd,
        run_dir=run_dir,
        promote=not args.no_promote,
        open_retries=args.open_retries,
        retry_delay_seconds=args.retry_delay_seconds,
        close_saved_live_target=not args.fail_if_live_open,
    )


def main() -> None:
    args = build_arg_parser().parse_args()
    cfg = config_from_args(args)

    started_at = time.perf_counter()
    cfg.effective_run_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logger(cfg.log_file)

    report: dict[str, Any] = {
        "status": "started",
        "target": cfg.target,
        "run_dir": cfg.effective_run_dir,
        "backup": cfg.backup,
        "sidecar": cfg.sidecar,
        "report": cfg.report,
        "log_file": cfg.log_file,
        "config": asdict(cfg),
        "python_version": platform.python_version(),
        "platform": platform.platform(),
        "openpyxl_version": openpyxl.__version__,
        "promoted": False,
        "final_readonly_reopen": False,
        "closed_live_workbooks": [],
        "exception": None,
        "traceback": None,
    }

    promoted = False
    exit_code = 1

    try:
        if not cfg.target.exists():
            raise FileNotFoundError(cfg.target)

        logger.info("Target workbook: %s", cfg.target)
        baseline = package_contract(cfg.target)
        report["baseline_contract"] = baseline
        report["baseline_sha256"] = sha256_file(cfg.target)

        closed_live = check_live_target_not_unsaved(
            target=cfg.target,
            close_saved=cfg.close_saved_live_target,
        )
        report["closed_live_workbooks"] = closed_live

        shutil.copy2(cfg.target, cfg.backup)
        shutil.copy2(cfg.target, cfg.sidecar)
        report["backup_sha256"] = sha256_file(cfg.backup)
        report["sidecar_pre_patch_sha256"] = sha256_file(cfg.sidecar)

        com_result = apply_patch_with_com(cfg.sidecar, cfg, logger)
        report["com_result"] = com_result

        sidecar_validation = validate(cfg.sidecar, baseline, com_result, cfg)
        report["sidecar_validation"] = sidecar_validation
        report["sidecar_post_patch_sha256"] = sha256_file(cfg.sidecar)

        if sidecar_validation["errors"]:
            report["status"] = "sidecar_validation_failed"
            logger.error("Sidecar validation failed with %d errors.", len(sidecar_validation["errors"]))
        elif not cfg.promote:
            report["status"] = "sidecar_validated_no_promote"
            exit_code = 0
            logger.info("Sidecar patch validated successfully. Promotion was skipped by flag.")
        else:
            promote_sidecar(cfg.sidecar, cfg.target)
            promoted = True
            report["promoted"] = True
            report["status"] = "promoted_pending_final_validation"

            final_validation = validate(cfg.target, baseline, com_result, cfg)
            report["final_validation"] = final_validation
            final_reopen = readonly_reopen(cfg.target, cfg)
            report["final_readonly_reopen"] = final_reopen
            report["final_sha256"] = sha256_file(cfg.target)

            if final_validation["errors"] or not final_reopen:
                logger.error("Final validation/reopen failed; restoring backup.")
                shutil.copy2(cfg.backup, cfg.target)
                report["status"] = "rolled_back_after_failed_final_validation"
                report["restored_backup_sha256"] = sha256_file(cfg.target)
                promoted = False
                report["promoted"] = False
            else:
                report["status"] = "success"
                exit_code = 0
                logger.info("Promotion succeeded and final validation passed.")

    except Exception as exc:
        logger.exception("Patch execution failed.")
        report["status"] = "exception"
        report["exception"] = repr(exc)
        report["traceback"] = traceback.format_exc()

        # If target was already replaced before the exception, restore backup.
        try:
            if promoted and cfg.backup.exists():
                shutil.copy2(cfg.backup, cfg.target)
                report["status"] = "rolled_back_after_exception"
                report["promoted"] = False
                report["restored_backup_sha256"] = sha256_file(cfg.target)
        except Exception as restore_exc:  # pragma: no cover - defensive
            report["restore_exception"] = repr(restore_exc)
            report["restore_traceback"] = traceback.format_exc()
    finally:
        report["duration_seconds"] = round(time.perf_counter() - started_at, 3)
        write_report(cfg.report, report)
        print(json.dumps(safe_json_value(report), indent=2, ensure_ascii=False))

    raise SystemExit(exit_code)


if __name__ == "__main__":
    main()
