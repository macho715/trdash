"""
patch_insert_apr25_openpyxl.py
Use openpyxl (no running Excel needed) to:
1. Insert 2 columns at K (col 11) → shifts Apr27+ right
2. Set K3=Apr25, L3=Apr26 with matching format
3. Fix subtitle formula (Row 2) to be dynamic
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, date

XLSM = r'C:\tr_dash-main\work\TR5_PreOp_Gantt_20260415\excel\TR5_PreOp_Gantt_20260415_162140.xlsm'
HEADER_ROW = 3
DATE_COL_START = 11   # K
NEW_FIRST = date(2026, 4, 25)

print("Loading workbook (keep_vba=True)...")
wb = load_workbook(XLSM, keep_vba=True)
ws = wb['Gantt_BASE']

# ── Inspect current K3 ────────────────────────────────────────────────────────
k3 = ws.cell(HEADER_ROW, DATE_COL_START)
print(f"K3 value={k3.value}  type={type(k3.value).__name__}  fmt={k3.number_format}")

# Read style reference from current K3 (which should be Apr27)
ref_cell = ws.cell(HEADER_ROW, DATE_COL_START)
ref_font = Font(
    bold=ref_cell.font.bold if ref_cell.font else True,
    size=ref_cell.font.size if ref_cell.font else 7,
    color=ref_cell.font.color.rgb if (ref_cell.font and ref_cell.font.color and ref_cell.font.color.type == 'rgb') else "FFFFFFFF"
)
ref_fill = PatternFill(
    fill_type=ref_cell.fill.fill_type if ref_cell.fill else "solid",
    fgColor=ref_cell.fill.fgColor.rgb if (ref_cell.fill and ref_cell.fill.fgColor and ref_cell.fill.fgColor.type == 'rgb') else "FF1F4E79"
)
ref_align = Alignment(horizontal='center', vertical='center')
ref_fmt = ref_cell.number_format or "MM-DD"

# ── Insert 2 columns at position 11 ──────────────────────────────────────────
print("Inserting 2 columns at K (col 11)...")
ws.insert_cols(DATE_COL_START, 2)
print("  Done.")

# ── Fill K3 = Apr 25, L3 = Apr 26 ───────────────────────────────────────────
for i in range(2):
    target = datetime(2026, 4, 25 + i)
    cell = ws.cell(HEADER_ROW, DATE_COL_START + i)
    cell.value = target
    cell.number_format = ref_fmt
    cell.font = Font(bold=True, size=ref_font.size, color=ref_font.color)
    cell.fill = ref_fill
    cell.alignment = ref_align
    print(f"  Set col {DATE_COL_START+i} ({get_column_letter(DATE_COL_START+i)}{HEADER_ROW}) = {target.date()}")

# ── Fix subtitle: Row 2, Col 1 ───────────────────────────────────────────────
# openpyxl stores formula strings starting with '='
subtitle_cell = ws.cell(2, 1)
print(f"\nCurrent subtitle: {subtitle_cell.value!r}")
subtitle_formula = (
    '=TEXT(TODAY(),"yyyy-mm-dd")'
    '&" | Source: TR5_Pre-Op_Simulation_v2_20260430.md | T+0=2026-04-30 | Load-out="'
    '&TEXT(D30,"yyyy-mm-dd")'
    '&" | AGI JD="'
    '&TEXT(D42,"yyyy-mm-dd")'
)
subtitle_cell.value = subtitle_formula
print(f"Subtitle formula set.")

# ── Verify ────────────────────────────────────────────────────────────────────
print("\nVerification (header row cols 11..16):")
for c in range(DATE_COL_START, DATE_COL_START + 6):
    cell = ws.cell(HEADER_ROW, c)
    print(f"  {get_column_letter(c)}{HEADER_ROW} = {cell.value!r}  fmt={cell.number_format}")

# ── Save ─────────────────────────────────────────────────────────────────────
print("\nSaving workbook...")
wb.save(XLSM)
print("Saved. Done.")
