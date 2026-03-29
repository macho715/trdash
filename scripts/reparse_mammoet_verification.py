from __future__ import annotations

import argparse
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
from pypdf import PdfReader


DPR_PDF_NAMES = [
    "15111578 - Samsung HVDC - DPRs - LOLI 1-3 TR Units.pdf",
    "15111578 - Samsung HVDC - DPRs - 24.02 - 26.02.pdf",
    "15111578 - Samsung HVDC - DPRs - 27.02 - 06.03.pdf",
]

CERT_PDF_NAMES = [
    "Certificate of Completion - AGI TR Units 1-3 Turning.pdf",
    "Certificate of Completion - AGI TR Units 1-4 Jack-down.pdf",
    "Certificate of Completion - AGI TR Units 1-4 Load-out.pdf",
    "Certificate of Completion - AGI TR Units 1-4 Transportation.pdf",
]

POSITION_RE = re.compile(
    r"(Project Manager|Project Engineer|SPMT SV(?:\s?\(MZP\))?|SPMT OP(?:\s?\(MZP\))?|"
    r"J&S SV|J&S OP|MARINE SV(?:\s?\(MZP\))?|MARINE OP(?:\s?\(MZP\))?)"
)


@dataclass
class Assignment:
    order: int
    name: str
    position: str
    start: str
    finish: str
    location: str | None
    equipment: str


@dataclass
class DPRPage:
    ref: str
    date: str
    notes: str
    mzp_activities: str
    agi_activities: str
    assignments: list[Assignment]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split())


def copy_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def find_header_row(ws, expected_headers: Iterable[str]) -> int:
    wanted = [normalize_text(x) for x in expected_headers]
    for row in range(1, ws.max_row + 1):
        values = [normalize_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if all(header in values for header in wanted):
            return row
    raise ValueError(f"Header row not found in sheet {ws.title}: {wanted}")


def header_map(ws, row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_text(ws.cell(row, col).value)
        if key:
            mapping[key] = col
    return mapping


def parse_assignments(text: str) -> list[Assignment]:
    match = re.search(r"START FINISH EXTENSION(.*?)S\.NO START TIME FINISH TIME", text, re.S)
    if not match:
        return []

    lines = [line.strip() for line in match.group(1).splitlines() if line.strip()]
    items: list[str] = []
    current = ""
    for line in lines:
        if re.match(r"^\d{1,2}(?:\s|$)", line):
            if current:
                items.append(current.strip())
            current = line
        else:
            current += f" {line}"
    if current:
        items.append(current.strip())

    parsed: list[Assignment] = []
    for item in items:
        item_match = re.match(r"^(\d{1,2})\s*(.*)$", item)
        if not item_match:
            continue

        order = int(item_match.group(1))
        body = item_match.group(2).strip().replace("(MZP ", "(MZP) ")
        time_match = re.search(r"(\d{2}:\d{2})\s+(\d{2}:\d{2})", body)
        if time_match:
            pre = body[: time_match.start()].strip()
            post = body[time_match.end() :].strip()
            start = time_match.group(1)
            finish = time_match.group(2)
        else:
            pre = body
            post = ""
            start = "—"
            finish = "—"

        pos_match = POSITION_RE.search(pre)
        if not pos_match:
            continue

        name = pre[: pos_match.start()].strip()
        if not name:
            continue

        position = pos_match.group(1)
        if not post:
            post = body[pos_match.end() :].strip()

        location = None
        equipment = post.strip()
        loc_match = re.search(r"\b(MZP|AGI)\b\s*$", equipment)
        if loc_match:
            location = loc_match.group(1)
            equipment = equipment[: loc_match.start()].strip()
        elif "(MZP)" in position:
            location = "MZP"

        parsed.append(
            Assignment(
                order=order,
                name=name,
                position=position,
                start=start,
                finish=finish,
                location=location,
                equipment=equipment,
            )
        )

    return parsed


def extract_section(text: str, start_pattern: str, end_pattern: str) -> str:
    match = re.search(start_pattern + r"(.*?)" + end_pattern, text, re.S)
    if not match:
        return ""
    return normalize_text(match.group(1))


def parse_dpr_page(text: str) -> DPRPage | None:
    ref_match = re.search(r"Doc Ref:\s*DPR\s*-\s*TR Units 1-7-LOLI\s*-\s*(\d+)", text)
    date_match = re.search(r"Project Date\s+(\d{2}/\d{2}/\d{4})", text)
    if not ref_match or not date_match:
        return None

    return DPRPage(
        ref=f"DPR-{ref_match.group(1)}",
        date=date_match.group(1),
        notes=extract_section(text, r"AREA OF CONCERN / NOTES\s*", r"\s*MZP Team"),
        mzp_activities=extract_section(text, r"MZP Team\s*", r"\s*ACTIVITIES REMARKS\s*AGI Team"),
        agi_activities=extract_section(text, r"AGI Team\s*", r"\s*S\.NO EMPLOYEE NAME POSITION"),
        assignments=parse_assignments(text),
    )


def load_dpr_pages(pdf_dir: Path) -> dict[str, DPRPage]:
    pages: dict[str, DPRPage] = {}
    for name in DPR_PDF_NAMES:
        reader = PdfReader(str(pdf_dir / name))
        for page in reader.pages:
            parsed = parse_dpr_page(page.extract_text() or "")
            if parsed:
                pages[parsed.ref] = parsed
    return pages


def build_existing_daily(ws) -> dict[str, dict[str, object]]:
    header_row = find_header_row(
        ws,
        [
            "Day#",
            "Date",
            "DPR Ref",
            "TR Unit",
            "Phase",
            "Work Phase Detail",
            "Location",
            "MZP Personnel",
            "AGI Personnel",
            "MZP Activities (DPR)",
            "AGI Activities (DPR)",
            "Key Events / LCT Status",
            "Waiting / Delay Recorded",
            "Δ vs Schedule",
        ],
    )
    cols = header_map(ws, header_row)
    data: dict[str, dict[str, object]] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        ref = ws.cell(row, cols["DPR Ref"]).value
        if not ref:
            continue
        ref_key = normalize_text(ref)
        data[ref_key] = {
            "row": row,
            "Day#": ws.cell(row, cols["Day#"]).value,
            "Date": ws.cell(row, cols["Date"]).value,
            "DPR Ref": ref,
            "TR Unit": ws.cell(row, cols["TR Unit"]).value,
            "Phase": ws.cell(row, cols["Phase"]).value,
            "Work Phase Detail": ws.cell(row, cols["Work Phase Detail"]).value,
            "Location": ws.cell(row, cols["Location"]).value,
            "MZP Personnel": ws.cell(row, cols["MZP Personnel"]).value,
            "AGI Personnel": ws.cell(row, cols["AGI Personnel"]).value,
            "MZP Activities (DPR)": ws.cell(row, cols["MZP Activities (DPR)"]).value,
            "AGI Activities (DPR)": ws.cell(row, cols["AGI Activities (DPR)"]).value,
            "Key Events / LCT Status": ws.cell(row, cols["Key Events / LCT Status"]).value,
            "Waiting / Delay Recorded": ws.cell(row, cols["Waiting / Delay Recorded"]).value,
            "Δ vs Schedule": ws.cell(row, cols["Δ vs Schedule"]).value,
        }
    return data


def build_existing_personnel(ws) -> tuple[dict[tuple[str, str, str, str], object], dict[str, list[dict[str, object]]]]:
    header_row = find_header_row(
        ws,
        ["Day#", "Date", "DPR Ref", "Employee Name", "Position", "Start", "Finish", "Location", "Notes"],
    )
    cols = header_map(ws, header_row)
    notes_by_key: dict[tuple[str, str, str, str], object] = {}
    rows_by_ref: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in range(header_row + 1, ws.max_row + 1):
        ref = normalize_text(ws.cell(row, cols["DPR Ref"]).value)
        if not ref:
            continue
        name = normalize_text(ws.cell(row, cols["Employee Name"]).value)
        start = normalize_text(ws.cell(row, cols["Start"]).value)
        finish = normalize_text(ws.cell(row, cols["Finish"]).value)
        notes_by_key[(ref, name, start, finish)] = ws.cell(row, cols["Notes"]).value
        rows_by_ref[ref].append(
            {
                "Day#": ws.cell(row, cols["Day#"]).value,
                "Date": ws.cell(row, cols["Date"]).value,
                "DPR Ref": ws.cell(row, cols["DPR Ref"]).value,
                "Employee Name": ws.cell(row, cols["Employee Name"]).value,
                "Position": ws.cell(row, cols["Position"]).value,
                "Start": ws.cell(row, cols["Start"]).value,
                "Finish": ws.cell(row, cols["Finish"]).value,
                "Location": ws.cell(row, cols["Location"]).value,
                "Notes": ws.cell(row, cols["Notes"]).value,
            }
        )
    return notes_by_key, rows_by_ref


def build_personnel_rows(
    dpr_pages: dict[str, DPRPage],
    existing_daily: dict[str, dict[str, object]],
    existing_person_rows_137: list[dict[str, object]],
    notes_by_key: dict[tuple[str, str, str, str], object],
) -> list[dict[str, object]]:
    refs_in_order = sorted(
        existing_daily.keys(),
        key=lambda ref: existing_daily[ref]["Day#"] if existing_daily[ref]["Day#"] is not None else 999,
    )
    output: list[dict[str, object]] = []
    for ref in refs_in_order:
        day = existing_daily[ref]["Day#"]
        date = existing_daily[ref]["Date"]
        if ref == "DPR-137":
            output.extend(existing_person_rows_137)
            continue
        page = dpr_pages.get(ref)
        if not page:
            continue
        for assignment in page.assignments:
            key = (ref, assignment.name, assignment.start, assignment.finish)
            output.append(
                {
                    "Day#": day,
                    "Date": page.date,
                    "DPR Ref": ref,
                    "Employee Name": assignment.name,
                    "Position": assignment.position,
                    "Start": assignment.start,
                    "Finish": assignment.finish,
                    "Location": assignment.location or "",
                    "Notes": notes_by_key.get(key),
                }
            )
    return output


def update_personnel_sheet(ws, rows: list[dict[str, object]]) -> None:
    header_row = find_header_row(
        ws,
        ["Day#", "Date", "DPR Ref", "Employee Name", "Position", "Start", "Finish", "Location", "Notes"],
    )
    cols = header_map(ws, header_row)
    template_row = header_row + 1
    max_row = max(ws.max_row, header_row + len(rows))
    for row in range(header_row + 1, max_row + 1):
        for col in cols.values():
            ws.cell(row, col).value = None
    for index, data in enumerate(rows, start=header_row + 1):
        for title, col in cols.items():
            copy_style(ws.cell(template_row, col), ws.cell(index, col))
            ws.cell(index, col).value = data.get(title)


def build_daily_rows(
    dpr_pages: dict[str, DPRPage],
    existing_daily: dict[str, dict[str, object]],
) -> list[dict[str, object]]:
    refs_in_order = sorted(
        existing_daily.keys(),
        key=lambda ref: existing_daily[ref]["Day#"] if existing_daily[ref]["Day#"] is not None else 999,
    )
    output: list[dict[str, object]] = []
    for ref in refs_in_order:
        existing = existing_daily[ref]
        if ref == "DPR-137" or ref not in dpr_pages:
            output.append(existing)
            continue

        page = dpr_pages[ref]
        mzp_count = sum(1 for assignment in page.assignments if assignment.location == "MZP")
        agi_count = sum(1 for assignment in page.assignments if assignment.location == "AGI")
        if mzp_count and agi_count:
            location = "MZP+AGI"
        elif mzp_count:
            location = "MZP"
        elif agi_count:
            location = "AGI"
        else:
            location = existing["Location"]

        output.append(
            {
                "Day#": existing["Day#"],
                "Date": page.date,
                "DPR Ref": ref,
                "TR Unit": existing["TR Unit"],
                "Phase": existing["Phase"],
                "Work Phase Detail": existing["Work Phase Detail"],
                "Location": location,
                "MZP Personnel": mzp_count,
                "AGI Personnel": agi_count,
                "MZP Activities (DPR)": page.mzp_activities or existing["MZP Activities (DPR)"],
                "AGI Activities (DPR)": page.agi_activities or existing["AGI Activities (DPR)"],
                "Key Events / LCT Status": existing["Key Events / LCT Status"] or page.notes,
                "Waiting / Delay Recorded": existing["Waiting / Delay Recorded"],
                "Δ vs Schedule": existing["Δ vs Schedule"],
            }
        )
    return output


def update_daily_sheet(ws, rows: list[dict[str, object]]) -> None:
    header_row = find_header_row(
        ws,
        [
            "Day#",
            "Date",
            "DPR Ref",
            "TR Unit",
            "Phase",
            "Work Phase Detail",
            "Location",
            "MZP Personnel",
            "AGI Personnel",
            "MZP Activities (DPR)",
            "AGI Activities (DPR)",
            "Key Events / LCT Status",
            "Waiting / Delay Recorded",
            "Δ vs Schedule",
        ],
    )
    cols = header_map(ws, header_row)
    template_row = header_row + 1
    max_row = max(ws.max_row, header_row + len(rows))
    for row in range(header_row + 1, max_row + 1):
        for col in cols.values():
            ws.cell(row, col).value = None
    for index, data in enumerate(rows, start=header_row + 1):
        for title, col in cols.items():
            copy_style(ws.cell(template_row, col), ws.cell(index, col))
            ws.cell(index, col).value = data.get(title)


def update_summary_sheet(ws) -> None:
    ws["C5"] = "18/03/2026"
    ws["G5"] = "26/01/2026 - 06/03/2026"

    milestone_header_row = find_header_row(ws, ["TR Unit", "Milestone", "Planned (Original)", "Actual (DPR)", "Δ Days", "Delay Reason"])
    cols = header_map(ws, milestone_header_row)
    milestone_rows = {
        25: {
            "TR Unit": "TR4",
            "Milestone": "Load-Out MZP",
            "Planned (Original)": "22/02/2026",
            "Actual (DPR)": "25/02/2026 @08:05",
            "Δ Days": "+3d",
            "Delay Reason": "Berth #5 occupied until 23.02; TR4 LO shifted to 25.02",
        },
        26: {
            "TR Unit": "TR4",
            "Milestone": "ETD MZP",
            "Planned (Original)": "24/02/2026",
            "Actual (DPR)": "27/02/2026 @08:42",
            "Δ Days": "+3d",
            "Delay Reason": "Bad weather hold after load-out; cast off 27.02 @08:42",
        },
        27: {
            "TR Unit": "TR4",
            "Milestone": "Load-In AGI",
            "Planned (Original)": "28/02/2026",
            "Actual (DPR)": "02/03/2026 @16:00",
            "Δ Days": "+2d",
            "Delay Reason": "Weather hold + safety evacuation / Force Majeure",
        },
        28: {
            "TR Unit": "TR4",
            "Milestone": "Jacking Down",
            "Planned (Original)": "05/03/2026",
            "Actual (DPR)": "05/03/2026 (full support)",
            "Δ Days": "0d",
            "Delay Reason": "Transport completed; ready for turning/final jack-down",
        },
    }
    for row, values in milestone_rows.items():
        for title, value in values.items():
            ws.cell(row, cols[title]).value = value

    cause_rows = {
        31: {"#": 1, "Root Cause": "FRA (Fire & Rescue Authority) Approval Delay", "Period / Time": "03.02-08.02 (6 days)", "Impact on Operations": "TR1 LI delayed 3d; TR1 Transport permit delayed", "Category": "External - Regulatory"},
        32: {"#": 2, "Root Cause": "Weather - FOG (11/02/2026)", "Period / Time": "11.02 07:00-09:00", "Impact on Operations": "TR2 LO cancelled; crew released; rescheduled to 12.02", "Category": "External - Weather"},
        33: {"#": 3, "Root Cause": "Weather - STRONG WIND (16/02/2026)", "Period / Time": "16.02 all day", "Impact on Operations": "All ops suspended at MZP; TR3 LO delayed", "Category": "External - Weather"},
        34: {"#": 4, "Root Cause": "Vessel Condition - High Tide (18/02/2026)", "Period / Time": "18.02 ~12:00-13:30", "Impact on Operations": "Load-out delayed 1.5h; required ballasting for lower freeboard", "Category": "External - Tidal/Vessel"},
        35: {"#": 5, "Root Cause": "MZP Berth #5 Occupied (19-23/02/2026)", "Period / Time": "19.02-23.02", "Impact on Operations": "LCT at anchorage 4 days; TR4 LO delayed; crew on standby", "Category": "External - Port Conflict"},
        36: {"#": 6, "Root Cause": "Gate Pass Renewal Delays (Ramadan)", "Period / Time": "23.02", "Impact on Operations": "Gate passes expired; renewal took longer during Ramadan", "Category": "External - Admin"},
        37: {"#": 7, "Root Cause": "Equipment Wait - Forklift (26/01/2026)", "Period / Time": "26.01", "Impact on Operations": "Marine equipment offloading delayed", "Category": "Operational"},
        38: {"#": 8, "Root Cause": "Force Majeure / Evacuation + Strong Wind", "Period / Time": "28.02-05.03", "Impact on Operations": "TR4 AGI load-in delayed to 02.03; transport/jack-down sequence held until 05.03", "Category": "External - Security / Weather"},
    }
    cause_header_row = find_header_row(ws, ["#", "Root Cause", "Period / Time", "Impact on Operations", "Category"])
    cause_cols = header_map(ws, cause_header_row)
    for row, values in cause_rows.items():
        for title, value in values.items():
            ws.cell(row, cause_cols[title]).value = value


def parse_certificate_rows(text: str) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    for unit, sn, actual in re.findall(r"Unit\s+(\d)\s+S/N\s+([A-Z0-9\-]+)\s+[–-]\s+([^\n]+)", text):
        rows.append((unit, sn, normalize_text(actual).rstrip(".")))
    return rows


def parse_certificate_date(text: str) -> str:
    match = re.search(r"Date:\s*([0-9]{2}[-/][0-9]{2}[-/][0-9]{4})", text)
    if not match:
        return ""
    raw = match.group(1).replace("/", "-")
    day, month, year = raw.split("-")
    return f"{year}-{month}-{day}"


def normalize_dotted_date(text: str) -> str:
    match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", text)
    if not match:
        return normalize_text(text)
    return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"


def update_certificate_sheet(ws, pdf_dir: Path) -> None:
    title_row = find_header_row(
        ws,
        ["No.", "완료 범위 (Scope)", "대상 TR Units", "TR Unit / S/N", "완료일 (Actual)", "발행일 (Cert. Date)", "서명자 (Mammoet)"],
    )
    cols = header_map(ws, title_row)

    cert_texts = {name: PdfReader(str(pdf_dir / name)).pages[0].extract_text() or "" for name in CERT_PDF_NAMES}
    turning_rows = parse_certificate_rows(cert_texts[CERT_PDF_NAMES[0]])
    jack_rows = parse_certificate_rows(cert_texts[CERT_PDF_NAMES[1]])
    load_rows = parse_certificate_rows(cert_texts[CERT_PDF_NAMES[2]])
    transport_rows = parse_certificate_rows(cert_texts[CERT_PDF_NAMES[3]])
    issue_dates = {name: parse_certificate_date(text) for name, text in cert_texts.items()}

    def fill_section(start_row: int, scope: str, rows: list[tuple[str, str, str]], cert_name: str) -> None:
        for offset, (unit, sn, actual) in enumerate(rows, start=1):
            row = start_row + offset
            ws.cell(row, cols["No."]).value = offset
            ws.cell(row, cols["완료 범위 (Scope)"]).value = scope
            ws.cell(row, cols["대상 TR Units"]).value = f"TR Unit {unit}"
            ws.cell(row, cols["TR Unit / S/N"]).value = f"S/N {sn}"
            ws.cell(row, cols["완료일 (Actual)"]).value = actual
            ws.cell(row, cols["발행일 (Cert. Date)"]).value = issue_dates[cert_name]
            ws.cell(row, cols["서명자 (Mammoet)"]).value = "Yulia Frolova"

    normalized_turning_rows = []
    for unit, sn, actual in turning_rows:
        if "require turning" in actual.lower():
            normalized_turning_rows.append((unit, sn, "Turning 불필요"))
        else:
            normalized_turning_rows.append((unit, sn, normalize_dotted_date(actual)))
    fill_section(5, "Turning", normalized_turning_rows, CERT_PDF_NAMES[0])

    normalized_jack_rows = [(unit, sn, normalize_dotted_date(actual)) for unit, sn, actual in jack_rows]
    fill_section(10, "Jack-down", normalized_jack_rows, CERT_PDF_NAMES[1])

    normalized_load_rows = []
    for unit, sn, actual in load_rows:
        if unit == "1" and "31.02.2026" in actual:
            normalized_load_rows.append((unit, sn, "2026-01-31*"))
        else:
            normalized_load_rows.append((unit, sn, normalize_dotted_date(actual)))
    fill_section(15, "Load-out (LCT)", normalized_load_rows, CERT_PDF_NAMES[2])

    normalized_transport_rows = []
    for unit, sn, actual in transport_rows:
        normalized_transport_rows.append((unit, sn, normalize_dotted_date(actual)))
    fill_section(21, "SPMT Transport", normalized_transport_rows, CERT_PDF_NAMES[3])

    # Summary block.
    ws["A29"] = "TR Unit 1"
    ws["B29"] = "2XBR62542-T"
    ws["C29"] = "2026-01-31"
    ws["D29"] = "2026-02-11"
    ws["E29"] = "불필요"
    ws["F29"] = "2026-02-12"
    ws["G29"] = "완료"

    ws["A30"] = "TR Unit 2"
    ws["B30"] = "2XBR62536-T"
    ws["C30"] = "2026-02-12"
    ws["D30"] = "2026-02-13"
    ws["E30"] = "2026-02-17"
    ws["F30"] = "2026-02-18"
    ws["G30"] = "완료"

    ws["A31"] = "TR Unit 3"
    ws["B31"] = "2XBR62537-T"
    ws["C31"] = "2026-02-18"
    ws["D31"] = "2026-02-20"
    ws["E31"] = "2026-02-22"
    ws["F31"] = "2026-02-23"
    ws["G31"] = "완료"

    ws["A32"] = "TR Unit 4"
    ws["B32"] = "2XBR62538-T"
    ws["C32"] = "2026-02-25"
    ws["D32"] = "2026-03-05"
    ws["E32"] = "—"
    ws["F32"] = "—"
    ws["G32"] = "FM delayed"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--pdf-dir", required=True)
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    pdf_dir = Path(args.pdf_dir)

    dpr_pages = load_dpr_pages(pdf_dir)
    wb = openpyxl.load_workbook(workbook_path)

    summary_ws = wb["📊 Summary"]
    cert_ws = wb["📜 완료증명서"]
    daily_ws = wb["📋 DPR Daily Log"]
    personnel_ws = wb["👷 Personnel Register"]

    existing_daily = build_existing_daily(daily_ws)
    personnel_notes, personnel_rows_by_ref = build_existing_personnel(personnel_ws)

    update_summary_sheet(summary_ws)
    update_certificate_sheet(cert_ws, pdf_dir)

    daily_rows = build_daily_rows(dpr_pages, existing_daily)
    update_daily_sheet(daily_ws, daily_rows)

    personnel_rows = build_personnel_rows(
        dpr_pages=dpr_pages,
        existing_daily=existing_daily,
        existing_person_rows_137=personnel_rows_by_ref.get("DPR-137", []),
        notes_by_key=personnel_notes,
    )
    update_personnel_sheet(personnel_ws, personnel_rows)

    wb.save(workbook_path)


if __name__ == "__main__":
    main()
