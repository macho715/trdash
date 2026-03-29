#!/usr/bin/env python3
"""Apply consistent, presentation-grade formatting to the Bushra review workbook."""

from __future__ import annotations

import math
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_FILE_NAME = "BUSHRA_DECKLOG_compiled_20260310_timeline_compare_reviewed.essential_only.xlsx"

NAVY = "1F4E78"
NAVY_DARK = "173A5E"
SKY = "D9EAF7"
SKY_LIGHT = "F6FAFD"
SLATE = "E9EFF5"
GRID = "D6DEE8"
TEXT = "1F2937"
TEXT_MUTED = "475569"
WHITE = "FFFFFF"
SUCCESS_FILL = "E2F0D9"
SUCCESS_TEXT = "2F6B2F"
WARN_FILL = "FFF2CC"
WARN_TEXT = "8A6D1D"
DANGER_FILL = "FCE4D6"
DANGER_TEXT = "8A3B12"
INFO_FILL = "DDEBF7"
INFO_TEXT = "1F4E78"
NEUTRAL_FILL = "EDEDED"
NEUTRAL_TEXT = "555555"

THIN_SIDE = Side(style="thin", color=GRID)
MEDIUM_SIDE = Side(style="medium", color=NAVY)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=MEDIUM_SIDE, bottom=MEDIUM_SIDE)

BODY_FONT = Font(name="Aptos", size=10, color=TEXT)
BODY_BOLD = Font(name="Aptos", size=10, bold=True, color=TEXT)
HEADER_FONT = Font(name="Aptos", size=10, bold=True, color=WHITE)
TITLE_FONT = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
SECTION_FONT = Font(name="Aptos", size=11, bold=True, color=NAVY_DARK)
LABEL_FONT = Font(name="Aptos", size=10, bold=True, color=TEXT_MUTED)
NOTE_FONT = Font(name="Aptos", size=9, italic=True, color=TEXT_MUTED)

TITLE_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
LABEL_FILL = PatternFill(fill_type="solid", fgColor=SLATE)
ALT_FILL = PatternFill(fill_type="solid", fgColor=SKY_LIGHT)
BLANK_FILL = PatternFill(fill_type="solid", fgColor=WHITE)


def resolve_workbook_path() -> Path:
    if len(sys.argv) > 1:
        path = Path(sys.argv[1]).expanduser().resolve()
        if path.exists():
            return path
        raise FileNotFoundError(f"Workbook not found: {path}")

    root = Path(__file__).resolve().parents[1]
    matches = list(root.rglob(DEFAULT_FILE_NAME))
    if not matches:
        raise FileNotFoundError(f"Workbook not found: {DEFAULT_FILE_NAME}")
    return matches[0]


def backup_workbook(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def iter_used_cells(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            yield cell


def apply_sheet_defaults(ws, tab_color: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = tab_color
    ws.print_options.horizontalCentered = False
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.55
    ws.page_margins.bottom = 0.55
    for cell in iter_used_cells(ws):
        cell.font = copy(BODY_FONT)
        cell.alignment = Alignment(vertical="top")
        cell.border = copy(THIN_BORDER)


def style_title_row(ws, end_col: int) -> None:
    if end_col < 1:
        return
    end_letter = get_column_letter(end_col)
    if f"A1:{end_letter}1" not in [str(rng) for rng in ws.merged_cells.ranges]:
        ws.merge_cells(f"A1:{end_letter}1")
    cell = ws["A1"]
    cell.fill = copy(TITLE_FILL)
    cell.font = copy(TITLE_FONT)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def style_header_row(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(HEADER_BORDER)
    ws.row_dimensions[row_idx].height = 22


def stripe_rows(ws, start_row: int, end_row: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else BLANK_FILL
        for cell in ws[row_idx]:
            if cell.value not in (None, ""):
                cell.fill = copy(fill)


def set_alignment(ws, cols: str, start_row: int, end_row: int, horizontal: str = "left", wrap: bool = False) -> None:
    for col in cols.split(","):
        col = col.strip()
        if not col:
            continue
        for row_idx in range(start_row, end_row + 1):
            ws[f"{col}{row_idx}"].alignment = Alignment(
                horizontal=horizontal,
                vertical="top",
                wrap_text=wrap,
            )


def set_column_widths(ws, widths: dict[str, float], default_width: float | None = None) -> None:
    if default_width is not None:
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = default_width
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def estimate_height(value, width: float, min_height: float = 20.0, line_height: float = 15.0) -> float:
    if value in (None, ""):
        return min_height
    text = str(value)
    lines = text.splitlines() or [text]
    wrapped_lines = 0
    chars_per_line = max(int(width * 1.15), 8)
    for line in lines:
        wrapped_lines += max(1, math.ceil(len(line) / chars_per_line))
    return max(min_height, wrapped_lines * line_height + 4)


def set_estimated_row_heights(ws, start_row: int, text_columns: dict[str, float], min_height: float = 20.0, max_height: float = 420.0) -> None:
    for row_idx in range(start_row, ws.max_row + 1):
        height = min_height
        for col, width in text_columns.items():
            value = ws[f"{col}{row_idx}"].value
            height = max(height, estimate_height(value, width, min_height=min_height))
        ws.row_dimensions[row_idx].height = min(max_height, height)


def apply_autofilter(ws, start_row: int = 1) -> None:
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def fill_cell(cell, fill_color: str, font_color: str, bold: bool = True) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
    cell.font = Font(name="Aptos", size=10, bold=bold, color=font_color)


def style_metadata_rows(ws, rows: range, value_col: str = "B") -> None:
    for row_idx in rows:
        label = ws[f"A{row_idx}"]
        value = ws[f"{value_col}{row_idx}"]
        if label.value not in (None, ""):
            label.fill = copy(LABEL_FILL)
            label.font = copy(LABEL_FONT)
            label.alignment = Alignment(horizontal="left", vertical="center")
        if value.value not in (None, ""):
            value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def format_final_report(ws) -> None:
    apply_sheet_defaults(ws, NAVY)
    style_title_row(ws, 6)
    style_metadata_rows(ws, range(2, 7))
    ws["B6"].font = copy(BODY_BOLD)
    ws["B6"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    style_header_row(ws, 8)
    stripe_rows(ws, 9, ws.max_row)
    apply_autofilter(ws, 8)
    ws.freeze_panes = "A8"
    set_column_widths(
        ws,
        {
            "A": 14,
            "B": 12,
            "C": 14,
            "D": 16,
            "E": 16,
            "F": 54,
        },
    )
    set_alignment(ws, "A,B,C,D,E", 8, ws.max_row, horizontal="center", wrap=True)
    set_alignment(ws, "F", 8, ws.max_row, horizontal="left", wrap=True)
    set_estimated_row_heights(ws, 2, {"B": 54, "F": 54}, min_height=20, max_height=90)


def format_verify_summary(ws) -> None:
    apply_sheet_defaults(ws, "5B9BD5")
    style_title_row(ws, 7)
    style_metadata_rows(ws, range(2, 7))
    for address in ["D2", "G2"]:
        ws[address].fill = copy(LABEL_FILL)
        ws[address].font = copy(LABEL_FONT)
        ws[address].alignment = Alignment(horizontal="left", vertical="center")
    for row_idx in range(3, 9):
        for address in [f"D{row_idx}", f"E{row_idx}", f"G{row_idx}"]:
            ws[address].border = copy(THIN_BORDER)
    stripe_rows(ws, 3, 8)
    set_column_widths(
        ws,
        {
            "A": 16,
            "B": 46,
            "C": 4,
            "D": 12,
            "E": 10,
            "F": 4,
            "G": 68,
        },
    )
    set_alignment(ws, "A,B,D,G", 2, ws.max_row, horizontal="left", wrap=True)
    set_alignment(ws, "E", 2, ws.max_row, horizontal="center", wrap=False)
    ws.freeze_panes = "A3"
    set_estimated_row_heights(ws, 2, {"B": 46, "G": 68}, min_height=20, max_height=96)


def format_verify_detail(ws) -> None:
    apply_sheet_defaults(ws, "70AD47")
    style_header_row(ws, 1)
    stripe_rows(ws, 2, ws.max_row)
    apply_autofilter(ws, 1)
    ws.freeze_panes = "A2"
    set_column_widths(
        ws,
        {
            "A": 12,
            "B": 6,
            "C": 12,
            "D": 12,
            "E": 42,
            "F": 18,
            "G": 34,
            "H": 22,
            "I": 24,
            "J": 46,
            "K": 28,
            "L": 22,
            "M": 42,
        },
    )
    set_alignment(ws, "A,B,C,D,F,H,I,L", 1, ws.max_row, horizontal="center", wrap=True)
    set_alignment(ws, "E,G,J,K,M", 1, ws.max_row, horizontal="left", wrap=True)
    for row_idx in range(2, ws.max_row + 1):
        status_cell = ws[f"F{row_idx}"]
        coverage_cell = ws[f"L{row_idx}"]
        status = str(status_cell.value or "").upper()
        coverage = str(coverage_cell.value or "").upper()
        if "VERIFIED_EXACT" in status:
            fill_cell(status_cell, SUCCESS_FILL, SUCCESS_TEXT)
        elif "PARTIAL" in status:
            fill_cell(status_cell, WARN_FILL, WARN_TEXT)
        elif "CONTEXT_ONLY" in status:
            fill_cell(status_cell, DANGER_FILL, DANGER_TEXT)
        elif "REVIEW" in status:
            fill_cell(status_cell, DANGER_FILL, DANGER_TEXT)
        if "WITHIN" in coverage:
            fill_cell(coverage_cell, INFO_FILL, INFO_TEXT, bold=False)
        elif coverage:
            fill_cell(coverage_cell, NEUTRAL_FILL, NEUTRAL_TEXT, bold=False)
    set_estimated_row_heights(ws, 2, {"E": 42, "G": 34, "J": 46, "M": 42}, min_height=22, max_height=110)


def format_daily_decklog(ws) -> None:
    apply_sheet_defaults(ws, "ED7D31")
    style_header_row(ws, 1)
    stripe_rows(ws, 2, ws.max_row)
    apply_autofilter(ws, 1)
    ws.freeze_panes = "A2"
    set_column_widths(
        ws,
        {
            "A": 6,
            "B": 12,
            "C": 12,
            "D": 14,
            "E": 26,
            "F": 16,
            "G": 10,
            "H": 16,
            "I": 86,
            "J": 10,
            "K": 10,
            "L": 10,
            "M": 10,
            "N": 10,
            "O": 10,
            "P": 10,
            "Q": 10,
            "R": 14,
            "S": 14,
            "T": 14,
            "U": 14,
            "V": 14,
            "W": 14,
            "X": 14,
            "Y": 14,
            "Z": 16,
            "AA": 16,
            "AB": 16,
            "AC": 16,
            "AD": 16,
            "AE": 16,
            "AF": 26,
            "AG": 14,
        },
    )
    set_alignment(ws, "A,B,C,D,F,G,H,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AG", 1, ws.max_row, horizontal="center", wrap=True)
    set_alignment(ws, "E,I,AF", 1, ws.max_row, horizontal="left", wrap=True)
    for row_idx in range(2, ws.max_row + 1):
        port_cell = ws[f"F{row_idx}"]
        activity_cell = ws[f"I{row_idx}"]
        port_value = str(port_cell.value or "").upper()
        activity_value = str(activity_cell.value or "").upper()
        if "SOURCE GAP" in port_value or "SOURCE GAP" in activity_value:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type="solid", fgColor="F4F4F4")
    set_estimated_row_heights(ws, 2, {"E": 26, "I": 86, "AF": 26}, min_height=22, max_height=410)


def format_whatsapp(ws) -> None:
    apply_sheet_defaults(ws, "4472C4")
    style_header_row(ws, 1)
    stripe_rows(ws, 2, ws.max_row)
    apply_autofilter(ws, 1)
    ws.freeze_panes = "A2"
    set_column_widths(
        ws,
        {
            "A": 18,
            "B": 12,
            "C": 22,
            "D": 92,
            "E": 16,
            "F": 34,
        },
    )
    set_alignment(ws, "A,B,E", 1, ws.max_row, horizontal="center", wrap=True)
    set_alignment(ws, "C,D,F", 1, ws.max_row, horizontal="left", wrap=True)
    for row_idx in range(2, ws.max_row + 1):
        relevance_cell = ws[f"E{row_idx}"]
        relevance = str(relevance_cell.value or "").strip().lower()
        if relevance == "yes":
            fill_cell(relevance_cell, SUCCESS_FILL, SUCCESS_TEXT)
        elif relevance == "no":
            fill_cell(relevance_cell, DANGER_FILL, DANGER_TEXT)
    set_estimated_row_heights(ws, 2, {"C": 22, "D": 92, "F": 34}, min_height=22, max_height=300)


def format_bushra_text(ws) -> None:
    apply_sheet_defaults(ws, "A5A5A5")
    style_header_row(ws, 1)
    stripe_rows(ws, 2, ws.max_row)
    apply_autofilter(ws, 1)
    ws.freeze_panes = "A2"
    set_column_widths(
        ws,
        {
            "A": 14,
            "B": 20,
            "C": 18,
            "D": 96,
            "E": 30,
            "F": 10,
        },
    )
    set_alignment(ws, "A,B,F", 1, ws.max_row, horizontal="center", wrap=True)
    set_alignment(ws, "C,D,E", 1, ws.max_row, horizontal="left", wrap=True)
    set_estimated_row_heights(ws, 2, {"C": 18, "D": 96, "E": 30}, min_height=22, max_height=160)


def main() -> int:
    workbook_path = resolve_workbook_path()
    backup_path = backup_workbook(workbook_path)
    wb = load_workbook(workbook_path)

    formatters = {
        "01_Final_Report": format_final_report,
        "02_Verify_Summary": format_verify_summary,
        "03_Verify_Detail": format_verify_detail,
        "04_Daily_DeckLog": format_daily_decklog,
        "WhatsApp_Parsed": format_whatsapp,
        "Bushra_Text_Parsed": format_bushra_text,
    }

    for name, formatter in formatters.items():
        if name in wb.sheetnames:
            formatter(wb[name])

    wb.save(workbook_path)
    print(f"Formatted workbook: {workbook_path}")
    print(f"Backup created: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
